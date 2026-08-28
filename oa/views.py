from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework import serializers
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.db import models, transaction
from django.conf import settings
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from loguru import logger
import os
import json
import uuid
import requests

from .models import (
    AttendanceRecord, ApprovalRequest, ApprovalType, ApprovalLog,
    ApprovalNode, ApprovalAssignee, WorkNotification, ApprovalCarbonCopy,
    ApprovalDeptConfig, AttendanceConfig, UserAttendanceConfig,
    SubsidyApplication, SubsidyPayment, SubsidyConfig, SubsidyWallet,
    SubsidyWithdrawal, SubsidyInvoiceVerifyRecord,
    MaterialItem, MaterialRequirement, MaterialRequisition,
    MaterialRequirementItem, MaterialRequisitionItem, DocumentSequence,
    WatermarkConfig, DEFAULT_WATERMARK_PAGES, PrintLog, DailyWorkSummary,
)
from .type_utils import (
    ensure_builtin_types, resolve_approval_type,
    collect_form_data, validate_form_data,
)


def _approval_type_label(approval):
    """审批类型展示名（动态类型表解析，兼容历史数据）"""
    t = resolve_approval_type(approval.approval_type, approval.tenant)
    return t.name if t else approval.approval_type


def _threshold_field_label(config):
    """阈值字段展示名：内置字段走映射，自定义类型回退到 schema 字段 label"""
    if not config or not config.threshold_field:
        return ''
    field_map = dict(ApprovalDeptConfig.THRESHOLD_FIELD_CHOICES)
    if config.threshold_field in field_map:
        return field_map[config.threshold_field]
    t = resolve_approval_type(config.approval_type, config.tenant)
    if t:
        for f in (t.form_schema or []):
            if f.get('key') == config.threshold_field:
                return f.get('label') or config.threshold_field
    return config.threshold_field
from .serializers import (
    AttendanceRecordSerializer,
    AttendanceClockSerializer,
    ApprovalRequestSerializer,
    ApprovalListSerializer,
    ApprovalCreateSerializer,
    ApprovalActionSerializer,
    ApprovalDraftSerializer,
    ApprovalLogSerializer,
)
from utils.encrypt_aes import encrypt_data
from utils.request_util import get_request_ip
from utils.cloud_permission_utils import get_effective_operation_permission


def send_work_notification(user_id, title, content, notification_type='system', related_url='', extra_data=None):
    """创建工作通知并通过WebSocket实时推送"""
    from accounts.models import CustomUser
    try:
        user = CustomUser.objects.get(id=user_id)
        note = WorkNotification.objects.create(
            recipient=user,
            notification_type=notification_type,
            title=title,
            content=content,
            related_url=related_url,
            extra_data=extra_data or {},
        )
        # WebSocket实时推送
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f'user_{user_id}_notifications',
            {
                'type': 'work.notification',
                'event_type': 'new',
                'notification': {
                    'id': note.id,
                    'type': notification_type,
                    'title': title,
                    'content': content,
                    'related_url': related_url,
                    'avatar_url': user.get_avatar_url() if hasattr(user, 'get_avatar_url') else '',
                    'created_at': note.created_at.isoformat() if note.created_at else '',
                }
            }
        )
        # Web Push：无论在线与否都推送，保证回到主屏幕 / 锁屏 / 关闭 App 时也能收到工作通知。
        # urgent=True → urgency=high：即时送达，避免 iOS 批量延迟。
        try:
            from django.conf import settings as dj_settings
            if dj_settings.PUSH_ENABLED:
                from chat.push_utils import build_work_push_payload
                from chat.tasks import send_push_task
                send_push_task.delay(user_id, build_work_push_payload(note, related_url), ttl=43200, urgent=True)
        except Exception as e:
            logger.warning(f'Web Push 工作通知失败: {e}')
        return note
    except CustomUser.DoesNotExist:
        logger.error(f'用户{user_id}不存在')
        return None
    except Exception as e:
        logger.error(f'创建工作通知失败: {e}')
        return None


class AttendanceViewSet(viewsets.ViewSet):
    """考勤打卡视图集"""
    permission_classes = [permissions.IsAuthenticated]

    def _get_managed_department_ids(self, user):
        """获取用户作为负责人/副负责人的部门 ID 集合（含所有子部门）"""
        from accounts.models import Department
        managed = list(Department.objects.filter(
            Q(manager=user) | Q(deputy_managers=user)
        ).values_list('id', flat=True))
        if not managed:
            return []
        # 向下收集所有子部门，使部门负责人可查看整个部门树成员的打卡记录
        ids = list(managed)
        frontier = list(managed)
        while frontier:
            children = list(Department.objects.filter(
                parent_id__in=frontier
            ).values_list('id', flat=True))
            new_ids = [c for c in children if c not in ids]
            if not new_ids:
                break
            ids.extend(new_ids)
            frontier = new_ids
        return ids

    def _can_view_record(self, user, record):
        """打卡记录查看权限：超管=全部；部门负责人=本部门成员；普通用户=仅自己"""
        if user.user_type == 'super_admin':
            return True
        if record.user_id == user.id:
            return True
        managed = self._get_managed_department_ids(user)
        if managed:
            from org.models import UserDepartment
            return UserDepartment.objects.filter(
                user_id=record.user_id, department_id__in=managed
            ).exists()
        return False

    def list(self, request):
        """打卡记录列表（分页+搜索，三级权限：超管/部门负责人/普通用户）"""
        user = request.user
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        search = request.query_params.get('search', '').strip()
        status_filter = request.query_params.get('status', '').strip()
        clock_type = request.query_params.get('clock_type', '').strip()

        if user.user_type == 'super_admin':
            # 超级管理员：可见全部成员（当前企业 + 子企业）
            qs = AttendanceRecord.objects.select_related('user__department').all()
            if tenant:
                tenant_ids = [tenant.id]
                try:
                    sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                    if sub_ids:
                        tenant_ids.extend(sub_ids)
                except Exception:
                    pass
                qs = qs.filter(tenant_id__in=tenant_ids)
        else:
            # 部门负责人：仅可见所管理部门的成员（含子部门）
            managed_dept_ids = self._get_managed_department_ids(user)
            if managed_dept_ids:
                qs = AttendanceRecord.objects.select_related('user__department').filter(
                    user__department_relations__department_id__in=managed_dept_ids
                ).distinct()
            else:
                # 普通用户：仅可见自己
                qs = AttendanceRecord.objects.select_related('user__department').filter(user=user)

        filter_tenant_id = request.query_params.get('tenant_id', '').strip()
        if filter_tenant_id:
            qs = qs.filter(tenant_id=filter_tenant_id)
        filter_org_dept_id = request.query_params.get('org_dept_id', '').strip()
        if filter_org_dept_id:
            qs = qs.filter(user__department_relations__department_id=filter_org_dept_id)

        if status_filter:
            qs = qs.filter(status=status_filter)

        if clock_type:
            qs = qs.filter(clock_type=clock_type)

        if search:
            qs = qs.filter(
                Q(user__username__icontains=search) |
                Q(user__real_name__icontains=search) |
                Q(location__icontains=search) |
                Q(date__icontains=search)
            )

        qs = qs.order_by('-clock_time')
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        start = (page - 1) * page_size
        items = qs[start:start + page_size]

        results = AttendanceRecordSerializer(items, many=True, context={'request': request}).data

        return Response({'encrypt': True, 'data': encrypt_data({
            'results': results,
            'count': total,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
        })})

    def retrieve(self, request, pk=None):
        try:
            record = AttendanceRecord.objects.select_related('user__department').get(id=pk)
            if not self._can_view_record(request.user, record):
                return Response({'error': '暂无查看权限'}, status=403)
            data = AttendanceRecordSerializer(record, context={'request': request}).data
            return Response({'encrypt': True, 'data': encrypt_data(data)})
        except AttendanceRecord.DoesNotExist:
            return Response({'error': '记录不存在'}, status=404)

    @action(detail=False, methods=['post'])
    def clock_in(self, request):
        serializer = AttendanceClockSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        now = timezone.localtime(timezone.now())
        today = now.date()
        # 检查考勤配置（含用户主部门）
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        from org.models import UserDepartment
        primary_dept = UserDepartment.objects.filter(user=request.user, is_primary=True).select_related('department').first()
        dept_id = primary_dept.department_id if primary_dept else None
        config = self._get_attendance_config(tenant, dept_id, request.user)
        if config and not config.clock_in_enabled:
            return Response({'error': '该时段无需打卡（已配置为不打卡）', 'skip': True}, status=200)

        existing = AttendanceRecord.objects.filter(user=request.user, date=today, clock_type='clock_in').first()
        if existing:
            data = AttendanceRecordSerializer(existing, context={'request': request}).data
            return Response({'encrypt': True, 'data': encrypt_data(data)})
        # 使用配置时间或默认 9:00
        if config and config.clock_in_time:
            deadline = now.replace(hour=config.clock_in_time.hour, minute=config.clock_in_time.minute, second=59, microsecond=0)
        else:
            deadline = now.replace(hour=9, minute=0, second=59, microsecond=0)
        status_val = 'late' if now > deadline else 'normal'
        logger.info(f'{request.user} 上班打卡 now: {now} deadline: {deadline} status: {status_val}')
        record = AttendanceRecord.objects.create(
            user=request.user, clock_type='clock_in', date=today,
            tenant=getattr(request, 'tenant', None) or request.user.get_active_tenant(),
            latitude=serializer.validated_data.get('latitude'),
            longitude=serializer.validated_data.get('longitude'),
            location=serializer.validated_data.get('location', ''),
            device=serializer.validated_data.get('device', ''),
            status=status_val, remark=serializer.validated_data.get('remark', ''),
            reverse_geocoding=serializer.validated_data.get('reverse_geocoding'),
            ip_address=serializer.validated_data.get('ip_address', '') or get_request_ip(request),
            user_agent=serializer.validated_data.get('user_agent', '') or request.META.get('HTTP_USER_AGENT', ''),
        )
        location_text = serializer.validated_data.get('location', '')
        logger.info(f'{request.user} 上班打卡 {today} status={status_val}')
        # 通知用户本人
        status_text = '迟到' if status_val == 'late' else '正常'
        loc = f' 位置：{location_text}' if location_text else ''
        send_work_notification(
            user_id=request.user.id,
            title='上班打卡成功',
            content=f'今日上班打卡 {status_text}{loc}',
            notification_type='attendance',
            related_url='/oa/attendance/',
            extra_data={'status': status_val, 'date': str(today)},
        )
        # 管理员上班打卡通知 - 通知部门负责人和所属企业超级管理员
        if status_val == 'late':
            tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
            if tenant:
                from accounts.models import CustomUser
                notified_ids = set()
                try:
                    # 通知部门负责人
                    from org.models import UserDepartment
                    dept_rels = UserDepartment.objects.filter(user=request.user).select_related('department__manager')
                    for ud_rel in dept_rels:
                        mgr = ud_rel.department.manager
                        if mgr and mgr.id not in notified_ids and mgr.id != request.user.id:
                            notified_ids.add(mgr.id)
                            send_work_notification(
                                user_id=mgr.id,
                                title='迟到提醒',
                                content=f'{request.user.real_name or request.user.username} 今日上班打卡迟到',
                                notification_type='attendance',
                                related_url='/oa/attendance/',
                                extra_data={'user_id': request.user.id, 'status': 'late', 'date': str(today)},
                            )
                except Exception:
                    pass
                # 通知企业超级管理员
                # admins = CustomUser.objects.filter(
                #     Q(user_type='super_admin') & Q(tenant_memberships__tenant=tenant, tenant_memberships__is_active=True)
                # ).exclude(id=request.user.id).distinct()
                # for admin in admins:
                #     if admin.id not in notified_ids:
                #         send_work_notification(
                #             user_id=admin.id,
                #             title='迟到提醒',
                #             content=f'{request.user.real_name or request.user.username} 今日上班打卡迟到',
                #             notification_type='attendance',
                #             related_url='/oa/attendance/',
                #             extra_data={'user_id': request.user.id, 'status': 'late', 'date': str(today)},
                #         )
        data = AttendanceRecordSerializer(record, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201)

    @action(detail=False, methods=['post'])
    def clock_out(self, request):
        serializer = AttendanceClockSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        now = timezone.localtime(timezone.now())
        today = now.date()
        # 检查考勤配置（含用户主部门）
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        from org.models import UserDepartment
        primary_dept = UserDepartment.objects.filter(user=request.user, is_primary=True).select_related('department').first()
        dept_id = primary_dept.department_id if primary_dept else None
        config = self._get_attendance_config(tenant, dept_id, request.user)
        if config and not config.clock_out_enabled:
            return Response({'error': '该时段无需打卡（已配置为不打卡）', 'skip': True}, status=200)

        # 下班卡可重复打（防止误打下班卡），以最后一次打卡时间为准，上限由配置控制
        out_limit = config.clock_out_limit if config and config.clock_out_limit else 3
        if out_limit < 1:
            out_limit = 1
        today_out_count = AttendanceRecord.objects.filter(
            user=request.user, date=today, clock_type='clock_out').count()
        if today_out_count >= out_limit:
            return Response({
                'error': f'今日下班打卡次数已达上限（{out_limit}次）',
                'clock_out_count': today_out_count,
                'clock_out_limit': out_limit,
            }, status=400)
        # 查找最近一次上班打卡（支持跨夜班次）
        from datetime import timedelta as td
        clock_in_record = AttendanceRecord.objects.filter(
            user=request.user, clock_type='clock_in'
        ).order_by('-clock_time').first()
        if not clock_in_record or (clock_in_record.date < today - td(days=1)):
            return Response({'error': '请先上班打卡'}, status=400)
        # 使用配置时间或默认 18:00；夜班下班截止时间为上班日次日（如 20:00 上班 → 次日 06:00 下班）
        is_night = config is not None and config.shift_type == 'night'
        if config and config.clock_out_time:
            if is_night:
                from datetime import datetime as dt
                base_date = (clock_in_record.date + td(days=1)) if clock_in_record else today
                deadline = timezone.make_aware(dt.combine(base_date, config.clock_out_time))
            else:
                deadline = now.replace(hour=config.clock_out_time.hour, minute=config.clock_out_time.minute, second=0, microsecond=0)
        else:
            deadline = now.replace(hour=18, minute=0, second=0, microsecond=0)
        status_val = 'early_leave' if now < deadline else 'normal'
        logger.info(f'{request.user} 下班打卡 now: {now} deadline: {deadline} status: {status_val}')
        record = AttendanceRecord.objects.create(
            user=request.user, clock_type='clock_out', date=today,
            tenant=getattr(request, 'tenant', None) or request.user.get_active_tenant(),
            latitude=serializer.validated_data.get('latitude'),
            longitude=serializer.validated_data.get('longitude'),
            location=serializer.validated_data.get('location', ''),
            device=serializer.validated_data.get('device', ''),
            status=status_val, remark=serializer.validated_data.get('remark', ''),
            reverse_geocoding=serializer.validated_data.get('reverse_geocoding'),
            ip_address=serializer.validated_data.get('ip_address', '') or get_request_ip(request),
            user_agent=serializer.validated_data.get('user_agent', '') or request.META.get('HTTP_USER_AGENT', ''),
        )
        logger.info(f'{request.user} 下班打卡 {today} status={status_val}')
        location_text = serializer.validated_data.get('location', '')
        status_text = '早退' if status_val == 'early_leave' else '正常'
        loc = f' 位置：{location_text}' if location_text else ''
        send_work_notification(
            user_id=request.user.id,
            title='下班打卡成功',
            content=f'今日下班打卡 {status_text}{loc}',
            notification_type='attendance',
            related_url='/oa/attendance/',
            extra_data={'status': status_val, 'date': str(today)},
        )
        data = AttendanceRecordSerializer(record, context={'request': request}).data
        data['clock_out_count'] = today_out_count + 1
        data['clock_out_limit'] = out_limit
        data['clock_out_remaining'] = max(0, out_limit - (today_out_count + 1))
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201)

    @action(detail=False, methods=['post'])
    def makeup(self, request):
        """补卡：仅可补当月漏打的上班/下班卡，次数上限由考勤配置 makeup_allowance 控制"""
        from datetime import datetime as dt_datetime, time as dt_time
        clock_type = request.data.get('clock_type', '')
        date_str = request.data.get('date', '')
        if clock_type not in ('clock_in', 'clock_out'):
            return Response({'error': '请选择补卡类型（上班卡/下班卡）'}, status=400)
        try:
            makeup_date = dt_datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD'}, status=400)
        now = timezone.localtime(timezone.now())
        today = now.date()
        if makeup_date > today:
            return Response({'error': '不能为未来日期补卡'}, status=400)
        if makeup_date.year != today.year or makeup_date.month != today.month:
            return Response({'error': '仅支持补当月漏打的打卡记录'}, status=400)

        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        from org.models import UserDepartment
        primary_dept = UserDepartment.objects.filter(user=request.user, is_primary=True).select_related('department').first()
        dept_id = primary_dept.department_id if primary_dept else None
        config = self._get_attendance_config(tenant, dept_id, request.user)
        allowance = config.makeup_allowance if config and config.makeup_allowance is not None else 3
        if allowance <= 0:
            return Response({'error': '当前考勤未开启补卡功能（每月补卡次数为0）'}, status=400)

        used = AttendanceRecord.objects.filter(
            user=request.user,
            date__year=today.year, date__month=today.month,
        ).filter(Q(remark='补卡') | Q(makeup_at__isnull=False)).count()
        if used >= allowance:
            return Response({
                'error': f'本月补卡次数已用完（{allowance}次）',
                'makeup_used': used,
                'makeup_allowance': allowance,
                'makeup_remaining': 0,
            }, status=400)

        if clock_type == 'clock_in':
            cfg_time = config.clock_in_time if config and config.clock_in_time else dt_time(9, 0)
            label = '上班'
        else:
            cfg_time = config.clock_out_time if config and config.clock_out_time else dt_time(18, 0)
            label = '下班'

        # 已存在该日期该类型的打卡记录：正常记录不可补卡；迟到/早退记录可补卡修正为正常
        existing = AttendanceRecord.objects.filter(user=request.user, date=makeup_date, clock_type=clock_type).first()
        if existing:
            if existing.status == 'normal':
                return Response({'error': '该日期已正常打卡，无需补卡'}, status=400)
            # 迟到/早退 → 补卡修正为正常，并记录补卡时间
            existing.status = 'normal'
            existing.makeup_at = now
            existing.remark = '补卡'
            existing.save(update_fields=['status', 'makeup_at', 'remark'])
            record = existing
            logger.info(f'{request.user} 补卡(修正迟到/早退) {clock_type} {makeup_date}')
        else:
            clock_dt = dt_datetime.combine(makeup_date, cfg_time)
            aware_dt = timezone.make_aware(clock_dt) if timezone.is_naive(clock_dt) else clock_dt
            record = AttendanceRecord.objects.create(
                user=request.user, clock_type=clock_type, date=makeup_date,
                tenant=tenant,
                status='normal', remark='补卡',
                location='', device='补卡',
                makeup_at=now,
            )
            # clock_time 为 auto_now_add，创建后更新为配置时间
            record.clock_time = aware_dt
            record.save(update_fields=['clock_time'])
            logger.info(f'{request.user} 补卡 {clock_type} {makeup_date} 时间:{aware_dt}')
        send_work_notification(
            user_id=request.user.id,
            title='补卡成功',
            content=f'已补{label}卡（{makeup_date}）' + ('，原迟到/早退记录已修正为正常' if existing else f'，时间 {cfg_time.strftime("%H:%M")}'),
            notification_type='attendance',
            related_url='/oa/attendance/',
            extra_data={'clock_type': clock_type, 'date': str(makeup_date), 'makeup': True},
        )
        data = AttendanceRecordSerializer(record, context={'request': request}).data
        data['makeup_used'] = used + 1
        data['makeup_allowance'] = allowance
        data['makeup_remaining'] = allowance - (used + 1)
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201)

    @action(detail=False, methods=['get'])
    def today(self, request):
        today = timezone.localtime(timezone.now()).date()
        records = AttendanceRecord.objects.filter(user=request.user, date=today).order_by('clock_time')
        clock_in = records.filter(clock_type='clock_in').first()
        clock_out = records.filter(clock_type='clock_out').first()
        clock_out_count = records.filter(clock_type='clock_out').count()
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        from org.models import UserDepartment
        primary_dept = UserDepartment.objects.filter(user=request.user, is_primary=True).select_related('department').first()
        dept_id = primary_dept.department_id if primary_dept else None
        config = self._get_attendance_config(tenant, dept_id, request.user)
        shift_type = config.shift_type if config else 'day'
        out_limit = config.clock_out_limit if config and config.clock_out_limit else 3
        if out_limit < 1:
            out_limit = 1
        allowance = config.makeup_allowance if config and config.makeup_allowance is not None else 3
        makeup_used = AttendanceRecord.objects.filter(
            user=request.user, remark='补卡',
            date__year=today.year, date__month=today.month,
        ).count()
        # 待下班状态：当天已上班打卡未下班；或夜班时前一天晚上已上班打卡（工作跨夜），次日尚未下班
        from datetime import timedelta as td
        has_pending_clock_out = clock_out is None and (
            clock_in is not None
            or (shift_type == 'night' and AttendanceRecord.objects.filter(
                user=request.user, date=today - td(days=1), clock_type='clock_in').exists())
        )
        return Response({'encrypt': True, 'data': encrypt_data({
            'date': today.isoformat(),
            'clock_in': AttendanceRecordSerializer(clock_in, context={'request': request}).data if clock_in else None,
            'clock_out': AttendanceRecordSerializer(clock_out, context={'request': request}).data if clock_out else None,
            'has_clock_in': clock_in is not None,
            'has_clock_out': clock_out is not None,
            'has_pending_clock_out': has_pending_clock_out,
            'shift_type': shift_type,
            'clock_out_count': clock_out_count,
            'clock_out_limit': out_limit,
            'clock_out_remaining': max(0, out_limit - clock_out_count),
            'makeup_enabled': allowance > 0,
            'makeup_allowance': allowance,
            'makeup_used': makeup_used,
            'makeup_remaining': max(0, allowance - makeup_used),
        })})

    @action(detail=True, methods=['get'])
    def convert_coords(self, request, pk=None):
        """获取打卡记录的BD09坐标（用于百度地图），如需转换则调用百度坐标转换接口"""
        try:
            record = AttendanceRecord.objects.get(id=pk)
            if not self._can_view_record(request.user, record):
                return Response({'error': '暂无查看权限'}, status=403)

            # 如果已有BD09坐标，直接返回
            if record.bd09_latitude is not None and record.bd09_longitude is not None:
                return Response({
                    'bd09_latitude': record.bd09_latitude,
                    'bd09_longitude': record.bd09_longitude,
                    'source': 'cache',
                })

            # 没有GPS坐标则无法转换
            if record.latitude is None or record.longitude is None:
                return Response({'error': '无位置信息'}, status=400)

            from utils.reverse_geocoding_to_city import baidu_convert_WGS84ToBD09
            ak = getattr(settings, 'BAIDU_MAP_SERVER_AK', '')
            if not ak:
                return Response({'error': '百度地图AK未配置'}, status=500)

            result = baidu_convert_WGS84ToBD09(record.longitude, record.latitude, ak)
            if result and result.get('status') == 0:
                coords = result['result'][0]
                bd09_lng = coords['x']
                bd09_lat = coords['y']
                # 保存到记录
                record.bd09_longitude = bd09_lng
                record.bd09_latitude = bd09_lat
                record.save(update_fields=['bd09_longitude', 'bd09_latitude'])
                return Response({
                    'bd09_latitude': bd09_lat,
                    'bd09_longitude': bd09_lng,
                    'source': 'converted',
                })
            return Response({'error': '坐标转换失败'}, status=502)
        except AttendanceRecord.DoesNotExist:
            return Response({'error': '记录不存在'}, status=404)
        except Exception as e:
            logger.error(f'坐标转换异常: {e}')
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def export(self, request):
        from django.http import HttpResponse
        from urllib.parse import quote
        import csv
        user = request.user
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if user.user_type == 'super_admin':
            qs = AttendanceRecord.objects.select_related('user__department').all()
            if tenant:
                tenant_ids = [tenant.id]
                try:
                    sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                    if sub_ids:
                        tenant_ids.extend(sub_ids)
                except Exception:
                    pass
                qs = qs.filter(tenant_id__in=tenant_ids)
        else:
            # 部门负责人：仅可见所管理部门的成员（含子部门）；普通用户：仅自己
            managed_dept_ids = self._get_managed_department_ids(user)
            if managed_dept_ids:
                qs = AttendanceRecord.objects.select_related('user__department').filter(
                    user__department_relations__department_id__in=managed_dept_ids
                ).distinct()
            else:
                qs = AttendanceRecord.objects.select_related('user__department').filter(user=user)

        # Filter by selected record IDs
        record_ids_str = request.query_params.get('record_ids', '').strip()
        if record_ids_str:
            ids = [int(x) for x in record_ids_str.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(id__in=ids)
        qs = qs.order_by('-clock_time')[:10000]
        # Determine fields to export
        fields_param = request.query_params.get('fields', '').strip()
        field_map = {
            'user_name': lambda r: r.user.real_name or r.user.username,
            'department_name': lambda r: r.user.department.name if r.user.department else '',
            'date': lambda r: str(r.date),
            'clock_type_display': lambda r: r.get_clock_type_display(),
            'clock_time': lambda r: r.clock_time.strftime('%Y-%m-%d %H:%M:%S') if r.clock_time else '',
            'status': lambda r: {'normal': '正常', 'late': '迟到', 'early_leave': '早退'}.get(r.status, r.status),
            'location': lambda r: r.location or '',
            'device': lambda r: r.device or '',
        }
        field_labels = {
            'user_name': '用户', 'department_name': '部门', 'date': '日期',
            'clock_type_display': '打卡类型', 'clock_time': '打卡时间', 'status': '状态',
            'location': '位置', 'device': '设备',
        }

        if fields_param:
            selected_fields = [f.strip() for f in fields_param.split(',') if f.strip() in field_map]
        else:
            selected_fields = list(field_map.keys())
        # Generate filename with datetime
        now_dt = timezone.localtime(timezone.now())
        date_str = now_dt.strftime('%Y%m%d_%H%M%S')
        filename = f'考勤记录_{date_str}.csv'
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        writer = csv.writer(response)
        writer.writerow([field_labels.get(f, f) for f in selected_fields])
        for r in qs:
            try:
                row = [field_map[f](r) for f in selected_fields]
                logger.info(f'row: {row}')
                writer.writerow(row)
            except Exception as e:
                logger.error(f'导出行数据异常: {e}, record={r.id}')
                continue
        return response

    @action(detail=False, methods=['get'])
    def calendar_stats(self, request):
        """日历统计：按月返回每日打卡状态"""
        now = timezone.localtime(timezone.now())
        year = int(request.query_params.get('year', now.year))
        month = int(request.query_params.get('month', now.month))
        import calendar as cal_mod
        _, days_in_month = cal_mod.monthrange(year, month)
        from datetime import date, timedelta
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        records = AttendanceRecord.objects.filter(
            user=request.user,
            date__year=year, date__month=month
        ).order_by('date', 'clock_time')
        daily = {}
        for r in records:
            d_str = r.date.isoformat()
            if d_str not in daily:
                daily[d_str] = {'date': d_str, 'clock_in': None, 'clock_out': None, 'status': []}
            entry = daily[d_str]
            if r.clock_type == 'clock_in':
                entry['clock_in'] = {
                    'time': timezone.localtime(r.clock_time).strftime('%H:%M') if r.clock_time else '',
                    'status': r.status, 'location': r.location or '',
                }
            elif r.clock_type == 'clock_out':
                entry['clock_out'] = {
                    'time': timezone.localtime(r.clock_time).strftime('%H:%M') if r.clock_time else '',
                    'status': r.status, 'location': r.location or '',
                }
        # 已批准的请假日期集合（工作日请假 → 标记为请假而非缺卡）
        from .holidays import is_rest_day
        month_start = date(year, month, 1)
        month_end = date(year, month, days_in_month)
        leave_dates = set()
        try:
            leaves = ApprovalRequest.objects.filter(
                applicant=request.user, approval_type='leave', status='approved',
                start_date__lte=month_end, end_date__gte=month_start,
            )
            for lv in leaves:
                s = max(lv.start_date or month_start, month_start)
                e = min(lv.end_date or month_end, month_end)
                dd = s
                while dd <= e:
                    leave_dates.add(dd.isoformat())
                    dd += timedelta(days=1)
        except Exception:
            pass

        # 构建每日状态
        result = []
        for d in range(1, days_in_month + 1):
            dt = date(year, month, d)
            d_str = dt.isoformat()
            today = now.date()
            entry = daily.get(d_str, {'date': d_str, 'clock_in': None, 'clock_out': None})
            ci = entry.get('clock_in')
            co = entry.get('clock_out')
            day_status = 'normal'
            day_label = '正常'
            if dt > today:
                day_status = 'future'
                day_label = '未到'
            elif d_str in leave_dates:
                day_status = 'leave'
                day_label = '请假'
            elif is_rest_day(dt):
                # 星期日/法定节假日为休息日，不算缺卡/缺勤
                day_status = 'rest'
                day_label = '休息'
            elif ci and co:
                if ci['status'] == 'late' or co['status'] == 'early_leave':
                    day_status = 'late'
                    day_label = '异常'
                    if ci['status'] == 'late' and co['status'] == 'early_leave':
                        day_label = '迟到+早退'
                    elif ci['status'] == 'late':
                        day_label = '迟到'
                    elif co['status'] == 'early_leave':
                        day_label = '早退'
                else:
                    day_status = 'normal'
                    day_label = '正常'
            elif ci and not co:
                day_status = 'miss_clock'
                day_label = '缺卡'
            elif not ci and dt < today:
                day_status = 'absent'
                day_label = '缺勤'
            else:
                day_status = 'none'
                day_label = ''
            entry['day_status'] = day_status
            entry['day_label'] = day_label
            result.append(entry)
        # 统计汇总
        summary = {
            'total': days_in_month,
            'normal': sum(1 for e in result if e['day_status'] == 'normal'),
            'late': sum(1 for e in result if e['day_status'] == 'late'),
            'miss_clock': sum(1 for e in result if e['day_status'] == 'miss_clock'),
            'absent': sum(1 for e in result if e['day_status'] == 'absent'),
            'future': sum(1 for e in result if e['day_status'] == 'future'),
            'leave': sum(1 for e in result if e['day_status'] == 'leave'),
            'rest': sum(1 for e in result if e['day_status'] == 'rest'),
        }
        # 当前用户生效考勤配置的打卡时间（补卡日历据此判断「今日」漏卡是否可补、并提示将按配置时间补卡）
        from org.models import UserDepartment as _UD
        _primary = _UD.objects.filter(user=request.user, is_primary=True).select_related('department').first()
        _dept_id = _primary.department_id if _primary else None
        _cfg = self._get_attendance_config(tenant, _dept_id, request.user)
        _cfg_time = {
            'clock_in': _cfg.clock_in_time.strftime('%H:%M') if _cfg and _cfg.clock_in_time else '09:00',
            'clock_out': _cfg.clock_out_time.strftime('%H:%M') if _cfg and _cfg.clock_out_time else '18:00',
            'shift_type': _cfg.shift_type if _cfg else 'day',
        }
        return Response({'encrypt': True, 'data': encrypt_data({
            'year': year, 'month': month, 'days': result, 'summary': summary,
            'config_time': _cfg_time,
        })})

    @action(detail=False, methods=['get'])
    def calendar_day_detail(self, request):
        """日历日期详情：返回该日打卡记录及加班审批信息"""
        date_str = request.query_params.get('date', '')
        if not date_str:
            return Response({'error': '缺少日期参数'}, status=400)
        from datetime import date as dt_date
        try:
            target_date = dt_date.fromisoformat(date_str)
        except ValueError:
            return Response({'error': '日期格式错误'}, status=400)
        records = AttendanceRecord.objects.filter(
            user=request.user, date=target_date
        ).order_by('clock_time')
        clock_in = None
        clock_out = None
        for r in records:
            if r.clock_type == 'clock_in':
                clock_in = AttendanceRecordSerializer(r, context={'request': request}).data
            elif r.clock_type == 'clock_out':
                clock_out = AttendanceRecordSerializer(r, context={'request': request}).data
        # 查找该日的加班审批（加班类型，日期覆盖该日）
        overtime_info = None
        try:
            overtime_reqs = ApprovalRequest.objects.filter(
                applicant=request.user,
                approval_type='overtime',
                status='approved',
                start_date__lte=target_date,
                end_date__gte=target_date,
            )
            for req in overtime_reqs:
                overtime_info = {
                    'id': req.id,
                    'title': req.title,
                    'duration': float(req.duration) if req.duration else 0,
                    'content': req.content or '',
                    'status': req.status,
                }
                break
        except Exception:
            pass
        result = {
            'date': date_str,
            'clock_in': clock_in,
            'clock_out': clock_out,
            'overtime': overtime_info,
        }
        return Response({'encrypt': True, 'data': encrypt_data(result)})

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        now = timezone.now()
        year = int(request.query_params.get('year', now.year))
        month = int(request.query_params.get('month', now.month))
        qs = AttendanceRecord.objects.filter(user=request.user, date__year=year, date__month=month)
        return Response({'encrypt': True, 'data': encrypt_data({
            'year': year, 'month': month,
            'total_days': qs.values('date').distinct().count(),
            'late_count': qs.filter(status='late').count(),
            'early_leave_count': qs.filter(status='early_leave').count(),
            'clock_in_count': qs.filter(clock_type='clock_in').count(),
            'clock_out_count': qs.filter(clock_type='clock_out').count(),
        })})

    # ──────── 考勤配置 ────────

    def _get_attendance_config(self, tenant, department_id=None, user=None):
        """按优先级获取考勤配置：个人配置 > 部门配置 > 子企业配置 > 企业默认 > 父级回溯"""
        if user:
            try:
                return UserAttendanceConfig.objects.get(user=user)
            except UserAttendanceConfig.DoesNotExist:
                pass
        if not tenant:
            return None
        # 1. 部门级配置
        if department_id:
            try:
                return AttendanceConfig.objects.get(tenant=tenant, department_id=department_id)
            except AttendanceConfig.DoesNotExist:
                pass
        # 2. 子企业专属配置（配置在 parent tenant 上，sub_tenant=当前企业）
        if tenant.parent:
            try:
                return AttendanceConfig.objects.get(
                    tenant=tenant.parent, sub_tenant=tenant, department__isnull=True)
            except AttendanceConfig.DoesNotExist:
                pass
        # 3. 当前企业默认配置
        try:
            return AttendanceConfig.objects.get(
                tenant=tenant, sub_tenant__isnull=True, department__isnull=True)
        except AttendanceConfig.DoesNotExist:
            pass
        # 4. 父级集团回溯
        if tenant.parent:
            parent = self._get_parent_tenant(tenant)
            if parent:
                return self._get_attendance_config(parent, user=user)
        return None

    def _get_parent_tenant(self, tenant):
        return tenant.parent if tenant.parent else None

    @action(detail=False, methods=['get'])
    def attendance_configs(self, request):
        """获取考勤配置列表 + 子公司列表。
        超管：可查看/配置集团、子公司、部门全部配置；
        普通管理员：仅可查看/配置本部门（含子部门）的部门级配置，集团/子公司配置不返回。"""
        from .serializers import AttendanceConfigSerializer
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'results': [], 'sub_tenants': []})
        from django.db.models import Q
        managed = None
        if request.user.user_type != 'super_admin':
            managed = self._get_managed_department_ids(request.user)
            query = Q(tenant=tenant, sub_tenant=None, department_id__in=managed) if managed else Q(pk__in=[])
        else:
            query = Q(tenant=tenant)
            try:
                sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                if sub_ids:
                    query |= Q(tenant_id__in=sub_ids)
            except Exception:
                pass
        configs = AttendanceConfig.objects.filter(query).select_related('department', 'sub_tenant')
        data = AttendanceConfigSerializer(configs, many=True).data
        extra = {'sub_tenants': list(tenant.sub_tenants.filter(is_active=True).values(
            'id', 'name', 'short_name', 'tenant_type'))} if hasattr(tenant, 'sub_tenants') else {'sub_tenants': []}
        if managed is not None:
            extra['managed_dept_ids'] = managed
        return Response({'results': data, **extra})

    @action(detail=False, methods=['post'])
    def save_attendance_config(self, request):
        """保存考勤配置"""
        if request.user.user_type not in ('super_admin', 'admin'):
            return Response({'error': '仅企业超级管理员或管理员可操作'}, status=403)
        from accounts.models import Department, Tenant
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'error': '未找到所属企业'}, status=400)
        sub_tenant_id = request.data.get('sub_tenant_id')
        sub_tenant_obj = None
        if sub_tenant_id:
            try:
                sub_tenant_obj = Tenant.objects.get(id=int(sub_tenant_id))
                if sub_tenant_obj.parent_id != tenant.id:
                    return Response({'error': '指定的子公司不属于当前企业集团'}, status=400)
            except (ValueError, Tenant.DoesNotExist):
                return Response({'error': '子公司不存在'}, status=400)
        department_id = request.data.get('department_id')
        dept = None
        if department_id:
            try:
                dept = Department.objects.get(id=int(department_id), tenant=tenant)
            except (ValueError, Department.DoesNotExist):
                return Response({'error': '部门不存在'}, status=400)
        # 普通管理员：只能配置本部门（含子部门）考勤规则，集团/子公司不可配置
        if request.user.user_type != 'super_admin':
            if sub_tenant_id:
                return Response({'error': '普通管理员无权配置子公司考勤规则'}, status=403)
            managed = self._get_managed_department_ids(request.user)
            if not dept or dept.id not in managed:
                return Response({'error': '普通管理员只能配置本部门（含子部门）的考勤规则'}, status=403)
        from .serializers import AttendanceConfigSerializer
        clock_in_time = request.data.get('clock_in_time')
        clock_out_time = request.data.get('clock_out_time')
        from datetime import time as dt_time
        if clock_in_time:
            try:
                parts = clock_in_time.split(':')
                clock_in_time = dt_time(int(parts[0]), int(parts[1]), 0)
            except (ValueError, IndexError):
                return Response({'error': '上班打卡时间格式错误，请使用 HH:MM 格式'}, status=400)
        if clock_out_time:
            try:
                parts = clock_out_time.split(':')
                clock_out_time = dt_time(int(parts[0]), int(parts[1]), 0)
            except (ValueError, IndexError):
                return Response({'error': '下班打卡时间格式错误，请使用 HH:MM 格式'}, status=400)
        makeup_allowance = request.data.get('makeup_allowance', 3)
        clock_out_limit = request.data.get('clock_out_limit', 3)
        try:
            makeup_allowance = int(makeup_allowance)
        except (ValueError, TypeError):
            return Response({'error': '每月补卡次数格式错误'}, status=400)
        if makeup_allowance < 0:
            return Response({'error': '每月补卡次数不能小于0（0表示禁用补卡）'}, status=400)
        try:
            clock_out_limit = int(clock_out_limit)
        except (ValueError, TypeError):
            return Response({'error': '下班卡最多打卡次数格式错误'}, status=400)
        if clock_out_limit < 1:
            return Response({'error': '下班卡最多打卡次数不能小于1'}, status=400)
        shift_type = request.data.get('shift_type', 'day')
        if shift_type not in ('day', 'night'):
            return Response({'error': '班次类型错误（白班/夜班）'}, status=400)
        defaults = {
            'clock_in_enabled': request.data.get('clock_in_enabled', True),
            'clock_in_time': clock_in_time,
            'clock_out_enabled': request.data.get('clock_out_enabled', True),
            'clock_out_time': clock_out_time,
            'makeup_allowance': makeup_allowance,
            'clock_out_limit': clock_out_limit,
            'shift_type': shift_type,
        }
        try:
            config, created = AttendanceConfig.objects.update_or_create(
                tenant=tenant,
                sub_tenant=sub_tenant_obj,
                department=dept,
                defaults=defaults,
            )
            data = AttendanceConfigSerializer(config).data
            return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201 if created else 200)
        except Exception as e:
            logger.error(f'保存考勤配置失败: {e}')
            return Response({'error': f'保存配置失败: {str(e)}'}, status=400)

    @action(detail=True, methods=['delete'])
    def delete_attendance_config(self, request, pk=None):
        """删除考勤配置（超管可删全部；普通管理员仅可删本部门（含子部门）配置）"""
        try:
            config = AttendanceConfig.objects.get(id=pk)
        except AttendanceConfig.DoesNotExist:
            return Response({'error': '配置不存在'}, status=404)
        if request.user.user_type != 'super_admin':
            managed = self._get_managed_department_ids(request.user)
            if config.sub_tenant_id or not config.department_id or config.department_id not in managed:
                return Response({'error': '普通管理员只能删除本部门（含子部门）的考勤配置'}, status=403)
        config.delete()
        return Response({'message': 'ok'})

    @action(detail=False, methods=['get'])
    def resolve_my_config(self, request):
        """获取当前用户的考勤配置（按优先级解析）"""
        from org.models import UserDepartment
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        primary_dept = UserDepartment.objects.filter(user=request.user, is_primary=True).select_related('department').first()
        dept_id = primary_dept.department_id if primary_dept else None
        config = self._get_attendance_config(tenant, dept_id, request.user)
        # 返回配置和当前企业的默认时间
        result = {
            'config': config,
            'default_clock_in_time': '09:00',
            'default_clock_out_time': '18:00',
        }
        if config:
            if isinstance(config, UserAttendanceConfig):
                from .serializers import UserAttendanceConfigSerializer
                result['config'] = UserAttendanceConfigSerializer(config).data
            else:
                from .serializers import AttendanceConfigSerializer
                result['config'] = AttendanceConfigSerializer(config).data
        return Response({'encrypt': True, 'data': encrypt_data(result)})

    @action(detail=False, methods=['get'])
    def members(self, request):
        """当前企业成员列表，供个人考勤配置选择成员（普通管理员仅返回本部门成员）"""
        from accounts.models import CustomUser
        from django.db.models import Q
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        keyword = request.query_params.get('search', '').strip()
        qs = CustomUser.objects.filter(is_active=True)
        if tenant:
            qs = qs.filter(tenant_memberships__tenant=tenant, tenant_memberships__is_active=True)
        if request.user.user_type != 'super_admin':
            managed = self._get_managed_department_ids(request.user)
            if managed:
                from org.models import UserDepartment
                member_ids = set(UserDepartment.objects.filter(department_id__in=managed).values_list('user_id', flat=True))
                qs = qs.filter(id__in=member_ids)
            else:
                qs = qs.none()
        if keyword:
            qs = qs.filter(Q(real_name__icontains=keyword) | Q(username__icontains=keyword))
        qs = qs.distinct().order_by('real_name', 'username')[:100]
        ids = [u.id for u in qs]
        from org.models import UserDepartment
        dept_map = dict(UserDepartment.objects.filter(user_id__in=ids, is_primary=True).select_related('department').values_list('user_id', 'department__name'))
        results = [{
            'id': u.id,
            'name': u.real_name or u.username,
            'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
            'position': u.position or '',
            'department_name': dept_map.get(u.id) or '',
        } for u in qs]
        return Response({'results': results})

    @action(detail=False, methods=['get'])
    def user_attendance_configs(self, request):
        """个人考勤配置列表（普通管理员仅返回本部门成员的配置，超管返回全部）"""
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        qs = UserAttendanceConfig.objects.select_related('user').all()
        if tenant:
            qs = qs.filter(user__tenant_memberships__tenant=tenant, user__tenant_memberships__is_active=True)
        if request.user.user_type != 'super_admin':
            managed = self._get_managed_department_ids(request.user)
            if managed:
                from org.models import UserDepartment
                member_ids = set(UserDepartment.objects.filter(department_id__in=managed).values_list('user_id', flat=True))
                qs = qs.filter(user_id__in=member_ids)
            else:
                qs = qs.none()
        qs = qs.distinct().order_by('-updated_at')
        from .serializers import UserAttendanceConfigSerializer
        data = UserAttendanceConfigSerializer(qs, many=True).data
        return Response({'results': data})

    @action(detail=False, methods=['post'])
    def save_user_attendance_config(self, request):
        """保存个人考勤配置（如部门内个别成员单独上夜班）"""
        if request.user.user_type not in ('super_admin', 'admin'):
            return Response({'error': '仅企业超级管理员或管理员可操作'}, status=403)
        from accounts.models import CustomUser
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        try:
            target_user = CustomUser.objects.get(id=int(request.data.get('user_id')))
        except (ValueError, TypeError, CustomUser.DoesNotExist):
            return Response({'error': '成员不存在'}, status=400)
        if tenant:
            if not target_user.tenant_memberships.filter(tenant=tenant, is_active=True).exists():
                return Response({'error': '该成员不属于当前企业'}, status=400)
        # 普通管理员：只能为本人所在部门（含子部门）的成员配置个人考勤规则
        if request.user.user_type != 'super_admin':
            managed = self._get_managed_department_ids(request.user)
            from org.models import UserDepartment
            if not managed or not UserDepartment.objects.filter(user_id=target_user.id, department_id__in=managed).exists():
                return Response({'error': '普通管理员只能配置本部门成员的考勤规则'}, status=403)

        shift_type = request.data.get('shift_type', 'day')
        if shift_type not in ('day', 'night'):
            return Response({'error': '班次类型错误（白班/夜班）'}, status=400)
        from datetime import time as dt_time
        clock_in_time = request.data.get('clock_in_time')
        clock_out_time = request.data.get('clock_out_time')
        if clock_in_time:
            try:
                parts = clock_in_time.split(':')
                clock_in_time = dt_time(int(parts[0]), int(parts[1]), 0)
            except (ValueError, IndexError):
                return Response({'error': '上班打卡时间格式错误，请使用 HH:MM 格式'}, status=400)
        if clock_out_time:
            try:
                parts = clock_out_time.split(':')
                clock_out_time = dt_time(int(parts[0]), int(parts[1]), 0)
            except (ValueError, IndexError):
                return Response({'error': '下班打卡时间格式错误，请使用 HH:MM 格式'}, status=400)
        try:
            makeup_allowance = int(request.data.get('makeup_allowance', 3))
        except (ValueError, TypeError):
            return Response({'error': '每月补卡次数格式错误'}, status=400)
        if makeup_allowance < 0:
            return Response({'error': '每月补卡次数不能小于0（0表示禁用补卡）'}, status=400)
        try:
            clock_out_limit = int(request.data.get('clock_out_limit', 3))
        except (ValueError, TypeError):
            return Response({'error': '下班卡最多打卡次数格式错误'}, status=400)
        if clock_out_limit < 1:
            return Response({'error': '下班卡最多打卡次数不能小于1'}, status=400)

        config, created = UserAttendanceConfig.objects.update_or_create(
            user=target_user,
            defaults={
                'shift_type': shift_type,
                'clock_in_enabled': request.data.get('clock_in_enabled', True),
                'clock_in_time': clock_in_time,
                'clock_out_enabled': request.data.get('clock_out_enabled', True),
                'clock_out_time': clock_out_time,
                'makeup_allowance': makeup_allowance,
                'clock_out_limit': clock_out_limit,
            },
        )
        from .serializers import UserAttendanceConfigSerializer
        data = UserAttendanceConfigSerializer(config).data
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201 if created else 200)

    @action(detail=True, methods=['delete'])
    def delete_user_attendance_config(self, request, pk=None):
        """删除个人考勤配置"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可操作'}, status=403)
        try:
            config = UserAttendanceConfig.objects.get(id=pk)
            config.delete()
            return Response({'message': 'ok'})
        except UserAttendanceConfig.DoesNotExist:
            return Response({'error': '配置不存在'}, status=404)


class ApprovalTypeViewSet(viewsets.ViewSet):
    """审批类型管理（内置 + 企业自定义）"""
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        """返回当前企业可用审批类型：全局内置 + 当前企业自定义（含 form_schema）。

        ?scope=manage：类型管理用，包含未启用的自定义类型（保证可重新编辑/启用）。
        默认仅返回启用类型（供新建审批选择）。
        """
        ensure_builtin_types()
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        scope = request.query_params.get('scope', '')
        if scope == 'manage':
            # 类型管理：内置 + 当前企业全部自定义（含未启用）
            qs = ApprovalType.objects.all()
            if tenant:
                qs = qs.filter(Q(tenant=tenant) | Q(tenant__isnull=True)).order_by('-tenant_id', 'sort_order', 'id')
            else:
                qs = qs.filter(tenant__isnull=True).order_by('sort_order', 'id')
        else:
            qs = ApprovalType.objects.filter(enabled=True)
            if tenant:
                qs = qs.filter(Q(tenant=tenant) | Q(tenant__isnull=True)).order_by('-tenant_id', 'sort_order', 'id')
            else:
                qs = qs.filter(tenant__isnull=True).order_by('sort_order', 'id')
        data = [{
            'id': t.id,
            'code': t.code,
            'name': t.name,
            'icon': t.icon,
            'color': t.color,
            'description': t.description or '',
            'is_builtin': t.is_builtin,
            'form_schema': t.form_schema or [],
            'sort_order': t.sort_order,
            'enabled': t.enabled,
        } for t in qs]
        return Response({'results': data, 'count': len(data)})

    @staticmethod
    def _validate_schema(schema):
        """校验 form_schema 结构，返回错误字符串或 None"""
        if not isinstance(schema, list):
            return '表单字段定义必须为数组'
        seen_keys = set()
        for f in schema:
            if not isinstance(f, dict):
                return '字段定义格式错误'
            if not f.get('key') or not f.get('label'):
                return '每个字段必须包含 key 和 label'
            if f['key'] in seen_keys:
                return f'字段编码(key)「{f["key"]}」存在重复'
            seen_keys.add(f['key'])
            ftype = f.get('type')
            if ftype not in dict(ApprovalType.FIELD_TYPES):
                return f'字段类型 {ftype} 不支持'
            if ftype in ('select', 'radio', 'checkbox') and not f.get('options'):
                return f'字段 {f.get("label")} 需要配置选项'
            if ftype == 'struct_table':
                cols = f.get('columns')
                if not isinstance(cols, list) or not cols:
                    return f'字段 {f.get("label")} 需要配置明细列'
                for c in cols:
                    if not isinstance(c, dict) or not c.get('key') or not c.get('label'):
                        return f'字段 {f.get("label")} 的明细列需包含 key 和 label'
                    if c.get('type') and c.get('type') not in ('text', 'number', 'amount', 'item'):
                        return f'字段 {f.get("label")} 的明细列类型 {c.get("type")} 不支持'
        return None

    @staticmethod
    def _normalize_schema(schema):
        """规范 form_schema：options 统一为数组，补齐 required/placeholder 等默认值"""
        if not isinstance(schema, list):
            return schema
        for f in schema:
            if not isinstance(f, dict):
                continue
            opts = f.get('options')
            if isinstance(opts, str):
                f['options'] = [s.strip() for s in opts.replace('，', ',').split(',') if s.strip()]
            if not isinstance(f.get('options'), list):
                f['options'] = []
            f.setdefault('required', False)
            f.setdefault('placeholder', '')
            # 结构化明细：列定义规范化，缺失时按报销项目默认三列
            if f.get('type') == 'struct_table':
                cols = f.get('columns')
                if not isinstance(cols, list) or not cols:
                    cols = [
                        {'key': 'name', 'label': '项目名称', 'type': 'text'},
                        {'key': 'amount', 'label': '金额', 'type': 'amount'},
                        {'key': 'remark', 'label': '备注', 'type': 'text'},
                    ]
                normalized = []
                for c in cols:
                    if not isinstance(c, dict) or not c.get('key'):
                        continue
                    normalized.append({
                        'key': c.get('key'),
                        'label': c.get('label') or c.get('key'),
                        'type': c.get('type') if c.get('type') in ('text', 'number', 'amount', 'item') else 'text',
                    })
                f['columns'] = normalized
        return schema

    @staticmethod
    def _admin(request):
        return request.user.user_type in ('super_admin', 'admin')

    def create(self, request):
        if not self._admin(request):
            return Response({'error': '仅企业超级管理员或管理员可操作'}, status=403)
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'error': '未找到所属企业'}, status=400)
        code = (request.data.get('code') or '').strip()
        name = (request.data.get('name') or '').strip()
        if not code or not name:
            return Response({'error': '类型编码和名称不能为空'}, status=400)
        if ApprovalType.objects.filter(tenant=tenant, code=code).exists():
            return Response({'error': '该企业下已存在此编码的审批类型'}, status=400)
        schema = self._normalize_schema(request.data.get('form_schema', []) or [])
        err = self._validate_schema(schema)
        if err:
            return Response({'error': err}, status=400)
        t = ApprovalType.objects.create(
            tenant=tenant, code=code, name=name,
            icon=(request.data.get('icon') or 'fa-file-lines'),
            color=(request.data.get('color') or '#409EFF'),
            description=(request.data.get('description') or ''),
            enabled=bool(request.data.get('enabled', True)),
            form_schema=schema,
            sort_order=int(request.data.get('sort_order') or 0),
        )
        return Response({'message': '创建成功', 'code': t.code}, status=201)

    def update(self, request, pk=None):
        if not self._admin(request):
            return Response({'error': '仅企业超级管理员或管理员可操作'}, status=403)
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        try:
            t = ApprovalType.objects.get(pk=pk)
        except ApprovalType.DoesNotExist:
            return Response({'error': '类型不存在'}, status=404)
        if t.is_builtin:
            # 内置类型：仅允许切换启用/禁用，其它字段（名称/编码/图标/颜色/表单）锁定
            if set(request.data.keys()) - {'enabled'}:
                return Response({'error': '内置类型仅可切换启用状态'}, status=403)
        elif tenant and t.tenant_id != tenant.id:
            return Response({'error': '无权限修改该类型'}, status=403)
        if 'code' in request.data and (request.data.get('code') or '').strip() != t.code:
            new_code = (request.data.get('code') or '').strip()
            if ApprovalType.objects.filter(tenant=t.tenant_id, code=new_code).exclude(pk=t.pk).exists():
                return Response({'error': '该编码已存在'}, status=400)
            t.code = new_code
        if 'name' in request.data and (request.data.get('name') or '').strip():
            t.name = request.data['name'].strip()
        if 'icon' in request.data:
            t.icon = request.data.get('icon') or t.icon
        if 'color' in request.data:
            t.color = request.data.get('color') or t.color
        if 'description' in request.data:
            t.description = request.data.get('description') or ''
        if 'enabled' in request.data:
            t.enabled = bool(request.data.get('enabled'))
        if 'form_schema' in request.data:
            schema = self._normalize_schema(request.data.get('form_schema') or [])
            err = self._validate_schema(schema)
            if err:
                return Response({'error': err}, status=400)
            t.form_schema = schema
        t.save()
        return Response({'message': '更新成功'})

    def destroy(self, request, pk=None):
        if not self._admin(request):
            return Response({'error': '仅企业超级管理员或管理员可操作'}, status=403)
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        try:
            t = ApprovalType.objects.get(pk=pk)
        except ApprovalType.DoesNotExist:
            return Response({'error': '类型不存在'}, status=404)
        if t.is_builtin or (tenant and t.tenant_id != tenant.id):
            return Response({'error': '内置类型不可删除或无权限'}, status=403)
        t.delete()
        return Response({'message': '删除成功'})


class ApprovalViewSet(viewsets.ViewSet):
    """OA审批视图集"""
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        user = request.user
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        search = request.query_params.get('search', '').strip()
        status_filter = request.query_params.get('status', '').strip()
        type_filter = request.query_params.get('type', '').strip()

        if user.user_type == 'super_admin':
            # 仅超级管理员可查看全部审批（含集团及子企业）
            qs = ApprovalRequest.objects.select_related('applicant', 'department').all()
            tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
            if tenant:
                tenant_ids = [tenant.id]
                try:
                    sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                    if sub_ids:
                        tenant_ids.extend(sub_ids)
                except Exception:
                    pass
                qs = qs.filter(tenant_id__in=tenant_ids)
        else:
            # 其他用户（含管理员）：仅可见自己发起的审批
            my_approval_ids = ApprovalRequest.objects.filter(applicant=user).values_list('id', flat=True)
            # 审批人可见：
            #  ① 已到自己的节点且审批进行中（并行全部节点、顺序仅当前节点）；
            #  ② 自己已通过/驳回过该审批（无论审批当前状态/节点）。
            # 审批人可见规则：
            #  ① 已通过：曾为该审批审批人的均可查看；
            #  ② 已驳回：仅驳回该审批的审批人可查看；
            #  ③ 已撤回：审批到达自己节点时被发起人撤回的审批人可查看；
            #  ④ 进行中：已到达自己节点（并行全部节点、顺序仅当前及已到达节点）。
            active_assignee_ids = ApprovalAssignee.objects.filter(
                user=user
            ).exclude(
                node__request__applicant=user
            ).filter(
                Q(node__request__status='approved') |
                Q(node__request__status='rejected', status='rejected') |
                Q(
                    node__request__status='cancelled',
                    node__request__approval_mode='parallel'
                ) |
                Q(
                    node__request__status='cancelled',
                    node__request__approval_mode='sequential',
                    node__order__lte=models.F('node__request__current_node_order')
                ) |
                Q(
                    node__request__status__in=['pending', 'deferred', 'processing'],
                    node__request__approval_mode='parallel'
                ) |
                Q(
                    node__request__status__in=['pending', 'deferred', 'processing'],
                    node__request__approval_mode='sequential',
                    node__order__lte=models.F('node__request__current_node_order')
                )
            ).values_list('node__request_id', flat=True).distinct()

            # 抄送人：仅在整个审核完成后（已通过/已驳回）才能看到
            cc_ids = ApprovalCarbonCopy.objects.filter(
                cc_user=user, request__status__in=['approved', 'rejected']
            ).values_list('request_id', flat=True).distinct()
            # 抄送部门负责人/副负责人：同样审核完成后可见
            cc_dept_ids = []
            try:
                from org.models import UserDepartment
                user_dept_ids = list(UserDepartment.objects.filter(user=user).values_list('department_id', flat=True))
                from accounts.models import Department
                managed_dept_ids = list(Department.objects.filter(
                    Q(manager=user) | Q(deputy_managers=user)
                ).values_list('id', flat=True))
                all_dept_ids = list(set(user_dept_ids + managed_dept_ids))
                if all_dept_ids:
                    cc_dept_ids = list(ApprovalCarbonCopy.objects.filter(
                        cc_department_id__in=all_dept_ids, request__status__in=['approved', 'rejected']
                    ).values_list('request_id', flat=True).distinct())
            except Exception:
                pass

            qs = ApprovalRequest.objects.select_related('applicant', 'department').filter(
                Q(id__in=my_approval_ids) | Q(id__in=active_assignee_ids) |
                Q(id__in=cc_ids) | Q(id__in=cc_dept_ids)
            )

        scope_filter = request.query_params.get('scope', '')
        if scope_filter == 'mine':
            # 我发起的：仅本人发起的审批
            qs = qs.filter(applicant=user)
        elif scope_filter == 'cc':
            # 抄送我的：审批结束后（已通过/已驳回）的抄送记录（含部门抄送）
            cc_user_ids = list(ApprovalCarbonCopy.objects.filter(
                cc_user=user, request__status__in=['approved', 'rejected']
            ).values_list('request_id', flat=True).distinct())
            cc_dept_ids = []
            try:
                from org.models import UserDepartment
                from accounts.models import Department
                user_dept_ids = list(UserDepartment.objects.filter(user=user).values_list('department_id', flat=True))
                managed_dept_ids = list(Department.objects.filter(
                    Q(manager=user) | Q(deputy_managers=user)
                ).values_list('id', flat=True))
                all_dept_ids = list(set(user_dept_ids + managed_dept_ids))
                if all_dept_ids:
                    cc_dept_ids = list(ApprovalCarbonCopy.objects.filter(
                        cc_department_id__in=all_dept_ids, request__status__in=['approved', 'rejected']
                    ).values_list('request_id', flat=True).distinct())
            except Exception:
                pass
            qs = qs.filter(Q(id__in=cc_user_ids) | Q(id__in=cc_dept_ids))
        elif scope_filter == 'todo':
            # 待我审批：我仍有未处理（待审批）的审批人节点（仅统计已到达/当前节点，排除顺序审批中尚未到达的后续节点）
            todo_ids = ApprovalAssignee.objects.filter(
                user=user, status='pending',
                node__order__lte=models.F('node__request__current_node_order')
            ).values_list('node__request_id', flat=True).distinct()
            qs = qs.filter(id__in=todo_ids)
        elif scope_filter == 'done':
            # 我已审批：我已处理过（通过/驳回/暂缓/办理中）的审批人节点
            done_ids = ApprovalAssignee.objects.filter(
                user=user
            ).exclude(status='pending').values_list('node__request_id', flat=True).distinct()
            qs = qs.filter(id__in=done_ids)

        if status_filter:
            qs = qs.filter(status=status_filter)
        if type_filter:
            qs = qs.filter(approval_type=type_filter)
        if search:
            qs = qs.filter(
                Q(title__icontains=search) |
                Q(applicant__username__icontains=search) |
                Q(applicant__real_name__icontains=search)
            )

        qs = qs.order_by('-updated_at')
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        start = (page - 1) * page_size
        items = qs[start:start + page_size]
        results = ApprovalListSerializer(items, many=True, context={'request': request}).data

        return Response({'encrypt': True, 'data': encrypt_data({
            'results': results, 'count': total, 'page': page,
            'page_size': page_size, 'total_pages': total_pages,
        })})

    def retrieve(self, request, pk=None):
        try:
            approval = ApprovalRequest.objects.select_related(
                'applicant', 'approver', 'department'
            ).prefetch_related('logs__operator', 'approval_nodes__assignees__user').get(id=pk)
            # 权限：仅超级管理员可查看全部；申请人可见自己发起；
            # 审批人：节点到达自己且审批进行中可见，或自己已通过/驳回过该审批可见；抄送人仅审核完成后可见。
            is_cc = approval.carbon_copies.filter(
                Q(cc_user=request.user) | Q(cc_department__manager=request.user) | Q(cc_department__deputy_managers=request.user)
            ).exists()
            # 可查看规则：①超管全量；②发起人全流程；③审批人到达自己节点（含后续流程）即可查看；
            # ④抄送人仅在审批结束后（已通过/已驳回）可查看。
            can_view = (
                request.user.user_type == 'super_admin'
                or approval.applicant == request.user
                or self._is_approver_reached(approval, request.user)
                or (is_cc and approval.status in ('approved', 'rejected'))
            )
            if not can_view:
                # 审批人是该审批的经办节点但未到当前节点（顺序审批时后续节点）→ 提示未到达节点
                is_any_assignee = ApprovalAssignee.objects.filter(
                    node__request=approval, user=request.user
                ).exists()
                if is_any_assignee:
                    return Response({'error': '您不在当前审批节点，无法查看该审批'}, status=403)
                return Response({'error': '您暂时没有权限查看该审批'}, status=403)
            data = ApprovalRequestSerializer(approval, context={'request': request}).data
            # 查询该审批类型是否需要手写签名：以审批所属企业为准（审批在哪家企业发起，
            # 就应用该企业的配置，子公司专属配置优先于集团默认配置）。
            try:
                tenant = approval.tenant or getattr(request, 'tenant', None) or request.user.get_active_tenant()
                config = self._get_config_for_tenant(tenant, approval.approval_type) if tenant else None
                data['require_signature'] = bool(config and config.require_signature)
                # 票据回传时限（小时，0=禁止回传；未配置时默认 24 小时）
                data['receipt_return_hours'] = int(config.receipt_return_hours) if config is not None and config.receipt_return_hours is not None else 24
                # 票据回传开关（默认关闭）
                data['enable_receipt_return'] = bool(config and config.enable_receipt_return)
            except Exception as e:
                logger.warning(f'解析审批签名配置失败: {e}')
                data['require_signature'] = False
                data['receipt_return_hours'] = 0
                data['enable_receipt_return'] = False

            # 收款信息可见性：仅超级管理员、申请人、最终审批节点的审批人可见（保护申请人隐私）
            try:
                is_final_node_approver = ApprovalAssignee.objects.filter(
                    node__request=approval, node__is_final_approver=True, user=request.user
                ).exists()
                data['can_view_payment'] = (
                    request.user.user_type == 'super_admin'
                    or approval.applicant_id == request.user.id
                    or is_final_node_approver
                )
            except Exception as e:
                logger.warning(f'计算收款信息可见性失败: {e}')
                data['can_view_payment'] = True  # 出错时默认可见，避免误隐藏

            # 标注最终审批人的配置来源（优先使用节点固化的来源；历史数据缺失时实时解析兜底）
            try:
                tenant = approval.tenant or getattr(request, 'tenant', None) or request.user.get_active_tenant()
                _, fa_source, fa_source_label = self._resolve_final_approver(tenant, approval.approval_type)
                if fa_source:
                    for n in (data.get('approval_nodes') or []):
                        if n.get('is_final_approver') and not n.get('final_approver_source'):
                            n['final_approver_source'] = fa_source
                            n['final_approver_source_label'] = fa_source_label
            except Exception as e:
                logger.warning(f'标注最终审批人配置来源失败: {e}')

            return Response({'encrypt': True, 'data': encrypt_data(data)})
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '该条审批不存在或者已经删除'}, status=404)

    @action(detail=True, methods=['post'])
    def send_private(self, request, pk=None):
        """审批发起人将待审批的审批以私聊卡片形式发送给指定审批人，对方点击卡片直达审批详情处理"""
        from chat.models import ChatRoom, Message
        user = request.user
        try:
            approval = ApprovalRequest.objects.select_related(
                'applicant'
            ).prefetch_related('approval_nodes__assignees').get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '该条审批不存在或者已经删除'}, status=404)

        if approval.applicant_id != user.id:
            return Response({'error': '只有审批发起人可以将审批私聊发送给审批人'}, status=403)
        if approval.status in ('approved', 'rejected', 'cancelled'):
            return Response({'error': '该审批已结束，无法再发送私聊提醒'}, status=400)

        target_id = request.data.get('assignee_user_id')
        if not target_id:
            return Response({'error': '缺少目标审批人'}, status=400)
        try:
            from accounts.models import CustomUser
            target_user = CustomUser.objects.get(id=target_id)
        except CustomUser.DoesNotExist:
            return Response({'error': '目标用户不存在'}, status=404)
        if target_user.id == user.id:
            return Response({'error': '不能发送给自己'}, status=400)

        # 仅当目标审批人仍处于待审批状态时允许发送；顺序审批时还需该审批人所在节点已到达
        pending_assignee = ApprovalAssignee.objects.filter(
            node__request=approval, user=target_user, status='pending'
        ).select_related('node').first()
        if not pending_assignee:
            return Response({'error': '该审批人已完成处理或不在待处理节点，无法发送'}, status=400)
        if approval.approval_mode == 'sequential' and approval.current_node_order:
            if pending_assignee.node.order > approval.current_node_order:
                return Response({'error': '该审批人所在节点尚未到达，暂不能发送私聊提醒'}, status=400)

        # get-or-create 私聊房间（与 ChatRoomViewSet.create 的私聊唯一性逻辑保持一致）
        room = ChatRoom.objects.filter(
            room_type='private', members=user
        ).filter(members__id=target_user.id).first()
        if not room:
            room = ChatRoom.objects.create(room_type='private', name='', creator=user)
            room.members.add(user, target_user)

        type_label = _approval_type_label(approval)
        card_data = {
            'approval_id': approval.id,
            'title': approval.title,
            'type_name': type_label,
            'applicant_name': user.real_name or user.username,
            'applicant_id': user.id,
            'amount': str(approval.amount) if approval.amount is not None else None,
            'status': approval.status,
            'status_display': approval.get_status_display(),
            'created_at': approval.created_at.isoformat() if approval.created_at else None,
            'sender_name': user.real_name or user.username,
        }
        preview_text = f"📨 {card_data['applicant_name']} 给您发送了一条审批: {approval.title}"

        message = Message.objects.create(
            chat_room=room,
            sender=user,
            content=json.dumps(card_data, ensure_ascii=False),
            message_type='approval_card'
        )

        # WebSocket 实时广播到该私聊房间
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f'chat_{room.id}',
            {
                'type': 'chat_message',
                'chat_room': room.id,
                'message_id': str(message.id),
                'sender': {
                    'id': user.id,
                    'username': user.username,
                    'real_name': user.real_name,
                    'avatar': user.avatar.url if getattr(user, 'avatar', None) else None,
                },
                'sender_id': user.id,
                'sender_name': user.real_name or user.username,
                'content': preview_text,
                'message_type': 'approval_card',
                'timestamp': message.timestamp.isoformat(),
                'approval_data': card_data,
            }
        )

        # 同步发一条工作通知，点击直达审批详情（复用已有跳转逻辑）
        send_work_notification(
            target_user.id,
            '审批私聊提醒',
            f'{user.real_name or user.username} 给您发送了一条审批「{approval.title}」，请及时处理',
            notification_type='approval',
            related_url=f'/oa/approval/?approval_id={approval.id}',
            extra_data={'approval_id': approval.id, 'action': 'private_remind'},
        )

        return Response({'encrypt': True, 'data': encrypt_data({'success': True, 'message': '已发送私聊提醒'})})

    def _get_receipt_return_hours(self, approval):
        """获取票据回传时限（小时，0=禁止回传；未配置默认 24 小时）"""
        try:
            tenant = approval.tenant or approval.applicant.get_active_tenant()
            config = self._get_config_for_tenant(tenant, approval.approval_type) if tenant else None
            if config is not None and config.receipt_return_hours is not None:
                return int(config.receipt_return_hours)
        except Exception as e:
            logger.warning(f'解析票据回传配置失败: {e}')
        return 24

    def _get_receipt_return_enabled(self, approval):
        """票据回传开关是否开启（配置 enable_receipt_return，默认关闭）"""
        try:
            tenant = approval.tenant or approval.applicant.get_active_tenant()
            config = self._get_config_for_tenant(tenant, approval.approval_type) if tenant else None
            if config is not None:
                return bool(config.enable_receipt_return)
        except Exception as e:
            logger.warning(f'解析票据回传开关配置失败: {e}')
        return False

    def _is_last_approver(self, approval, user):
        """是否为最后审批人（最终审批节点审批人；未标记最终节点时取最高 order 节点）"""
        final_node = approval.approval_nodes.filter(is_final_approver=True).first()
        if not final_node:
            final_node = approval.approval_nodes.order_by('-order').first()
        if not final_node:
            return False
        return final_node.assignees.filter(user=user).exists()

    def _last_approver_ids(self, approval):
        """最后审批节点的全部审批人 id"""
        final_node = approval.approval_nodes.filter(is_final_approver=True).first()
        if not final_node:
            final_node = approval.approval_nodes.order_by('-order').first()
        if not final_node:
            return []
        return list(final_node.assignees.values_list('user_id', flat=True))

    def _current_pending_approver_ids(self, approval):
        """当前需处理该审批的审批人 id（顺序=当前节点待审批；并行=所有待审批节点）"""
        if approval.approval_mode == 'sequential':
            node = approval.approval_nodes.filter(order=approval.current_node_order).first()
            if not node:
                return []
            return list(node.assignees.filter(status='pending').values_list('user_id', flat=True))
        return list(ApprovalAssignee.objects.filter(
            node__request=approval, status='pending'
        ).values_list('user_id', flat=True))

    @action(detail=True, methods=['post'])
    def upload_receipt(self, request, pk=None):
        """票据回传：审批发起人/最后审批人补传票据。
        发起人上传：审批进行中→通知当前节点审批人；审批已走完→通知最后审批人；未选择票据→不发送通知。
        最后审批人上传：发送工作通知给审批发起人。"""
        try:
            approval = ApprovalRequest.objects.select_related('applicant').prefetch_related(
                'approval_nodes__assignees').get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)
        is_applicant = approval.applicant_id == request.user.id
        is_last_approver = self._is_last_approver(approval, request.user)
        if not (is_applicant or is_last_approver):
            return Response({'error': '仅审批发起人或最后审批人可回传票据'}, status=403)
        if not self._get_receipt_return_enabled(approval):
            return Response({'error': '当前审批未开启票据回传'}, status=400)
        if approval.status not in ('pending', 'deferred', 'processing', 'approved'):
            return Response({'error': '当前审批状态不可回传票据'}, status=400)
        # 已通过时按配置的回传时限校验
        if approval.status == 'approved':
            hours = self._get_receipt_return_hours(approval)
            if hours <= 0:
                return Response({'error': '当前审批未开启票据回传'}, status=400)
            if not approval.receipt_deadline:
                approval.receipt_deadline = timezone.now() + timezone.timedelta(hours=hours)
                approval.save(update_fields=['receipt_deadline'])
            if timezone.now() > approval.receipt_deadline:
                return Response({'error': '已超过票据回传截止时间，无法回传'}, status=400)

        files = request.data.get('files') or request.data.get('receipts') or []
        if not isinstance(files, list):
            files = [files]
        receipts = list(approval.receipts or [])
        now_iso = timezone.now().isoformat()
        # 回传方角色：发起人回传=对审批人的反馈；最后审批人回传=对发起人的反馈
        uploader_role = 'applicant' if is_applicant else ('last_approver' if is_last_approver else 'other')
        added = 0
        added_files = []
        for f in files:
            if not isinstance(f, dict):
                continue
            url = (f.get('url') or '').strip()
            if not url:
                continue
            name = (f.get('name') or '').strip() or url.split('/')[-1]
            if any(r.get('url') == url for r in receipts):
                continue
            receipts.append({'url': url, 'name': name, 'uploaded_at': now_iso, 'uploader_role': uploader_role})
            added_files.append({'url': url, 'name': name})
            added += 1
        approval.receipts = receipts
        approval.save(update_fields=['receipts'])

        receipt_comment = (request.data.get('comment') or '').strip()
        # 既未选择票据也未填写审批意见 → 无操作，不发送通知
        if added == 0 and not receipt_comment:
            return Response({'encrypt': True, 'data': encrypt_data({
                'success': True, 'message': '未选择票据，未发送通知', 'receipts': receipts, 'added': 0})})

        # 记录审批日志：审批意见（可单独回传）+ 回传的票据文件显示在「审批记录」最下面
        try:
            ApprovalLog.objects.create(
                request=approval,
                operator=request.user,
                action='receipt_return',
                comment=receipt_comment or '回传了付款凭证/票据',
                attachments=added_files,
            )
        except Exception as e:
            logger.warning(f'记录票据回传日志失败: {e}')

        uploader_name = request.user.real_name or request.user.username
        related_url = f'/oa/approval/?approval_id={approval.id}'
        if added > 0:
            notify_act = 'receipt_uploaded'
            last_approver_notify = f'{uploader_name} 已回传 {added} 个票据/凭证，请及时查看'
            applicant_notify_done = f'{uploader_name} 已补传 {added} 个票据/凭证，请及时查看'
            applicant_notify_ing = f'{uploader_name} 已补传 {added} 个票据/凭证，请及时处理'
        else:
            notify_act = 'receipt_comment'
            last_approver_notify = f'{uploader_name} 提交了审批意见，请及时查看'
            applicant_notify_done = f'{uploader_name} 提交了审批意见，请及时查看'
            applicant_notify_ing = f'{uploader_name} 提交了审批意见，请及时处理'
        # 最后审批人回传/提交意见 → 通知审批发起人
        if is_last_approver and not is_applicant:
            send_work_notification(
                approval.applicant.id,
                '票据已回传',
                last_approver_notify,
                notification_type='approval',
                related_url=related_url,
                extra_data={'approval_id': approval.id, 'action': notify_act, 'by': 'last_approver'},
            )
        elif is_applicant:
            # 发起人回传/提交意见：审批已走完→通知最后审批人；进行中→通知当前节点审批人
            if approval.status == 'approved':
                targets = self._last_approver_ids(approval)
                content = applicant_notify_done
            else:
                targets = self._current_pending_approver_ids(approval)
                content = applicant_notify_ing
            for uid in set(targets):
                if uid == request.user.id:
                    continue
                send_work_notification(
                    uid,
                    '发起人已补传票据',
                    content,
                    notification_type='approval',
                    related_url=related_url,
                    extra_data={'approval_id': approval.id, 'action': notify_act, 'by': 'applicant'},
                )
        return Response({'encrypt': True, 'data': encrypt_data({
            'success': True, 'message': '回传成功', 'receipts': receipts, 'added': added})})

    @action(detail=True, methods=['post'])
    def delete_receipt(self, request, pk=None):
        """删除已回传的票据（仅回传本人可删除自己的票据）"""
        try:
            approval = ApprovalRequest.objects.get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)
        # 当前用户的回传角色：发起人删自己回传的；最后审批人删自己回传的
        user_role = 'applicant' if approval.applicant_id == request.user.id else ('last_approver' if self._is_last_approver(approval, request.user) else 'other')
        if user_role == 'other':
            return Response({'error': '您无权限删除回传票据'}, status=403)
        url = (request.data.get('url') or '').strip()
        if not url:
            return Response({'error': '缺少票据地址'}, status=400)
        receipts = list(approval.receipts or [])
        target = next((r for r in receipts if (r.get('url') or '') == url), None)
        if not target:
            return Response({'error': '未找到该回传票据'}, status=400)
        # 仅能删除本人上传的票据（防止删除对方的反馈票据）
        if (target.get('uploader_role') or 'applicant') != user_role:
            return Response({'error': '仅能删除本人回传的票据'}, status=403)
        new_receipts = [r for r in receipts if (r.get('url') or '') != url]
        approval.receipts = new_receipts
        approval.save(update_fields=['receipts'])
        # 同步从「审批记录」中对应的票据回传日志里移除该附件
        try:
            for log in ApprovalLog.objects.filter(request=approval, action='receipt_return'):
                atts = list(log.attachments or [])
                new_atts = [a for a in atts if (a.get('url') or '') != url]
                if len(new_atts) != len(atts):
                    log.attachments = new_atts
                    log.save(update_fields=['attachments'])
        except Exception as e:
            logger.warning(f'同步删除票据回传日志附件失败: {e}')
        return Response({'encrypt': True, 'data': encrypt_data({
            'success': True, 'message': '已删除', 'receipts': new_receipts})})

    @staticmethod
    def _custom_pm_to_dict(m):
        return {
            'id': m.id, 'type': 'custom',
            'payee_name': m.payee_name, 'bank_card': m.bank_card, 'bank_name': m.bank_name,
            'bank_address': m.bank_address, 'alipay_account': m.alipay_account,
            'wechat_account': m.wechat_account, 'alipay_qr': m.alipay_qr, 'wechat_qr': m.wechat_qr,
        }

    @action(detail=False, methods=['get', 'post'])
    def custom_payment_methods(self, request):
        """获取/保存用户自定义收款方式库（记忆功能）"""
        from oa.models import CustomPaymentMethod
        if request.method == 'GET':
            methods = CustomPaymentMethod.objects.filter(user=request.user).order_by('-updated_at')
            data = [self._custom_pm_to_dict(m) for m in methods]
            return Response({'results': data})
        # POST：保存一条自定义收款方式（相同内容去重）
        pm = request.data.get('payment_method') or request.data
        if not isinstance(pm, dict):
            pm = {}
        payee = (pm.get('payee_name') or '').strip()
        if not payee:
            return Response({'error': '请填写收款人姓名'}, status=400)
        m, _ = CustomPaymentMethod.objects.get_or_create(
            user=request.user,
            payee_name=payee,
            bank_card=(pm.get('bank_card') or '').strip(),
            bank_name=(pm.get('bank_name') or '').strip(),
            bank_address=(pm.get('bank_address') or '').strip(),
            alipay_account=(pm.get('alipay_account') or '').strip(),
            wechat_account=(pm.get('wechat_account') or '').strip(),
            alipay_qr=pm.get('alipay_qr') or '',
            wechat_qr=pm.get('wechat_qr') or '',
        )
        return Response(self._custom_pm_to_dict(m))

    @action(detail=True, methods=['delete'])
    def custom_payment_method(self, request, pk=None):
        """删除用户的一条自定义收款方式"""
        from oa.models import CustomPaymentMethod
        try:
            m = CustomPaymentMethod.objects.get(id=pk, user=request.user)
        except CustomPaymentMethod.DoesNotExist:
            return Response({'error': '收款方式不存在'}, status=404)
        m.delete()
        return Response({'message': 'ok'})

    def _resolve_payment_method(self, user, payment_method, approval_type, type_obj=None):
        """解析审批付款方式（报销/采购/自定义类型，收款方式可选）。

        default：使用用户默认收款账号（CustomUser，提交时快照）；未设置则返回需要完善收款信息的提示。
        custom：用户自定义收款方式（需收款人姓名 + 至少一种收款方式），并记忆到用户收款方式库。
        未选择收款方式时返回 ({}, None)，审批详情不展示收款信息。
        返回 (payment_method_dict, error_response)；非付款类型返回 ({}, None)。
        """
        if not isinstance(payment_method, dict):
            payment_method = {}
        need_payment = approval_type in ('expense', 'purchase')
        if not need_payment and type_obj and not type_obj.is_builtin:
            # 自定义类型：仅当 schema 配置了「收款方式」字段时才需要
            need_payment = any(f.get('type') == 'payment_method' for f in (type_obj.form_schema or []))
        if not need_payment:
            return {}, None
        pm_type = (payment_method.get('type') or '').strip()
        # 收款方式可选：未选择则不保存收款方式
        if not pm_type:
            return {}, None
        if pm_type == 'custom':
            payee = (payment_method.get('payee_name') or '').strip()
            has_any = any((payment_method.get(k) or '').strip()
                          for k in ('bank_card', 'alipay_account', 'wechat_account'))
            if not payee or not has_any:
                return None, Response({'error': '请填写自定义收款方式（收款人姓名 + 至少一种收款方式）'}, status=400)
            result = {
                'type': 'custom', 'payee_name': payee,
                'bank_card': (payment_method.get('bank_card') or '').strip(),
                'bank_name': (payment_method.get('bank_name') or '').strip(),
                'bank_address': (payment_method.get('bank_address') or '').strip(),
                'alipay_account': (payment_method.get('alipay_account') or '').strip(),
                'wechat_account': (payment_method.get('wechat_account') or '').strip(),
                'alipay_qr': payment_method.get('alipay_qr') or '',
                'wechat_qr': payment_method.get('wechat_qr') or '',
            }
            # 记忆自定义收款方式：保存到用户收款方式库，供下次选择复用（相同内容去重）
            try:
                from oa.models import CustomPaymentMethod
                CustomPaymentMethod.objects.get_or_create(
                    user=user,
                    payee_name=result['payee_name'],
                    bank_card=result['bank_card'],
                    bank_name=result['bank_name'],
                    bank_address=result['bank_address'],
                    alipay_account=result['alipay_account'],
                    wechat_account=result['wechat_account'],
                    alipay_qr=result['alipay_qr'],
                    wechat_qr=result['wechat_qr'],
                )
            except Exception as e:
                logger.warning(f'保存自定义收款方式失败: {e}')
            return result, None
        # 默认：使用用户默认收款账号；未设置则提示先完善
        if not SubsidyViewSet._user_has_payment_info(user):
            return None, Response({'error': '您还未填写收款账号，请先完善收款信息',
                                   'need_payment_info': True}, status=400)
        return self._snapshot_default_payment(user), None

    def _snapshot_default_payment(self, user):
        """快照用户默认收款账号（CustomUser 收款信息，含开户行与收款码）"""
        return {
            'type': 'default',
            'payee_name': user.payee_name or '',
            'bank_card': user.bank_card or '',
            'bank_name': user.bank_name or '',
            'bank_address': user.bank_address or '',
            'alipay_account': user.alipay_account or '',
            'wechat_account': user.wechat_account or '',
            'alipay_qr': user.alipay_qr or '',
            'wechat_qr': user.wechat_qr or '',
        }

    def _extract_payment_from_form_data(self, form_data, type_obj):
        """自定义类型：若 schema 配置了「收款方式」字段，从 form_data 提取其值并从 form_data 移除。
        返回 (payment_method_dict, 是否已提取)。"""
        if not type_obj or type_obj.is_builtin or not isinstance(form_data, dict):
            return {}, False
        pm_field = next((f for f in (type_obj.form_schema or []) if f.get('type') == 'payment_method'), None)
        if not pm_field:
            return {}, False
        key = pm_field.get('key')
        pm = form_data.get(key)
        if isinstance(pm, dict):
            form_data.pop(key, None)
            return pm, True
        return {}, False

    def _resolve_payment_method_for_draft(self, user, payment_method, approval_type):
        """草稿用的软解析：未选择→({})；默认方式→快照用户收款账号（无收款信息也不报错，字段为空）；
        自定义方式→原样存储（不强制校验）。保证草稿详情也能展示收款信息。"""
        if not isinstance(payment_method, dict):
            payment_method = {}
        pm_type = (payment_method.get('type') or '').strip()
        # 收款方式可选：未选择则不保存收款方式
        if not pm_type:
            return {}, None
        if pm_type == 'custom':
            return {
                'type': 'custom',
                'payee_name': (payment_method.get('payee_name') or '').strip(),
                'bank_card': (payment_method.get('bank_card') or '').strip(),
                'bank_name': (payment_method.get('bank_name') or '').strip(),
                'bank_address': (payment_method.get('bank_address') or '').strip(),
                'alipay_account': (payment_method.get('alipay_account') or '').strip(),
                'wechat_account': (payment_method.get('wechat_account') or '').strip(),
                'alipay_qr': payment_method.get('alipay_qr') or '',
                'wechat_qr': payment_method.get('wechat_qr') or '',
            }
        return self._snapshot_default_payment(user)

    def create(self, request):
        serializer = ApprovalCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        approval_type = serializer.validated_data['approval_type']
        department_id = serializer.validated_data.get('department_id')
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()

        # 统一 form_data：内置类型镜像 legacy 字段，自定义类型按 schema 校验
        form_data = collect_form_data(
            serializer.validated_data,
            serializer.validated_data.get('form_data') or {},
            approval_type,
        )
        type_obj = resolve_approval_type(approval_type, tenant)
        # 类型存在但已停用时拒绝发起审批（内置/自定义均适用）
        if type_obj is None:
            tq = ApprovalType.objects.filter(code=approval_type)
            if tenant:
                existing = tq.filter(tenant=tenant).first() or tq.filter(tenant__isnull=True).first()
            else:
                existing = tq.filter(tenant__isnull=True).first()
            if existing and not existing.enabled:
                return Response({'error': '该审批类型已停用，无法发起审批'}, status=400)
        if type_obj and not type_obj.is_builtin:
            verr = validate_form_data(type_obj.form_schema or [], form_data)
            if verr:
                return Response({'error': '表单校验失败', 'fields': verr}, status=400)

        # 审批人自动根据所选部门和审批类型生成
        approver_nodes = serializer.validated_data.get('approver_nodes', [])
        if not approver_nodes:
            threshold_values = self._gather_threshold_values(serializer.validated_data, form_data)
            approver_nodes = self._auto_determine_approvers(
                request.user, tenant,
                approval_type=approval_type,
                department_id=department_id,
                threshold_values=threshold_values or None,
            )

        # 所属部门
        from accounts.models import Department
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                department = request.user.department
        else:
            department = request.user.department

        # 内置类型结构化数据
        purchase_items = serializer.validated_data.get('purchase_items', [])
        expense_items = serializer.validated_data.get('expense_items', [])
        leave_type = serializer.validated_data.get('leave_type', '')
        trip_data = serializer.validated_data.get('trip_data', {}) or {}
        # 根据物项/项目自动计算总金额
        amount = serializer.validated_data.get('amount')
        auto_amount = None
        if approval_type == 'purchase' and purchase_items:
            try:
                auto_amount = sum(float(i.get('total') or 0) for i in purchase_items)
            except (ValueError, TypeError):
                auto_amount = None
        elif approval_type == 'expense' and expense_items:
            try:
                auto_amount = sum(float(i.get('amount') or 0) for i in expense_items)
            except (ValueError, TypeError):
                auto_amount = None
        elif approval_type == 'trip' and trip_data.get('amount'):
            try:
                auto_amount = float(trip_data.get('amount'))
            except (ValueError, TypeError):
                auto_amount = None
        if auto_amount is not None:
            amount = auto_amount
        # 出差天数
        duration = serializer.validated_data.get('duration')
        if approval_type == 'trip' and trip_data.get('days'):
            try:
                duration = float(trip_data.get('days'))
            except (ValueError, TypeError):
                pass

        # 收款方式：报销/采购用 serializer.payment_method；自定义类型从 form_data 的「收款方式」字段提取
        payment_method = serializer.validated_data.get('payment_method') or {}
        if type_obj and not type_obj.is_builtin:
            fd_pm, _ = self._extract_payment_from_form_data(form_data, type_obj)
            if fd_pm:
                payment_method = fd_pm
        payment_method, pm_err = self._resolve_payment_method(
            request.user, payment_method, approval_type, type_obj)
        if pm_err:
            return pm_err

        with transaction.atomic():
            approval = ApprovalRequest.objects.create(
                applicant=request.user,
                tenant=getattr(request, 'tenant', None) or request.user.get_active_tenant(),
                department=department,
                approval_type=approval_type,
                title=serializer.validated_data['title'],
                content=serializer.validated_data.get('content', ''),
                start_date=serializer.validated_data.get('start_date'),
                end_date=serializer.validated_data.get('end_date'),
                duration=duration,
                amount=amount,
                expense_type=serializer.validated_data.get('expense_type', ''),
                expense_date=serializer.validated_data.get('expense_date'),
                attachments=serializer.validated_data.get('attachments', []),
                recruit_data=serializer.validated_data.get('recruit_data', {}),
                purchase_items=purchase_items,
                expense_items=expense_items,
                leave_type=leave_type,
                trip_data=trip_data,
                form_data=form_data,
                sign_type=serializer.validated_data.get('sign_type', 'countersign'),
                approval_mode=serializer.validated_data.get('approval_mode', 'sequential'),
                payment_method=payment_method,
            )
            # 物资单据：创建业务记录 + 自动生成单据号（失败则回滚整单）；金额镜像到审批金额字段以支持阈值审批
            if approval_type in ('material_requirement', 'material_requisition'):
                from decimal import Decimal as _D, InvalidOperation as _IO
                from .material_utils import ensure_material_record
                _mrec, merr = ensure_material_record(approval, form_data)
                if merr:
                    raise serializers.ValidationError(merr)
                approval.form_data = form_data
                if form_data.get('amount'):
                    try:
                        approval.amount = _D(str(form_data['amount']))
                    except (_IO, ValueError, TypeError):
                        pass
                approval.save(update_fields=['form_data', 'amount'] if form_data.get('amount') else ['form_data'])
            from accounts.models import CustomUser, Department

            # 创建发起人节点（order 0，展示用）
            ApprovalNode.objects.create(
                request=approval, node_type='initiator',
                user=request.user, order=0,
            )

            if approver_nodes:
                for idx, node_data in enumerate(approver_nodes):
                    node_type = node_data.get('type', 'user')
                    node = ApprovalNode.objects.create(
                        request=approval,
                        node_type=node_type,
                        order=idx + 1,
                    )
                    if node_type == 'user':
                        user_id = node_data.get('id')
                        try:
                            u = CustomUser.objects.get(id=user_id)
                            node.user = u
                            node.save()
                            ApprovalAssignee.objects.create(node=node, user=u)
                        except CustomUser.DoesNotExist:
                            pass
                    elif node_type == 'department':
                        dept_id = node_data.get('id')
                        try:
                            dept = Department.objects.get(id=dept_id)
                            node.department = dept
                            node.save()
                            # 获取部门下所有管理员/超级管理员
                            dept_admins = CustomUser.objects.filter(
                                department=dept,
                                user_type__in=['super_admin', 'admin']
                            ).exclude(id=request.user.id)
                            for au in dept_admins:
                                ApprovalAssignee.objects.create(node=node, user=au)
                        except Department.DoesNotExist:
                            pass
            else:
                # 默认：自动分配第一个管理员
                admin = CustomUser.objects.filter(
                    Q(user_type='super_admin') | Q(user_type='admin')
                ).exclude(id=request.user.id).first()
                if admin:
                    node = ApprovalNode.objects.create(
                        request=approval, node_type='user',
                        user=admin, order=1,
                    )
                    ApprovalAssignee.objects.create(node=node, user=admin)

            # 若配置了最终审批人，标记对应节点
            self._mark_final_approver_node(approval, request)

            # 如果是顺序审批，只激活第一个审批人节点（order=1跳过发起人节点）
            if approval.approval_mode == 'sequential':
                approval.current_node_order = 1
                approval.save()

        # 通知审批人（顺序审批仅通知第一个节点）
        from accounts.models import CustomUser
        assignee_qs = ApprovalAssignee.objects.filter(node__request=approval, status='pending')
        if approval.approval_mode == 'sequential':
            first_node = approval.approval_nodes.filter(order=1).first()
            assignee_qs = assignee_qs.filter(node=first_node) if first_node else assignee_qs.none()
        for asgn in assignee_qs.select_related('user'):
            send_work_notification(
                user_id=asgn.user.id,
                title='审批待处理',
                content=f'{request.user.real_name or request.user.username} 提交了{_approval_type_label(approval)}申请：“{approval.title}”',
                notification_type='approval',
                related_url=f'/oa/approval/?approval_id={approval.id}',
                extra_data={'approval_id': approval.id, 'action': 'pending'},
            )

        # 处理抄送（用户+部门）
        cc_user_ids = serializer.validated_data.get('cc_users', [])
        cc_dept_ids = serializer.validated_data.get('cc_departments', [])
        self._create_cc_records(approval, request, cc_user_ids, cc_dept_ids)

        # 关联审批（自关联，如报销关联采购）
        related_ids = serializer.validated_data.get('related_approvals', [])
        if related_ids:
            approval.related_approvals.set(related_ids)

        # 加载默认配置中的抄送
        if request.tenant and approval.approval_type:
            try:
                config = ApprovalDeptConfig.objects.get(tenant=request.tenant, approval_type=approval.approval_type, sub_tenant__isnull=True)
                if config.cc_users:
                    self._create_cc_records(approval, request, config.cc_users, [])
                if config.cc_departments:
                    self._create_cc_records(approval, request, [], config.cc_departments)
            except ApprovalDeptConfig.DoesNotExist:
                pass

        logger.info(f'{request.user} 提交审批 {approval.title}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201)

    def _get_level_approvers(self, dept, user, seen_ids):
        """获取部门级别的审批人：先副负责人，后主负责人

        部门未设置负责人/副负责人时返回空，由审批链向上回溯到最近的
        有负责人的上级部门（即该部门的有效负责人），不再回退到普通成员，
        避免把与申请人同级别的部门成员误选为审批人。
        """
        result = []
        # Deputy managers first (from Department.deputy_managers M2M)
        for dm in dept.deputy_managers.all():
            if dm.id not in seen_ids and dm.id != user.id and dm.is_active:
                seen_ids.add(dm.id)
                result.append({
                    'type': 'user', 'id': dm.id,
                    'label': f'{dm.real_name or dm.username}（副负责人）',
                    'user_position': dm.position or '副负责人',
                })
        # Primary manager (from Department.manager FK)
        if dept.manager and dept.manager.id not in seen_ids and dept.manager.id != user.id and dept.manager.is_active:
            seen_ids.add(dept.manager.id)
            result.append({
                'type': 'user', 'id': dept.manager.id,
                'label': f'{dept.manager.real_name or dept.manager.username}（负责人）',
                'user_position': dept.manager.position or '负责人',
            })

        return result

    def _get_parent_tenant(self, tenant):
        """获取上级集团企业"""
        if not tenant:
            return None
        return tenant.parent if tenant.parent else None

    def _get_root_tenant(self, tenant):
        """获取根集团企业"""
        if not tenant:
            return None
        return tenant.get_root_tenant()

    def _get_final_approval_dept(self, tenant, approval_type, traverse_parent=True):
        """获取指定审批类型的最终审批部门（支持向上级集团查找）"""
        if not tenant or not approval_type:
            return None
        # 当前企业配置
        try:
            config = ApprovalDeptConfig.objects.get(tenant=tenant, approval_type=approval_type, sub_tenant__isnull=True)
            return config.department
        except ApprovalDeptConfig.DoesNotExist:
            pass
        # 上级集团配置兜底
        if traverse_parent:
            parent = self._get_parent_tenant(tenant)
            if parent:
                return self._get_final_approval_dept(parent, approval_type, traverse_parent=True)
        return None

    def _final_approver_source_label(self, config):
        """最终审批人配置来源展示文案"""
        if config and config.sub_tenant_id:
            return '子公司自定义配置'
        try:
            if config and config.tenant and config.tenant.sub_tenants.exists():
                return '集团默认配置'
        except Exception:
            pass
        return '默认配置'

    def _resolve_final_approver(self, tenant, approval_type):
        """解析配置的最终审批人及其来源（允许置空）

        返回 (user, source, source_label)：
        source ∈ ('sub' 子公司自定义配置, 'default' 集团/企业默认配置, None)
        优先级：集团子公司自定义配置 > 集团默认配置；
        子公司配置存在但未设置最终审批人时，回退到集团默认配置的最终审批人。
        """
        if not tenant or not approval_type:
            return None, None, ''
        config = self._get_config_for_tenant(tenant, approval_type)
        if not config:
            return None, None, ''
        if config.final_approver_id:
            source = 'sub' if config.sub_tenant_id else 'default'
            return config.final_approver, source, self._final_approver_source_label(config)
        # 子公司专属配置未设置最终审批人：回退到集团默认配置
        if config.sub_tenant_id:
            default_config = ApprovalDeptConfig.objects.filter(
                tenant=config.tenant_id,
                approval_type=approval_type,
                sub_tenant__isnull=True
            ).order_by('-updated_at').first()
            if default_config and default_config.final_approver_id:
                return default_config.final_approver, 'default', self._final_approver_source_label(default_config)
        return None, None, ''

    def _get_final_approver(self, tenant, approval_type):
        """获取配置的最终审批人（允许置空）"""
        user, _, _ = self._resolve_final_approver(tenant, approval_type)
        return user

    def _mark_final_approver_node(self, approval, request):
        """若该审批类型配置了最终审批人，则将对应审批节点标记为最终审批人（供前端流程标识），并固化配置来源"""
        try:
            tenant = approval.tenant or getattr(request, 'tenant', None) or request.user.get_active_tenant()
            final_user, final_source, _ = self._resolve_final_approver(tenant, approval.approval_type)
            if not final_user:
                return
            match = approval.approval_nodes.filter(
                user=final_user, node_type='user'
            ).order_by('-order').first()
            if match:
                match.is_final_approver = True
                match.final_approver_source = final_source or ''
                match.save(update_fields=['is_final_approver', 'final_approver_source'])
        except Exception as e:
            logger.warning(f'标记最终审批人节点失败: {e}')

    def _get_config_for_tenant(self, tenant, approval_type, traverse_parent=True):
        """获取审批配置，支持子企业专属配置和向父级集团回溯

        优先级：
        1. 子企业专属配置（集团配置中 sub_tenant=当前企业）
        2. 当前企业的默认配置（或集团的默认配置）
        3. 向父级集团回溯查找
        """
        if not tenant or not approval_type:
            return None
        # 确定配置所属的 tenant（子企业配置存在集团上，公司配置存在自己上）
        config_tenant = tenant.parent if tenant.parent else tenant

        # 第1优先：子企业专属配置（集团配置中 sub_tenant=当前企业）
        sub_config = ApprovalDeptConfig.objects.filter(
            tenant=config_tenant,
            approval_type=approval_type,
            sub_tenant=tenant if tenant.parent else None,
        ).order_by('-updated_at').first()
        if sub_config:
            return sub_config

        # 第2优先：当前 tenant 的默认配置（无 sub_tenant）
        default_config = ApprovalDeptConfig.objects.filter(
            tenant=config_tenant,
            approval_type=approval_type,
            sub_tenant__isnull=True
        ).order_by('-updated_at').first()
        if default_config:
            return default_config

        # 第3优先：向父级集团回溯
        if traverse_parent:
            parent = self._get_parent_tenant(tenant)
            if parent:
                return self._get_config_for_tenant(parent, approval_type, traverse_parent=True)

        return None

    def _fallback_approvers(self, user, tenant):
        """兜底：查找企业管理员（支持向父级集团回溯）"""
        from accounts.models import CustomUser
        from django.db.models import Q

        def _find_admins(t):
            if not t:
                return []
            qs = CustomUser.objects.filter(
                Q(user_type='super_admin') | Q(user_type='admin')
            ).exclude(id=user.id)
            qs = qs.filter(
                tenant_memberships__tenant=t,
                tenant_memberships__is_active=True
            )
            return list(qs.distinct()[:3])

        result = []
        seen = set()
        # 当前企业
        admins = _find_admins(tenant)
        for admin in admins:
            if admin.id not in seen:
                seen.add(admin.id)
                result.append({
                    'type': 'user', 'id': admin.id,
                    'label': admin.real_name or admin.username,
                    'user_position': admin.position or '',
                })
        # 父级集团兜底
        if not result:
            parent = self._get_parent_tenant(tenant)
            if parent:
                parent_admins = _find_admins(parent)
                for admin in parent_admins:
                    if admin.id not in seen:
                        seen.add(admin.id)
                        result.append({
                            'type': 'user', 'id': admin.id,
                            'label': f'{admin.real_name or admin.username}（集团）',
                            'user_position': admin.position or '',
                        })
        return result

    def _gather_threshold_values(self, validated_data, form_data=None):
        """收集用于阈值判断的数字字段：form_data（自定义/镜像）+ legacy（内置兜底）"""
        threshold_values = {}
        fd = form_data or {}
        for k, v in fd.items():
            if isinstance(v, bool):
                continue
            if isinstance(v, (int, float)):
                threshold_values[k] = float(v)
            elif isinstance(v, str) and v.strip():
                # 动态表单数字/金额字段收集为字符串（如 "1234.56"），也纳入阈值判断
                try:
                    threshold_values[k] = float(v)
                except (ValueError, TypeError):
                    pass
        duration = validated_data.get('duration')
        if duration:
            try:
                threshold_values['duration'] = float(duration)
            except (ValueError, TypeError):
                pass
        amount = validated_data.get('amount')
        if amount:
            try:
                threshold_values['amount'] = float(amount)
            except (ValueError, TypeError):
                pass
        recruit_data = validated_data.get('recruit_data', {})
        if recruit_data and isinstance(recruit_data, dict):
            hc = recruit_data.get('headcount', 0)
            if hc:
                try:
                    threshold_values['headcount'] = float(hc)
                except (ValueError, TypeError):
                    pass
        return threshold_values if threshold_values else None

    def _create_cc_records(self, approval, request, user_ids=None, dept_ids=None):
        """创建抄送记录（支持抄送用户和部门，幂等去重）。
        注意：仅创建记录，不在创建时发送抄送通知；抄送人需在审批结束后才收到抄送通知（见 _notify_cc_after_end）。"""
        user_ids = user_ids or []
        dept_ids = dept_ids or []
        from accounts.models import CustomUser, Department

        # 抄送用户（get_or_create 去重）
        seen_users = set()
        if user_ids:
            cc_users = CustomUser.objects.filter(id__in=user_ids, is_active=True).exclude(id=request.user.id)
            for cc_user in cc_users:
                if cc_user.id in seen_users:
                    continue
                seen_users.add(cc_user.id)
                ApprovalCarbonCopy.objects.get_or_create(
                    request=approval, cc_type='user', cc_user=cc_user,
                    defaults={'cc_department': None},
                )

        # 抄送部门（get_or_create 去重）
        seen_depts = set()
        if dept_ids:
            depts = Department.objects.filter(id__in=dept_ids, is_active=True)
            for dept in depts:
                if dept.id in seen_depts:
                    continue
                seen_depts.add(dept.id)
                ApprovalCarbonCopy.objects.get_or_create(
                    request=approval, cc_type='department', cc_department=dept,
                    defaults={'cc_user': None},
                )

    def _notify_cc_after_end(self, approval):
        """审批结束后（已通过/已驳回）才向抄送人发送抄送通知"""
        applicant_name = approval.applicant.real_name or approval.applicant.username if approval.applicant else ''
        status_label = approval.get_status_display()
        for cc in approval.carbon_copies.all():
            if cc.cc_type == 'user' and cc.cc_user and cc.cc_user.id != approval.applicant_id:
                send_work_notification(
                    user_id=cc.cc_user.id,
                    title='审批抄送通知',
                    content=f'{applicant_name} 的{_approval_type_label(approval)}申请“{approval.title}”已完成（{status_label}），请知悉',
                    notification_type='approval',
                    related_url=f'/oa/approval/?approval_id={approval.id}',
                    extra_data={'approval_id': approval.id, 'action': 'cc'},
                )
            elif cc.cc_type == 'department' and cc.cc_department:
                targets = []
                if cc.cc_department.manager and cc.cc_department.manager.id != approval.applicant_id:
                    targets.append(cc.cc_department.manager)
                for dm in cc.cc_department.deputy_managers.all():
                    if dm.id != approval.applicant_id:
                        targets.append(dm)
                for mgr in targets:
                    send_work_notification(
                        user_id=mgr.id,
                        title='审批抄送通知',
                        content=f'{applicant_name} 的{_approval_type_label(approval)}申请“{approval.title}”已完成（{status_label}，抄送部门：{cc.cc_department.name}），请知悉',
                        notification_type='approval',
                        related_url=f'/oa/approval/?approval_id={approval.id}',
                        extra_data={'approval_id': approval.id, 'action': 'cc'},
                    )

    def _auto_determine_approvers(self, user, tenant, approval_type=None, department_id=None,
                                   threshold_values=None):
        """根据所选部门自动确定审批人链（三级审批 + 阈值超额审批）

        规则：
        1. 第一级：申请人所选部门的负责人（先副负责人，后主负责人）
        2. 第二级：申请人所选部门的上一级部门负责人
        3. 第三级：超级管理员设置的指定最终审批部门负责人
        4. 若配置了阈值且超过阈值，则增加一级阈值超额审批部门
        如果某一级没有负责人，则跳过该级到下一级

        threshold_values: dict with keys 'duration', 'amount', 'headcount'
        """
        from accounts.models import Department, CustomUser
        from org.models import UserDepartment

        if not department_id:
            return self._fallback_approvers(user, tenant)

        # Load config for this type (支持向父级集团回溯查找配置)
        config = self._get_config_for_tenant(tenant, approval_type)

        try:
            selected_dept = Department.objects.get(id=department_id, tenant=tenant)
        except Department.DoesNotExist:
            logger.error(f'Department not found for id {department_id} and tenant {tenant}')
            return self._fallback_approvers(user, tenant)

        all_approvers = []
        seen_ids = set()

        # Level 1: Selected department's heads (deputy first, then manager)
        level1 = self._get_level_approvers(selected_dept, user, seen_ids)
        all_approvers.extend(level1)

        # Level 2: Parent department's heads (only if level 1 has someone)
        if selected_dept.parent:
            level2 = self._get_level_approvers(selected_dept.parent, user, seen_ids)
            all_approvers.extend(level2)
        elif not level1:
            # Selected dept has no approvers and no parent, try siblings or tenant admin
            pass

        # If still no approvers after checking selected + parent, walk up chain
        if not all_approvers:
            current = selected_dept.parent
            while current:
                level = self._get_level_approvers(current, user, seen_ids)
                if level:
                    all_approvers.extend(level)
                    break
                current = current.parent

        # Level 3: Use configured approver_users if set, otherwise use final approval dept's manager/deputy
        if config and config.approver_users:
            cfg_users = CustomUser.objects.filter(id__in=config.approver_users, is_active=True)
            for u in cfg_users:
                if u.id not in seen_ids and u.id != user.id:
                    seen_ids.add(u.id)
                    all_approvers.append({
                        'type': 'user', 'id': u.id,
                        'label': f'{u.real_name or u.username}',
                        'user_position': u.position or '',
                    })
        else:
            final_dept = self._get_final_approval_dept(tenant, approval_type)
            if final_dept:
                level3 = self._get_level_approvers(final_dept, user, seen_ids)
                all_approvers.extend(level3)

        if not all_approvers:
            # Final attempt: try final approval dept directly
            if not config or not config.approver_users:
                final_dept = self._get_final_approval_dept(tenant, approval_type)
                if final_dept:
                    level3 = self._get_level_approvers(final_dept, user, seen_ids)
                    all_approvers.extend(level3)

        # Level 4: Threshold exceeded — check sub-tenant config first, then group config
        threshold_dept = None
        if config and config.threshold_enabled and config.threshold_value is not None:
            threshold_val = None
            if threshold_values:
                tv_field = config.threshold_field
                if tv_field in threshold_values:
                    threshold_val = threshold_values[tv_field]
            if threshold_val is not None:
                try:
                    threshold_val = float(threshold_val)
                except (ValueError, TypeError):
                    threshold_val = None
            if threshold_val is not None and threshold_val > config.threshold_value:
                threshold_dept = config.threshold_department
        # 如果子企业配置没有启用阈值，尝试集团默认配置
        if not threshold_dept and config and config.approver_users is not None:
            try:
                fallback_config = ApprovalDeptConfig.objects.get(
                    tenant=tenant if not tenant.parent else tenant.parent,
                    approval_type=approval_type,
                    sub_tenant__isnull=True
                )
                if fallback_config and fallback_config.threshold_enabled and fallback_config.threshold_value is not None:
                    threshold_val = None
                    if threshold_values:
                        tv_field = fallback_config.threshold_field
                        if tv_field in threshold_values:
                            threshold_val = threshold_values[tv_field]
                    if threshold_val is not None:
                        try:
                            threshold_val = float(threshold_val)
                        except (ValueError, TypeError):
                            threshold_val = None
                    if threshold_val is not None and threshold_val > fallback_config.threshold_value:
                        threshold_dept = fallback_config.threshold_department
            except ApprovalDeptConfig.DoesNotExist:
                pass
        if threshold_dept:
            level4 = self._get_level_approvers(threshold_dept, user, seen_ids)
            if level4:
                for l in level4:
                    l['_threshold_level'] = True
                all_approvers.extend(level4)

        # Level 5: 最终审批人（配置里指定，追加到审批链最后一级；子公司配置 > 集团默认配置）
        final_approver, final_source, final_source_label = self._resolve_final_approver(tenant, approval_type)
        if final_approver:
            if final_approver.id not in seen_ids and final_approver.id != user.id and final_approver.is_active:
                seen_ids.add(final_approver.id)
                all_approvers.append({
                    'type': 'user', 'id': final_approver.id,
                    'label': f'{final_approver.real_name or final_approver.username}',
                    'user_position': final_approver.position or '',
                    '_final_approver': True,
                    '_final_approver_source': final_source,
                    '_final_approver_source_label': final_source_label,
                })

        if not all_approvers:
            return self._fallback_approvers(user, tenant)

        return all_approvers

    def _check_user_can_approve(self, approval, user):
        """检查用户是否有权限审批当前节点（待审批/暂缓/办理中均可操作）"""
        if approval.approval_mode == 'sequential':
            nodes = approval.approval_nodes.filter(order=approval.current_node_order)
        else:
            nodes = approval.approval_nodes.all()
        return ApprovalAssignee.objects.filter(
            node__in=nodes, user=user, status__in=('pending', 'deferred', 'processing')
        ).exists()

    def _check_user_can_view(self, approval, user):
        """检查用户是否有权限查看审批（不限制审批状态）"""
        if approval.approval_mode == 'sequential':
            nodes = approval.approval_nodes.filter(order=approval.current_node_order)
        else:
            nodes = approval.approval_nodes.all()
        return ApprovalAssignee.objects.filter(
            node__in=nodes, user=user
        ).exists()

    def _is_approver_reached(self, approval, user):
        """审批人可查看规则：
        - 已通过：曾为该审批审批人的均可查看；
        - 已驳回：仅驳回该审批的审批人可查看；
        - 已撤回：审批到达自己节点时被发起人撤回的审批人可查看；
        - 进行中：已到达自己节点（并行全部节点、顺序仅当前及已到达节点）。
        """
        assignee = ApprovalAssignee.objects.filter(
            node__request=approval, user=user
        ).select_related('node').first()
        if not assignee:
            return False
        if approval.approval_mode == 'parallel':
            return True
        if approval.status == 'approved':
            return True
        if approval.status == 'rejected':
            # 仅驳回该审批的审批人可查看
            return assignee.status == 'rejected'
        if approval.status == 'cancelled':
            # 审批到达自己节点时被发起人撤回
            return assignee.node.order <= (approval.current_node_order or 0)
        # 进行中（pending/deferred/processing）：已到达自己节点
        return assignee.node.order <= (approval.current_node_order or 0)

    def _check_user_processed(self, approval, user):
        """检查用户是否已作为审批人通过/驳回过该审批（此类审批人可查看）"""
        return ApprovalAssignee.objects.filter(
            node__request=approval, user=user, status__in=['approved', 'rejected']
        ).exists()

    def _process_node_approval(self, approval, node, action, user, comment, attachments=None, signature=''):
        """处理单个节点的审批（支持通过/驳回/暂缓/办理中，可附带手写签名）"""
        now = timezone.now()
        assignees = node.assignees.filter(user=user)
        if not assignees.exists():
            return False
        assignee = assignees.first()
        # 可操作状态：pending(待审批), deferred(暂缓), processing(办理中)
        if assignee.status not in ('pending', 'deferred', 'processing'):
            return False
        # 只有通过/驳回真正改变节点完成状态
        if action in ('approve', 'reject'):
            assignee.status = 'approved' if action == 'approve' else 'rejected'
        elif action == 'deferred':
            assignee.status = 'deferred'
        elif action == 'processing':
            assignee.status = 'processing'
        else:
            return False
        assignee.comment = comment or ''
        assignee.operated_at = now
        assignee.save()

        # 记录日志（附带手写签名）
        ApprovalLog.objects.create(
            request=approval, operator=user,
            action=action, comment=comment or '',
            attachments=attachments or [],
            signature=signature or '',
        )

        return True

    def _check_node_completed(self, node):
        """检查节点是否已完成审批"""
        total = node.assignees.count()
        approved = node.assignees.filter(status='approved').count()
        rejected = node.assignees.filter(status='rejected').count()

        # 如果没有审批人（如部门无管理员），自动跳过该节点
        if total == 0:
            return True, 'approved'

        if node.request.sign_type == 'orsign':
            # 或签：任一通过=节点通过，任一驳回=整体驳回
            if approved > 0:
                return True, 'approved'
            if rejected > 0:
                return True, 'rejected'
            return False, None
        else:
            # 会签：全部通过=节点通过，任一驳回=整体驳回
            if rejected > 0:
                return True, 'rejected'
            if approved >= total:
                return True, 'approved'
            return False, None

    def _finalize_approval(self, approval):
        """完成审批：更新整体状态"""
        now = timezone.now()
        all_approved = True
        all_rejected = False

        for node in approval.approval_nodes.all():
            completed, result = self._check_node_completed(node)
            if result == 'rejected':
                all_rejected = True
                all_approved = False
                break
            if result != 'approved':
                all_approved = False

        if all_rejected:
            approval.status = 'rejected'
            approval.save()
            return True

        if all_approved:
            approval.status = 'approved'
            # 显式刷新 updated_at：save(update_fields=[...]) 不会更新 auto_now 字段，
            # 否则审批通过后更新时间仍停留在倒数第二个审批人的操作时间
            approval.updated_at = now
            if approval.resume_node_order is not None:
                approval.resume_node_order = None
                approval.save(update_fields=['status', 'resume_node_order', 'updated_at'])
            else:
                approval.save(update_fields=['status', 'updated_at'])
            # 物资单据通过后的业务流转（领用单回写台账防超领 / 需求单置为待采购）
            try:
                self._on_material_approved(approval)
            except Exception as e:
                logger.warning(f'物资单据审批通过后处理失败: {approval.id} {e}')
            return True

        return False

    def _on_material_approved(self, approval):
        """物资单据审批通过：领用单回写需求单已领台账；需求单进入待采购状态"""
        from .material_utils import write_requisition_ledger
        from .models import MaterialRequirement, MaterialRequisition
        if approval.approval_type == 'material_requisition':
            rec = MaterialRequisition.objects.filter(request=approval).first()
            if rec and rec.status != 'approved':
                rec.status = 'approved'
                rec.save(update_fields=['status'])
                write_requisition_ledger(rec)
                # 通知申请人：领用单已通过，凭领用单号可领料
                if rec.created_by_id:
                    send_work_notification(
                        user_id=rec.created_by_id,
                        title='物资领用单已通过',
                        content=f'领用单 {rec.doc_no}（关联需求单 {rec.requirement_doc_no or ""}）已通过审核，可凭单领用',
                        notification_type='approval',
                        related_url=f'/oa/approval/?approval_id={approval.id}',
                        extra_data={'approval_id': approval.id},
                    )
        elif approval.approval_type == 'material_requirement':
            rec = MaterialRequirement.objects.filter(request=approval).first()
            if rec and rec.status == 'pending':
                rec.status = 'approved'
                rec.save(update_fields=['status'])

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        try:
            approval = ApprovalRequest.objects.select_related('applicant').prefetch_related(
                'approval_nodes__assignees'
            ).get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)

        if approval.status not in ('pending', 'deferred', 'processing'):
            return Response({'error': '该审批已处理'}, status=400)

        can_approve = self._check_user_can_approve(approval, request.user)
        if not can_approve:
            return Response({'error': '您不在当前审批节点中'}, status=403)

        serializer = ApprovalActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comment = serializer.validated_data.get('comment', '')
        attachments = serializer.validated_data.get('attachments', [])
        signature = serializer.validated_data.get('signature', '')

        with transaction.atomic():
            if approval.approval_mode == 'sequential':
                nodes = approval.approval_nodes.filter(order=approval.current_node_order)
            else:
                nodes = approval.approval_nodes.all()

            processed = False
            for node in nodes:
                if self._process_node_approval(approval, node, 'approve', request.user, comment, attachments, signature):
                    processed = True
                    break

            if not processed:
                return Response({'error': '您已审批过，无需重复操作'}, status=400)

            # 检查当前节点是否完成
            all_finished = True
            for node in nodes:
                completed, result = self._check_node_completed(node)
                if not completed:
                    all_finished = False
                if completed and approval.approval_mode == 'sequential':
                    # 顺序审批：进入下一节点
                    next_order = approval.current_node_order + 1
                    if approval.approval_nodes.filter(order=next_order).exists():
                        approval.current_node_order = next_order
                        approval.save()
                    break

            if all_finished:
                # 所有相关节点完成：调用收尾逻辑（通过/驳回/仍有未完成节点则保持待审批）
                finished = self._finalize_approval(approval)
                if not finished:
                    # 节点完成但整体审批未完成（如还有后续节点/并行未完成）→ 恢复待审批
                    if approval.status in ('deferred', 'processing'):
                        approval.status = 'pending'
                        approval.save(update_fields=['status'])
            else:
                # 当前节点通过但还有待审批的节点/审批人 → 恢复待审批
                if approval.status in ('deferred', 'processing'):
                    approval.status = 'pending'
                    approval.save(update_fields=['status'])

        # 通知申请人
        timestamp = timezone.localtime(timezone.now()).strftime('%m-%d %H:%M')
        if approval.status == 'approved':
            send_work_notification(
                user_id=approval.applicant.id,
                title='审批已通过',
                content=f'您的{_approval_type_label(approval)}申请“{approval.title}”已通过 [{timestamp}]',
                notification_type='approval',
                related_url=f'/oa/approval/?approval_id={approval.id}',
                extra_data={'approval_id': approval.id, 'action': 'approved'},
            )
            # 票据回传：按配置的回传时限计算截止时间；仅当审批人开启「通知发起人回传票据」时才发送
            # 「请及时回传票据/凭证」工作通知给发起人（默认关闭则不发送）
            notify_receipt_return = bool(request.data.get('notify_receipt_return'))
            hours = 24  # 未配置时默认 24 小时
            try:
                r_tenant = approval.tenant or getattr(request, 'tenant', None) or approval.applicant.get_active_tenant()
                r_config = self._get_config_for_tenant(r_tenant, approval.approval_type) if r_tenant else None
                if r_config is not None and r_config.receipt_return_hours is not None:
                    hours = int(r_config.receipt_return_hours)
            except Exception as e:
                logger.warning(f'解析票据回传配置失败: {e}')
            if self._get_receipt_return_enabled(approval) and hours > 0:
                deadline = timezone.now() + timezone.timedelta(hours=hours)
                approval.receipt_deadline = deadline
                approval.save(update_fields=['receipt_deadline'])
                if notify_receipt_return:
                    send_work_notification(
                        user_id=approval.applicant.id,
                        title='请及时回传票据/凭证',
                        content=f'您的审批已通过，请在 {hours} 小时内（{timezone.localtime(deadline).strftime("%m-%d %H:%M")} 前）回传付款凭证或相关票据',
                        notification_type='approval',
                        related_url=f'/oa/approval/?approval_id={approval.id}',
                        extra_data={'approval_id': approval.id, 'action': 'receipt_return'},
                    )
            elif approval.receipt_deadline is not None:
                approval.receipt_deadline = None
                approval.save(update_fields=['receipt_deadline'])
        # 审批通过后才向抄送人发送抄送通知（驳回不通知抄送人）
        if approval.status == 'approved':
            self._notify_cc_after_end(approval)
        # 通知下一节点审批人（顺序审批）
        if approval.status == 'pending' and approval.approval_mode == 'sequential':
            next_order = approval.current_node_order
            next_assignees = ApprovalAssignee.objects.filter(
                node__request=approval, node__order=next_order, status='pending'
            ).select_related('user')
            for asgn in next_assignees:
                send_work_notification(
                    user_id=asgn.user.id,
                    title='审批待处理',
                    content=f'有新的审批需要您处理：“{approval.title}”',
                    notification_type='approval',
                    related_url=f'/oa/approval/?approval_id={approval.id}',
                    extra_data={'approval_id': approval.id, 'action': 'pending'},
                )

        logger.info(f'{request.user} 通过审批 {approval.title}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        try:
            approval = ApprovalRequest.objects.select_related('applicant').prefetch_related(
                'approval_nodes__assignees'
            ).get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)

        if approval.status not in ('pending', 'deferred', 'processing'):
            return Response({'error': '该审批已处理'}, status=400)


        can_approve = self._check_user_can_approve(approval, request.user)
        if not can_approve:
            return Response({'error': '您不在当前审批节点中'}, status=403)

        serializer = ApprovalActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comment = serializer.validated_data.get('comment', '')
        attachments = serializer.validated_data.get('attachments', [])
        signature = serializer.validated_data.get('signature', '')

        with transaction.atomic():
            if approval.approval_mode == 'sequential':
                nodes = approval.approval_nodes.filter(order=approval.current_node_order)
            else:
                nodes = approval.approval_nodes.all()

            processed = False
            rejected_node_order = None
            for node in nodes:
                if self._process_node_approval(approval, node, 'reject', request.user, comment, attachments, signature):
                    processed = True
                    rejected_node_order = node.order
                    break

            if not processed:
                return Response({'error': '您已审批过，无需重复操作'}, status=400)

            # 任何人驳回，整体驳回；记录驳回节点，便于重新提交时从该节点继续审批
            approval.status = 'rejected'
            approval.resume_node_order = rejected_node_order
            approval.save()

        # 通知申请人（驳回时不给抄送人/抄送部门发送通知，仅通知发起人）
        timestamp = timezone.localtime(timezone.now()).strftime('%m-%d %H:%M')
        send_work_notification(
            user_id=approval.applicant.id,
            title='审批已驳回',
            content=f'您的{_approval_type_label(approval)}申请“{approval.title}”已被驳回 [{timestamp}]' + (f' 原因：{comment}' if comment else ''),
            notification_type='approval',
            related_url=f'/oa/approval/?approval_id={approval.id}',
            extra_data={'approval_id': approval.id, 'action': 'rejected', 'comment': comment},
        )

        logger.info(f'{request.user} 驳回审批 {approval.title}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=True, methods=['post'])
    def deferred(self, request, pk=None):
        """暂缓审批"""
        return self._handle_status_action(request, pk, 'deferred', '暂缓')

    @action(detail=True, methods=['post'])
    def processing(self, request, pk=None):
        """正在办理"""
        return self._handle_status_action(request, pk, 'processing', '办理中')

    def _handle_status_action(self, request, pk, action, action_label):
        """处理暂缓/办理中等中间状态操作"""
        try:
            approval = ApprovalRequest.objects.select_related('applicant').prefetch_related(
                'approval_nodes__assignees'
            ).get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)

        can_approve = self._check_user_can_approve(approval, request.user)
        if not can_approve:
            return Response({'error': '您不在当前审批节点中'}, status=403)

        serializer = ApprovalActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comment = serializer.validated_data.get('comment', '')
        attachments = serializer.validated_data.get('attachments', [])
        signature = serializer.validated_data.get('signature', '')

        with transaction.atomic():
            if approval.approval_mode == 'sequential':
                nodes = approval.approval_nodes.filter(order=approval.current_node_order)
            else:
                nodes = approval.approval_nodes.all()

            processed = False
            for node in nodes:
                if self._process_node_approval(approval, node, action, request.user, comment, attachments, signature):
                    processed = True
                    break

            if not processed:
                return Response({'error': '操作失败，当前状态不允许此操作'}, status=400)

            # 更新审批整体状态为暂缓/办理中
            if action in ('deferred', 'processing'):
                approval.status = action
                approval.save(update_fields=['status'])

        # 通知申请人
        timestamp = timezone.localtime(timezone.now()).strftime('%m-%d %H:%M')
        status_info = {
            'deferred': '暂缓（审批人稍后会继续处理）',
            'processing': '正在办理中',
        }
        send_work_notification(
            user_id=approval.applicant.id,
            title=f'审批{action_label}',
            content=f'您的{_approval_type_label(approval)}申请“{approval.title}”已被审批人{action_label} [{timestamp}]' + (f' 意见：{comment}' if comment else ''),
            notification_type='approval',
            related_url=f'/oa/approval/?approval_id={approval.id}',
            extra_data={'approval_id': approval.id, 'action': action, 'comment': comment},
        )

        logger.info(f'{request.user} {action_label}审批 {approval.title}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=False, methods=['post'])
    def draft(self, request):
        """保存审批草稿"""
        serializer = ApprovalDraftSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        from accounts.models import Department
        department_id = serializer.validated_data.get('department_id')
        approval_type = serializer.validated_data.get('approval_type', 'other')
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        form_data = collect_form_data(serializer.validated_data,
                                      serializer.validated_data.get('form_data') or {},
                                      approval_type)
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                department = request.user.department
        else:
            department = request.user.department
        # 草稿也软解析收款方式：默认方式→快照用户收款账号（便于草稿详情展示收款信息），自定义→原样存储
        # 自定义类型：从 form_data 的「收款方式」字段提取
        from .type_utils import resolve_approval_type
        _tobj = resolve_approval_type(approval_type, tenant)
        _pm = serializer.validated_data.get('payment_method') or {}
        if _tobj and not _tobj.is_builtin:
            fd_pm, _ = self._extract_payment_from_form_data(form_data, _tobj)
            if fd_pm:
                _pm = fd_pm
        payment_method = self._resolve_payment_method_for_draft(request.user, _pm, approval_type) if isinstance(_pm, dict) and _pm.get('type') else {}
        approval = ApprovalRequest.objects.create(
            applicant=request.user,
            tenant=tenant,
            status='draft',
            department=department,
            approval_type=approval_type,
            title=serializer.validated_data.get('title', '未命名草稿'),
            content=serializer.validated_data.get('content', ''),
            start_date=serializer.validated_data.get('start_date'),
            end_date=serializer.validated_data.get('end_date'),
            duration=serializer.validated_data.get('duration'),
            amount=serializer.validated_data.get('amount'),
            expense_type=serializer.validated_data.get('expense_type', ''),
            expense_date=serializer.validated_data.get('expense_date'),
            attachments=serializer.validated_data.get('attachments', []),
            recruit_data=serializer.validated_data.get('recruit_data', {}),
            purchase_items=serializer.validated_data.get('purchase_items', []) or [],
            expense_items=serializer.validated_data.get('expense_items', []) or [],
            leave_type=serializer.validated_data.get('leave_type', '') or '',
            trip_data=serializer.validated_data.get('trip_data', {}) or {},
            payment_method=payment_method,
            form_data=form_data,
            sign_type=serializer.validated_data.get('sign_type', 'countersign'),
            approval_mode=serializer.validated_data.get('approval_mode', 'sequential'),
        )
        # 保存审批人节点（草稿也保留配置）
        approver_nodes = serializer.validated_data.get('approver_nodes', [])
        if not approver_nodes:
            approver_nodes = self._auto_determine_approvers(
                request.user, tenant,
                approval_type=approval_type,
                department_id=serializer.validated_data.get('department_id'),
                threshold_values=self._gather_threshold_values(serializer.validated_data, form_data),
            )
        from accounts.models import CustomUser, Department
        # 创建发起人节点
        ApprovalNode.objects.create(
            request=approval, node_type='initiator',
            user=request.user, order=0,
        )
        for idx, node_data in enumerate(approver_nodes):
            node_type = node_data.get('type', 'user')
            node = ApprovalNode.objects.create(
                request=approval, node_type=node_type, order=idx + 1,
            )
            if node_type == 'user':
                user_id = node_data.get('id')
                try:
                    u = CustomUser.objects.get(id=user_id)
                    node.user = u
                    node.save()
                    ApprovalAssignee.objects.create(node=node, user=u)
                except CustomUser.DoesNotExist:
                    pass
            elif node_type == 'department':
                dept_id = node_data.get('id')
                try:
                    dept = Department.objects.get(id=dept_id)
                    node.department = dept
                    node.save()
                    dept_admins = CustomUser.objects.filter(
                        department=dept,
                        user_type__in=['super_admin', 'admin']
                    ).exclude(id=request.user.id)
                    for au in dept_admins:
                        ApprovalAssignee.objects.create(node=node, user=au)
                except Department.DoesNotExist:
                    pass
        # 若配置了最终审批人，标记对应节点
        self._mark_final_approver_node(approval, request)
        # 草稿也保存抄送人
        cc_user_ids = serializer.validated_data.get('cc_users', [])
        cc_dept_ids = serializer.validated_data.get('cc_departments', [])
        self._create_cc_records(approval, request, cc_user_ids, cc_dept_ids)
        # 草稿也保存关联审批
        related_ids = serializer.validated_data.get('related_approvals', [])
        if related_ids:
            approval.related_approvals.set(related_ids)

        logger.info(f'{request.user} 保存审批草稿 {approval.id}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """撤销待审批的申请"""
        try:
            approval = ApprovalRequest.objects.get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)
        if approval.status != 'pending':
            return Response({'error': '仅待审批的申请可以撤销'}, status=400)
        if approval.applicant != request.user:
            return Response({'error': '只能撤销自己的申请'}, status=403)
        approval.status = 'cancelled'
        # 撤回视为全新流程：清除驳回续审标记
        if approval.resume_node_order is not None:
            approval.resume_node_order = None
            approval.save(update_fields=['status', 'resume_node_order'])
        else:
            approval.save(update_fields=['status'])
        logger.info(f'{request.user} 撤销审批 {approval.title}')
        # 记录撤销日志
        ApprovalLog.objects.create(
            request=approval,
            operator=request.user,
            action='cancel',
            comment='已撤回审批申请',
        )
        # 撤销通知对象：已审批过的用户 + 到达审批节点的用户；未到达节点且未审批的用户不通知。
        # 到达节点判定：
        #   并行 → 全部节点到达；
        #   顺序且 current_node_order>0 → order <= current_node_order 的节点；
        #   顺序且 current_node_order 缺失(旧数据=0) → 推断第一个“未最终处理”的节点为到达节点（新提交即第一个审批节点）。
        # 到达节点上的审批人（含 assignee 与 node.user）均通知，但严格限定在到达节点内，避免漏发/多发。
        notify_ids = set()
        # 1) 已审批过的审批人（无论节点）
        processed_qs = ApprovalAssignee.objects.filter(
            node__request=approval,
            status__in=['approved', 'rejected', 'deferred', 'processing'],
        ).exclude(node__request__applicant=request.user)
        notify_ids.update(processed_qs.values_list('user_id', flat=True))

        # 2) 到达节点
        if approval.approval_mode == 'parallel':
            reached_node_qs = ApprovalNode.objects.filter(request=approval)
        elif approval.current_node_order and approval.current_node_order > 0:
            reached_node_qs = ApprovalNode.objects.filter(
                request=approval, order__lte=approval.current_node_order)
        else:
            # 顺序且缺失 current_node_order：第一个有“未最终处理”审批人的节点
            first_unfinished = ApprovalNode.objects.filter(
                request=approval,
                assignees__status__in=['pending', 'deferred', 'processing'],
            ).distinct().order_by('order').first()
            reached_node_qs = ApprovalNode.objects.filter(pk=first_unfinished.pk) if first_unfinished else ApprovalNode.objects.none()
        reached_node_ids = list(reached_node_qs.values_list('id', flat=True))

        # 到达节点上的审批人（assignee 记录）
        reached_assignees = ApprovalAssignee.objects.filter(node_id__in=reached_node_ids).exclude(
            node__request__applicant=request.user)
        notify_ids.update(reached_assignees.values_list('user_id', flat=True))
        # 到达节点上的审批人（node.user，assignee 记录缺失时的兜底，仅限到达节点）
        reached_node_users = ApprovalNode.objects.filter(id__in=reached_node_ids, user__isnull=False).exclude(
            user=request.user)
        notify_ids.update(reached_node_users.values_list('user_id', flat=True))
        notify_ids.discard(request.user.id)
        logger.info(f'{request.user} 撤销审批 {approval.title} 待通知候选 notify_ids:{notify_ids} '
                    f'(assignees={ApprovalAssignee.objects.filter(node__request=approval).count()}, '
                    f'nodes={ApprovalNode.objects.filter(request=approval).count()}, '
                    f'mode={approval.approval_mode}, cur_node_order={approval.current_node_order}, '
                    f'reached_node_ids={reached_node_ids})')
        for uid in notify_ids:
            send_work_notification(
                user_id=uid,
                title='审批已撤销',
                content=f'{request.user.real_name or request.user.username} 撤销了审批申请“{approval.title}”',
                notification_type='approval',
                related_url=f'/oa/approval/?approval_id={approval.id}',
                extra_data={'approval_id': approval.id, 'action': 'cancelled'},
            )

        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=True, methods=['post'])
    def re_edit(self, request, pk=None):
        """重新编辑已撤回或已驳回的审批"""
        try:
            approval = ApprovalRequest.objects.get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)
        if approval.status not in ('cancelled', 'rejected', 'draft'):
            return Response({'error': '仅草稿、已撤回或已驳回的审批可以重新编辑'}, status=400)
        if approval.applicant != request.user:
            return Response({'error': '只能编辑自己的审批'}, status=403)
        serializer = ApprovalDraftSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        # 更新字段
        from accounts.models import Department
        department_id = serializer.validated_data.get('department_id')
        if department_id:
            try:
                approval.department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                pass
        for field in ['approval_type', 'title', 'content', 'start_date', 'end_date',
                      'duration', 'amount', 'expense_type', 'expense_date',
                      'sign_type', 'approval_mode']:
            val = serializer.validated_data.get(field)
            if val is not None:
                setattr(approval, field, val)
        if serializer.validated_data.get('attachments') is not None:
            approval.attachments = serializer.validated_data['attachments']
        if serializer.validated_data.get('recruit_data') is not None:
            approval.recruit_data = serializer.validated_data['recruit_data']
        if serializer.validated_data.get('purchase_items') is not None:
            approval.purchase_items = serializer.validated_data['purchase_items']
        if serializer.validated_data.get('expense_items') is not None:
            approval.expense_items = serializer.validated_data['expense_items']
        if serializer.validated_data.get('leave_type') is not None:
            approval.leave_type = serializer.validated_data['leave_type']
        if serializer.validated_data.get('trip_data') is not None:
            approval.trip_data = serializer.validated_data['trip_data']
        # 付款方式：重新编辑提交时同样校验/快照（报销/采购/自定义类型）
        if 'payment_method' in serializer.validated_data or approval.approval_type in ('expense', 'purchase'):
            from .type_utils import resolve_approval_type
            t_obj = resolve_approval_type(approval.approval_type, request.tenant or request.user.get_active_tenant())
            pm, pm_err = self._resolve_payment_method(
                request.user, serializer.validated_data.get('payment_method'), approval.approval_type, t_obj)
            if pm_err:
                return pm_err
            approval.payment_method = pm
        # 根据物项/项目自动计算总金额与出差天数
        purchase_items = approval.purchase_items or []
        expense_items = approval.expense_items or []
        trip_data = approval.trip_data or {}
        auto_amount = None
        if approval.approval_type == 'purchase' and purchase_items:
            try:
                auto_amount = sum(float(i.get('total') or 0) for i in purchase_items)
            except (ValueError, TypeError):
                auto_amount = None
        elif approval.approval_type == 'expense' and expense_items:
            try:
                auto_amount = sum(float(i.get('amount') or 0) for i in expense_items)
            except (ValueError, TypeError):
                auto_amount = None
        elif approval.approval_type == 'trip' and trip_data.get('amount'):
            try:
                auto_amount = float(trip_data.get('amount'))
            except (ValueError, TypeError):
                auto_amount = None
        if auto_amount is not None:
            approval.amount = auto_amount
        if approval.approval_type == 'trip' and trip_data.get('days'):
            try:
                approval.duration = float(trip_data.get('days'))
            except (ValueError, TypeError):
                pass
        form_data = collect_form_data(serializer.validated_data,
                                      serializer.validated_data.get('form_data') or {},
                                      approval.approval_type)
        approval.form_data = form_data
        # 物资单据：重新提交时同步业务记录（保留原单据号），校验失败则拒绝重新提交
        if approval.approval_type in ('material_requirement', 'material_requisition'):
            from decimal import Decimal as _D, InvalidOperation as _IO
            from .material_utils import ensure_material_record
            _mrec, merr = ensure_material_record(approval, form_data)
            if merr:
                return Response({'error': merr}, status=400)
            approval.form_data = form_data
            if form_data.get('amount'):
                try:
                    approval.amount = _D(str(form_data['amount']))
                except (_IO, ValueError, TypeError):
                    pass
        # 被驳回后重新提交：从驳回节点继续审批（已通过节点不再重审）。
        # 优先使用驳回时持久化的 resume_node_order（兼容“驳回→存草稿→再提交”路径，
        # 因为 update_draft 会重建审批链并清空节点状态）；同时收集当前链中驳回节点之前的已通过节点，
        # 供直接重新提交时校验审批人是否一致。
        resume_order = approval.resume_node_order
        if not resume_order:
            # 兜底：从当前审批链中被驳回的审批人推导驳回节点（兼容旧数据直接重新提交）
            rejected_asgn = ApprovalAssignee.objects.filter(
                node__request=approval, status='rejected'
            ).select_related('node').order_by('node__order').first()
            if rejected_asgn:
                resume_order = rejected_asgn.node.order
        old_approved = {}  # {node_order: set(user_id)}
        if resume_order:
            for n in ApprovalNode.objects.filter(
                request=approval, order__lt=resume_order
            ).prefetch_related('assignees'):
                asgns = list(n.assignees.all())
                if asgns and all(a.status == 'approved' for a in asgns):
                    old_approved[n.order] = set(a.user_id for a in asgns)
        approval.status = 'pending'
        approval.resume_node_order = None  # 重新提交后清空
        approval.save()
        # 重建审批人节点（根据所选部门自动生成）
        approver_nodes = serializer.validated_data.get('approver_nodes', [])
        if not approver_nodes:
            approver_nodes = self._auto_determine_approvers(
                request.user, request.tenant or request.user.get_active_tenant(),
                approval_type=approval.approval_type,
                department_id=department_id or getattr(approval.department, 'id', None),
                threshold_values=self._gather_threshold_values(serializer.validated_data, form_data),
            )
        ApprovalAssignee.objects.filter(node__request=approval).delete()
        approval.approval_nodes.all().delete()
        from accounts.models import CustomUser
        # 创建发起人节点
        ApprovalNode.objects.create(
            request=approval, node_type='initiator',
            user=request.user, order=0,
        )
        for idx, node_data in enumerate(approver_nodes):
            node_type = node_data.get('type', 'user')
            node = ApprovalNode.objects.create(
                request=approval, node_type=node_type, order=idx + 1,
            )
            if node_type == 'user':
                user_id = node_data.get('id')
                try:
                    u = CustomUser.objects.get(id=user_id)
                    node.user = u
                    node.save()
                    ApprovalAssignee.objects.create(node=node, user=u)
                except CustomUser.DoesNotExist:
                    pass
            elif node_type == 'department':
                dept_id = node_data.get('id')
                try:
                    dept = Department.objects.get(id=dept_id)
                    node.department = dept
                    node.save()
                    dept_admins = CustomUser.objects.filter(
                        department=dept,
                        user_type__in=['super_admin', 'admin']
                    ).exclude(id=request.user.id)
                    for au in dept_admins:
                        ApprovalAssignee.objects.create(node=node, user=au)
                except Department.DoesNotExist:
                    pass
        # 若配置了最终审批人，标记对应节点
        self._mark_final_approver_node(approval, request)
        # 被驳回后重新提交：从驳回节点继续审批，之前已通过的节点不再重新审批。
        # - 直接重新提交（未存草稿）：当前链中驳回节点之前的已通过节点仍在，校验审批人一致后保持通过；
        # - 驳回→存草稿→再提交：update_draft 已重建审批链并清空状态，此时顺序审批下驳回节点之前的节点必然已通过，直接保持通过。
        carry_resume = False
        if resume_order:
            if old_approved:
                carry_ok = True
                for order, old_users in old_approved.items():
                    node = ApprovalNode.objects.filter(request=approval, order=order).first()
                    if not node:
                        carry_ok = False
                        break
                    new_users = set(node.assignees.values_list('user_id', flat=True))
                    if new_users != old_users:
                        carry_ok = False
                        break
                if carry_ok:
                    # 之前已通过的节点保持通过，不再重新审批
                    for order in old_approved:
                        node = ApprovalNode.objects.filter(request=approval, order=order).first()
                        if node:
                            node.assignees.update(status='approved')
                    carry_resume = True
            else:
                # 经草稿中转：驳回节点之前的节点保持通过
                for node in ApprovalNode.objects.filter(request=approval, order__lt=resume_order):
                    node.assignees.update(status='approved')
                carry_resume = True
        if carry_resume:
            # 顺序审批：从驳回节点开始
            if approval.approval_mode == 'sequential':
                approval.current_node_order = resume_order
            approval.save()
        else:
            # 顺序审批重置到第一个节点
            if approval.approval_mode == 'sequential':
                approval.current_node_order = 1
            approval.save()

        # 记录重新提交的审批日志
        ApprovalLog.objects.create(
            request=approval,
            operator=request.user,
            action='resubmit',
            comment='重新提交审批申请',
        )

        # 通知审批人（顺序审批仅通知当前活动节点；被驳回后继续审批时即驳回节点）
        from accounts.models import CustomUser, Department
        assignee_qs = ApprovalAssignee.objects.filter(node__request=approval, status='pending')
        if approval.approval_mode == 'sequential':
            active_order = approval.current_node_order or 1
            active_node = approval.approval_nodes.filter(order=active_order).first()
            assignee_qs = assignee_qs.filter(node=active_node) if active_node else assignee_qs.none()
        for asgn in assignee_qs.select_related('user'):
            send_work_notification(
                user_id=asgn.user.id,
                title='审批待处理',
                content=f'{request.user.real_name or request.user.username} 提交了{_approval_type_label(approval)}申请：“{approval.title}”',
                notification_type='approval',
                related_url=f'/oa/approval/?approval_id={approval.id}',
                extra_data={'approval_id': approval.id, 'action': 'pending'},
            )

        # 更新抄送人
        approval.carbon_copies.all().delete()
        cc_user_ids = serializer.validated_data.get('cc_users', [])
        cc_dept_ids = serializer.validated_data.get('cc_departments', [])
        self._create_cc_records(approval, request, cc_user_ids, cc_dept_ids)
        # 更新关联审批
        related_ids = serializer.validated_data.get('related_approvals', [])
        if related_ids:
            approval.related_approvals.set(related_ids)
        else:
            approval.related_approvals.clear()

        logger.info(f'{request.user} 重新编辑审批 {approval.title}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=False, methods=['get'])
    def drafts(self, request):
        """获取当前用户的草稿列表"""
        qs = ApprovalRequest.objects.filter(applicant=request.user, status='draft').order_by('-created_at')
        results = ApprovalListSerializer(qs, many=True, context={'request': request}).data
        return Response({'results': results})

    @action(detail=True, methods=['delete'])
    def delete_draft(self, request, pk=None):
        """删除草稿"""
        try:
            approval = ApprovalRequest.objects.get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '审批不存在'}, status=404)
        if approval.status not in ('draft', 'cancelled'):
            return Response({'error': '仅草稿和已撤回的审批可以删除'}, status=400)
        if approval.applicant != request.user:
            return Response({'error': '只能删除自己的草稿'}, status=403)
        approval.delete()
        logger.info(f'{request.user} 删除草稿 {pk}')
        return Response({'message': 'ok'})

    @action(detail=True, methods=['post'])
    def update_draft(self, request, pk=None):
        """更新已有草稿（保持ID不变）"""
        try:
            approval = ApprovalRequest.objects.get(id=pk)
        except ApprovalRequest.DoesNotExist:
            return Response({'error': '草稿不存在'}, status=404)
        if approval.status not in ('draft', 'cancelled', 'rejected'):
            return Response({'error': '仅草稿、已撤回或已驳回的审批可以更新'}, status=400)
        if approval.applicant != request.user:
            return Response({'error': '只能更新自己的草稿'}, status=403)
        serializer = ApprovalDraftSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        from accounts.models import Department, CustomUser
        department_id = serializer.validated_data.get('department_id')
        if department_id:
            try:
                approval.department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                pass
        else:
            approval.department = request.user.department
        for field in ['approval_type', 'title', 'content', 'start_date', 'end_date',
                      'duration', 'amount', 'expense_type', 'expense_date',
                      'sign_type', 'approval_mode']:
            val = serializer.validated_data.get(field)
            if val is not None:
                setattr(approval, field, val)
        if serializer.validated_data.get('recruit_data') is not None:
            approval.recruit_data = serializer.validated_data['recruit_data']
        if serializer.validated_data.get('attachments') is not None:
            approval.attachments = serializer.validated_data['attachments']
        if serializer.validated_data.get('purchase_items') is not None:
            approval.purchase_items = serializer.validated_data['purchase_items'] or []
        if serializer.validated_data.get('expense_items') is not None:
            approval.expense_items = serializer.validated_data['expense_items'] or []
        if serializer.validated_data.get('leave_type') is not None:
            approval.leave_type = serializer.validated_data['leave_type'] or ''
        if serializer.validated_data.get('trip_data') is not None:
            approval.trip_data = serializer.validated_data['trip_data'] or {}
        form_data = collect_form_data(serializer.validated_data,
                                      serializer.validated_data.get('form_data') or {},
                                      approval.approval_type)
        # 收款方式：报销/采购用 serializer.payment_method；自定义类型从 form_data 的「收款方式」字段提取
        _pm = serializer.validated_data.get('payment_method') or {}
        from .type_utils import resolve_approval_type
        _tobj = resolve_approval_type(approval.approval_type, request.tenant or request.user.get_active_tenant())
        if _tobj and not _tobj.is_builtin:
            fd_pm, _ = self._extract_payment_from_form_data(form_data, _tobj)
            if fd_pm:
                _pm = fd_pm
        if _pm:
            approval.payment_method = self._resolve_payment_method_for_draft(request.user, _pm, approval.approval_type)
        elif serializer.validated_data.get('payment_method') is not None:
            approval.payment_method = {}
        approval.form_data = form_data
        # 重新编辑时保存为草稿
        if approval.status in ('cancelled', 'rejected'):
            approval.status = 'draft'
        approval.save()
        approver_nodes = serializer.validated_data.get('approver_nodes', [])
        if not approver_nodes:
            approver_nodes = self._auto_determine_approvers(
                request.user, request.tenant or request.user.get_active_tenant(),
                approval_type=approval.approval_type,
                department_id=department_id or getattr(approval.department, 'id', None),
                threshold_values=self._gather_threshold_values(serializer.validated_data, form_data),
            )
        from oa.models import ApprovalAssignee
        ApprovalAssignee.objects.filter(node__request=approval).delete()
        approval.approval_nodes.all().delete()
        # 创建发起人节点
        ApprovalNode.objects.create(
            request=approval, node_type='initiator',
            user=request.user, order=0,
        )
        for idx, node_data in enumerate(approver_nodes):
            node_type = node_data.get('type', 'user')
            node = ApprovalNode.objects.create(
                request=approval, node_type=node_type, order=idx + 1,
            )
            if node_type == 'user':
                user_id = node_data.get('id')
                try:
                    u = CustomUser.objects.get(id=user_id)
                    node.user = u
                    node.save()
                    ApprovalAssignee.objects.create(node=node, user=u)
                except CustomUser.DoesNotExist:
                    pass
            elif node_type == 'department':
                dept_id = node_data.get('id')
                try:
                    dept = Department.objects.get(id=dept_id)
                    node.department = dept
                    node.save()
                    dept_admins = CustomUser.objects.filter(
                        department=dept,
                        user_type__in=['super_admin', 'admin']
                    ).exclude(id=request.user.id)
                    for au in dept_admins:
                        ApprovalAssignee.objects.create(node=node, user=au)
                except Department.DoesNotExist:
                    pass
        # 若配置了最终审批人，标记对应节点
        self._mark_final_approver_node(approval, request)
        # 顺序审批重置到第一个节点
        if approval.approval_mode == 'sequential':
            approval.current_node_order = 1
            approval.save()
        # 更新抄送人
        approval.carbon_copies.all().delete()
        cc_user_ids = serializer.validated_data.get('cc_users', [])
        cc_dept_ids = serializer.validated_data.get('cc_departments', [])
        self._create_cc_records(approval, request, cc_user_ids, cc_dept_ids)
        # 更新关联审批
        related_ids = serializer.validated_data.get('related_approvals', [])
        if related_ids:
            approval.related_approvals.set(related_ids)
        else:
            approval.related_approvals.clear()
        logger.info(f'{request.user} 更新草稿 {approval.id}')
        data = ApprovalRequestSerializer(approval, context={'request': request}).data
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=False, methods=['get'])
    def admins(self, request):
        """获取可选的审批人列表"""
        from accounts.models import CustomUser
        admins_qs = CustomUser.objects.filter(
            Q(user_type='super_admin') | Q(user_type='admin')
        ).exclude(id=request.user.id).select_related('department').values(
            'id', 'username', 'real_name', 'department__name'
        )
        results = []
        for a in admins_qs:
            dept_name = a.get('department__name') or ''
            label = a['real_name'] or a['username']
            if dept_name:
                label += f' ({dept_name})'
            results.append({'id': a['id'], 'name': label, 'department': dept_name})
        return Response({'results': results})

    @action(detail=False, methods=['get'])
    def departments(self, request):
        """获取可选部门列表（仅含管理员以上的部门）"""
        from accounts.models import CustomUser, Department
        dept_ids = CustomUser.objects.filter(
            Q(user_type='super_admin') | Q(user_type='admin')
        ).exclude(id=request.user.id).values_list('department_id', flat=True).distinct()
        depts = Department.objects.filter(id__in=dept_ids).values('id', 'name')
        return Response({'results': list(depts)})

    @action(detail=False, methods=['get'])
    def all_departments(self, request):
        """获取所有部门列表（用于表单选择）"""
        from accounts.models import Department
        depts = Department.objects.all().values('id', 'name')
        return Response({'results': list(depts)})

    @action(detail=False, methods=['get'])
    def org_departments(self, request):
        """获取组织架构部门（链式结构）。

        ?scope=all：返回当前用户所属的所有企业下的所有部门（供自定义类型"部门"字段选择，
        非超管/管理员仅返回自己所属部门）。
        默认：新建审批"所属部门"下拉框，按角色限定可选范围：
          super_admin → 当前企业+子企业全部部门；
          admin → 自己所属的所有部门；
          其他用户 → 自己所属的主部门（无主部门时取最低级部门）。
        返回 default_id：默认选中部门（主部门 → 无主部门时最低级部门 → 都没有则 None）。
        """
        from accounts.models import Department, Tenant
        from org.models import UserDepartment
        tenant = getattr(request, 'tenant', None)
        if tenant is None:
            tenant = request.user.get_active_tenant()
        user = request.user
        scope = request.query_params.get('scope', '')
        if scope == 'all':
            # 当前用户所属的所有企业（多企业隔离原则：只取用户所属企业的部门）
            tenant_ids = list(Tenant.objects.filter(
                memberships__user=user, memberships__is_active=True
            ).distinct().values_list('id', flat=True))
            if not tenant_ids and tenant:
                tenant_ids = [tenant.id]
        else:
            tenant_ids = [tenant.id] if tenant else []
            try:
                sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                if user.user_type in ('super_admin', 'admin') and sub_ids:
                    tenant_ids.extend(sub_ids)
            except Exception:
                pass
        if not tenant_ids:
            return Response({'results': [], 'default_id': None})

        depts = list(Department.objects.filter(tenant_id__in=tenant_ids, is_active=True).values(
            'id', 'name', 'parent_id', 'full_path', 'tenant_id', 'level', 'department_type'))

        default_id = None
        if scope == 'all':
            # 供自定义类型"部门"字段使用：非超管/管理员只返回自己所属部门
            if user.user_type not in ('super_admin', 'admin') and depts:
                user_dept_ids = list(UserDepartment.objects.filter(user=user).values_list('department_id', flat=True))
                if user_dept_ids:
                    depts = [d for d in depts if d['id'] in user_dept_ids]
        else:
            user_dept_ids = list(UserDepartment.objects.filter(user=user).values_list('department_id', flat=True))
            if user.user_type == 'super_admin':
                # 全部部门，不做限制
                pass
            elif user.user_type == 'admin':
                # 自己所属的所有部门
                depts = [d for d in depts if d['id'] in user_dept_ids]
            else:
                # 其他用户：仅主部门（无主部门时取最低级部门）
                primary_ids = list(UserDepartment.objects.filter(
                    user=user, is_primary=True).values_list('department_id', flat=True))
                if primary_ids:
                    depts = [d for d in depts if d['id'] in primary_ids]
                elif user_dept_ids:
                    deep_rel = (UserDepartment.objects.filter(user=user)
                                .select_related('department')
                                .order_by('-department__level', 'id').first())
                    if deep_rel:
                        depts = [d for d in depts if d['id'] == deep_rel.department_id]
                    else:
                        depts = []
                else:
                    depts = []
            # 默认选中：主部门 → 最低级部门
            if depts:
                primary_rel = (UserDepartment.objects.filter(user=user, is_primary=True)
                               .select_related('department').first())
                if primary_rel and any(d['id'] == primary_rel.department_id for d in depts):
                    default_id = primary_rel.department_id
                else:
                    deep_rel = (UserDepartment.objects.filter(
                        user=user, department_id__in=[d['id'] for d in depts])
                        .select_related('department')
                        .order_by('-department__level', 'id').first())
                    if deep_rel:
                        default_id = deep_rel.department_id

        # 多企业时用企业名作前缀，便于区分
        if len(tenant_ids) > 1 and depts:
            tenant_names = {}
            for t in Tenant.objects.filter(id__in=tenant_ids):
                tenant_names[t.id] = t.short_name or t.name
            for d in depts:
                tn = tenant_names.get(d['tenant_id'], '')
                if tn:
                    d['name'] = f'[{tn}] {d["name"]}'
                    d['_tenant_name'] = tn
        return Response({'results': depts, 'default_id': default_id})

    @action(detail=False, methods=['get'])
    def geocode(self, request):
        """反向地理编码（百度地图API）"""
        lat = request.query_params.get('lat')
        lng = request.query_params.get('lng')
        if not lat or not lng:
            return Response({'error': '缺少经纬度参数'}, status=400)
        try:
            from utils.reverse_geocoding_to_city import baidu_geocoding
            ak = getattr(settings, 'BAIDU_MAP_SERVER_AK', '')
            if not ak:
                return Response({'error': '百度地图AK未配置'}, status=500)
            result = baidu_geocoding(float(lat), float(lng), ak, coordtype='wgs84ll')
            if result and result.get('status') == 0:
                data = result['result']
                formatted_poi = data.get('formatted_address_poi', '')
                formatted = data.get('formatted_address', '')
                comp = data.get('addressComponent', {})
                province = comp.get('province', '')
                city = comp.get('city', '')
                district = comp.get('district', '')
                street = comp.get('street', '')
                sematic = data.get('sematic_description', '')
                location_name = formatted_poi or formatted or sematic or f'{province}{city}{district}{street}'
                return Response({
                    'location': location_name,
                    'province': province, 'city': city,
                    'district': district, 'street': street,
                    'lng': lng, 'lat': lat,
                    'reverse_geocoding': result
                })
            err_msg = result.get('msg', '地理编码失败') if result else '请求百度地图API失败'
            return Response({'error': err_msg}, status=502)
        except Exception as e:
            logger.error(f'反向地理编码异常: {e}')
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_attachment(self, request):
        """上传审批附件"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请选择文件'}, status=400)
        if file.size > 500 * 1024 * 1024:
            return Response({'error': '文件大小不能超过500MB'}, status=400)
        ext = os.path.splitext(file.name)[1].lower()
        allowed = ['.jpg', '.jpeg', '.png', '.gif', '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.zip', '.mp4', '.avi', '.mov', '.mp3', '.wav']
        if ext not in allowed:
            return Response({'error': '不支持的文件格式'}, status=400)
        from django.core.files.storage import default_storage
        filename = f'approval_attachments/{uuid.uuid4().hex}{ext}'
        saved_path = default_storage.save(filename, file)
        file_url = settings.MEDIA_URL + saved_path
        logger.info(f'{request.user} 上传审批附件 {file.name}')
        return Response({'url': file_url, 'name': file.name})

    @action(detail=False, methods=['get'])
    def approval_chain(self, request):
        """获取当前用户的审批链预览（根据所选部门和审批类型动态生成）"""


        department_id = request.query_params.get('department_id', '').strip()
        approval_type = request.query_params.get('approval_type', '').strip()
        if department_id:
            try:
                department_id = int(department_id)
            except ValueError:
                department_id = None

        # Read threshold values from query params for preview
        threshold_values = {}
        for key in ('duration', 'amount', 'headcount'):
            val = request.query_params.get(key, '').strip()
            if val:
                try:
                    threshold_values[key] = float(val)
                except ValueError:
                    pass
        # 自定义类型：接收 form_data JSON，收集数字字段用于阈值预览
        fd_raw = request.query_params.get('form_data', '')
        if fd_raw:
            try:
                import json as _json
                fd = _json.loads(fd_raw)
                if isinstance(fd, dict):
                    for k, v in fd.items():
                        if isinstance(v, (int, float)) and not isinstance(v, bool):
                            threshold_values[str(k)] = float(v)
            except Exception:
                pass

        chain = self._auto_determine_approvers(
            request.user, request.user.get_active_tenant(),
            approval_type=approval_type or None,
            department_id=department_id,
            threshold_values=threshold_values if threshold_values else None,
        )

        # 添加级别标记用于前端展示
        from accounts.models import Department
        result = []
        level = 1
        dept_map = {}
        if department_id:
            try:
                sel_dept = Department.objects.get(id=department_id)
                dept_map['selected'] = sel_dept.name
                if sel_dept.parent:
                    dept_map['parent'] = sel_dept.parent.name
            except Department.DoesNotExist:
                pass
        tenant = request.tenant or request.user.get_active_tenant()
        final_dept = self._get_final_approval_dept(tenant, approval_type or None)
        if final_dept:
            dept_map['final'] = final_dept.name

        # 阈值审批部门信息（支持集团回溯）
        if approval_type:
            config = self._get_config_for_tenant(tenant, approval_type)
            if config and config.threshold_enabled and config.threshold_department:
                dept_map['threshold'] = config.threshold_department.name
                dept_map['threshold_field'] = _threshold_field_label(config)
                dept_map['threshold_value'] = config.threshold_value
            # 是否要求手写签名
            dept_map['require_signature'] = bool(config and config.require_signature)

        # 集团企业信息
        if tenant and tenant.parent:
            dept_map['group_tenant'] = tenant.parent.name
            dept_map['group_tenant_id'] = tenant.parent.id

        for item in chain:
            is_threshold = item.pop('_threshold_level', False)
            is_final_approver = item.pop('_final_approver', False)
            final_source = item.pop('_final_approver_source', '')
            final_source_label = item.pop('_final_approver_source_label', '')
            if is_final_approver:
                level_label = '最终审批人'
            elif is_threshold:
                level_label = '阈值审批'
            else:
                level_label = f'第{level}级'
            result.append({
                **item,
                'level': level,
                'level_label': level_label,
                'is_final_approver': is_final_approver,
                'final_approver_source': final_source,
                'final_approver_source_label': final_source_label,
            })
            level += 1
        # 添加部门信息
        return Response({'results': result, 'count': len(result), 'departments': dept_map})

    @action(detail=False, methods=['get'])
    def dept_configs(self, request):
        """获取当前企业的最终审批部门配置列表（集团可查看子企业配置及子公司列表）"""
        from .serializers import ApprovalDeptConfigSerializer
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'results': []})
        # 集团企业：查看当前企业+子企业配置
        query = Q(tenant=tenant)
        try:
            sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
            if sub_ids:
                query |= Q(tenant_id__in=sub_ids)
        except Exception:
            pass
        configs = ApprovalDeptConfig.objects.filter(query).select_related('department', 'sub_tenant')
        data = ApprovalDeptConfigSerializer(configs, many=True).data
        # 附加子企业列表，方便集团配置选择
        extra = {}
        try:
            sub_tenants = tenant.sub_tenants.filter(is_active=True).values('id', 'name', 'short_name', 'tenant_type')
            extra['sub_tenants'] = list(sub_tenants)
        except Exception:
            extra['sub_tenants'] = []
        return Response({'results': data, **extra})

    @action(detail=False, methods=['post'])
    def save_dept_config(self, request):
        """保存审批类型配置（企业超管/管理员用）"""
        user_type = request.user.user_type
        if user_type not in ('super_admin', 'admin'):
            return Response({'error': '仅企业超级管理员或管理员可操作'}, status=403)
        approval_type = request.data.get('approval_type', '').strip()
        if not approval_type:
            return Response({'error': '缺少审批类型参数'}, status=400)
        department_id = request.data.get('department_id')
        cc_departments = request.data.get('cc_departments', [])
        cc_users = request.data.get('cc_users', [])
        approver_users = request.data.get('approver_users', [])
        from accounts.models import Department, CustomUser, Tenant
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'error': '未找到所属企业'}, status=400)
        # 配置归属企业：子企业配置统一存放到集团（与 _get_config_for_tenant 的 config_tenant 逻辑一致）
        config_tenant = tenant.parent if tenant.parent else tenant

        # 验证部门存在且属于该企业
        dept = None
        if department_id:
            try:
                dept = Department.objects.get(id=department_id, tenant=tenant)
            except Department.DoesNotExist:
                return Response({'error': '最终审批部门不存在'}, status=400)

        # 验证抄送部门、抄送人、审批人都是当前企业成员
        if cc_departments:
            valid_depts = Department.objects.filter(id__in=cc_departments, tenant=tenant)
            cc_departments = list(valid_depts.values_list('id', flat=True))
        if cc_users:
            valid_users = CustomUser.objects.filter(
                id__in=cc_users, tenant_memberships__tenant=tenant,
                tenant_memberships__is_active=True
            ).distinct()
            cc_users = list(valid_users.values_list('id', flat=True))
        if approver_users:
            valid_approvers = CustomUser.objects.filter(
                id__in=approver_users, tenant_memberships__tenant=tenant,
                tenant_memberships__is_active=True
            ).distinct()
            approver_users = list(valid_approvers.values_list('id', flat=True))

        # 最终审批人（可选，允许置空；必须是当前企业成员）
        final_approver_id = request.data.get('final_approver')
        final_approver_obj = None
        if final_approver_id:
            final_approver_obj = CustomUser.objects.filter(
                id=final_approver_id, is_active=True,
                tenant_memberships__tenant=tenant,
                tenant_memberships__is_active=True
            ).first()
            if not final_approver_obj:
                return Response({'error': '最终审批人必须是当前企业成员'}, status=400)

        # Sub-tenant config (集团模式下为子公司配置，存放到集团 config_tenant 下)
        sub_tenant_id = request.data.get('sub_tenant_id')
        sub_tenant_obj = None
        if sub_tenant_id:
            try:
                sub_tenant_obj = Tenant.objects.get(id=int(sub_tenant_id))
                if sub_tenant_obj.parent_id != config_tenant.id:
                    return Response({'error': '指定的子公司不属于当前企业集团'}, status=400)
            except (ValueError, Tenant.DoesNotExist):
                return Response({'error': '子公司不存在'}, status=400)

        # Threshold fields
        threshold_enabled = request.data.get('threshold_enabled', False)
        threshold_field = request.data.get('threshold_field', '')
        threshold_value = request.data.get('threshold_value')
        threshold_department_id = request.data.get('threshold_department_id')
        threshold_dept = None
        if threshold_department_id:
            try:
                threshold_dept = Department.objects.get(id=threshold_department_id, tenant=tenant)
            except Department.DoesNotExist:
                pass

        defaults = {}
        if dept:
            defaults['department'] = dept
        else:
            defaults['department'] = None
        defaults['cc_departments'] = cc_departments
        defaults['cc_users'] = cc_users
        defaults['approver_users'] = approver_users
        defaults['final_approver'] = final_approver_obj
        sign_type = request.data.get('sign_type', '').strip()
        if sign_type in ('countersign', 'orsign'):
            defaults['default_sign_type'] = sign_type
        approval_mode = request.data.get('approval_mode', '').strip()
        if approval_mode in ('sequential', 'parallel'):
            defaults['default_approval_mode'] = approval_mode
        defaults['threshold_enabled'] = bool(threshold_enabled)
        defaults['threshold_field'] = threshold_field
        if threshold_value is not None:
            try:
                defaults['threshold_value'] = float(threshold_value)
            except (ValueError, TypeError):
                pass
        if threshold_dept:
            defaults['threshold_department'] = threshold_dept
        else:
            defaults['threshold_department'] = None
        # 手写签名开关
        if 'require_signature' in request.data:
            defaults['require_signature'] = bool(request.data.get('require_signature'))
        # 票据回传时限（小时，0=禁止回传）
        if 'receipt_return_hours' in request.data:
            try:
                defaults['receipt_return_hours'] = int(request.data.get('receipt_return_hours') or 0)
            except (ValueError, TypeError):
                pass
        # 票据回传开关（默认关闭）
        if 'enable_receipt_return' in request.data:
            defaults['enable_receipt_return'] = bool(request.data.get('enable_receipt_return'))

        try:
            config, created = ApprovalDeptConfig.objects.update_or_create(
                tenant=config_tenant,
                approval_type=approval_type,
                sub_tenant=sub_tenant_obj,
                defaults=defaults,
            )
        except Exception as e:
            logger.error(f'保存审批配置失败: {e}')
            return Response({'error': f'保存配置失败: {str(e)}'}, status=400)

        from .serializers import ApprovalDeptConfigSerializer
        data = ApprovalDeptConfigSerializer(config).data
        return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201 if created else 200)

    @action(detail=True, methods=['delete'])
    def delete_dept_config(self, request, pk=None):
        """删除最终审批部门配置"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可操作'}, status=403)
        try:
            config = ApprovalDeptConfig.objects.get(id=pk)
            config.delete()
            return Response({'message': 'ok'})
        except ApprovalDeptConfig.DoesNotExist:
            return Response({'error': '配置不存在'}, status=404)

    @action(detail=False, methods=['get'])
    def search_cc_users(self, request):
        """搜索可抄送的用户（当前企业成员）"""
        from accounts.models import CustomUser
        from django.db.models import Q
        keyword = request.query_params.get('search', '').strip()
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        # qs = CustomUser.objects.filter(is_active=True).exclude(id=request.user.id)
        qs = CustomUser.objects.filter(is_active=True)
        if tenant:
            qs = qs.filter(tenant_memberships__tenant=tenant, tenant_memberships__is_active=True)
        if keyword:
            qs = qs.filter(
                Q(real_name__icontains=keyword) | Q(username__icontains=keyword)
            )
        qs = qs.distinct()[:20]
        results = [{
            'id': u.id,
            'name': u.real_name or u.username,
            'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
            'position': u.position or '',
        } for u in qs]
        return Response({'results': results})

    @action(detail=False, methods=['get'])
    def search_cc_departments(self, request):
        """搜索可抄送的部门（当前企业部门）"""
        from accounts.models import Department
        from django.db.models import Q
        keyword = request.query_params.get('search', '').strip()
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        qs = Department.objects.filter(tenant=tenant, is_active=True)
        if keyword:
            qs = qs.filter(name__icontains=keyword)
        qs = qs.order_by('name')[:20]
        results = [{
            'id': d.id,
            'name': d.name,
            'manager_name': d.manager.real_name or d.manager.username if d.manager else '',
        } for d in qs]
        return Response({'results': results})

    @action(detail=False, methods=['get'])
    def my_pending(self, request):
        """获取当前用户待审批的列表"""
        assignee_ids = ApprovalAssignee.objects.filter(
            user=request.user, status='pending'
        ).values_list('node__request_id', flat=True).distinct()
        qs = ApprovalRequest.objects.filter(
            id__in=assignee_ids, status='pending'
        ).select_related('applicant').order_by('-created_at')
        results = ApprovalListSerializer(qs, many=True, context={'request': request}).data
        return Response({'results': results, 'count': len(results)})


class MaterialViewSet(viewsets.ViewSet):
    """物资管理：物品库 + 需求/领用联动（关联需求单搜索/自动带出）+ 入库确认"""
    permission_classes = [permissions.IsAuthenticated]

    def _tenant(self, request):
        return getattr(request, 'tenant', None) or request.user.get_active_tenant()

    def _admin(self, request):
        return request.user.user_type in ('super_admin', 'admin')

    # ===== 物品库 =====
    def items(self, request):
        """物品库列表（可搜索+分页）；含现有库存 = 已入库需求单数量 - 已通过领用单数量"""
        tenant = self._tenant(request)
        qs = MaterialItem.objects.filter(tenant=tenant, is_active=True)
        keyword = request.query_params.get('search', '').strip()
        if keyword:
            from django.db.models import Q
            qs = qs.filter(Q(name__icontains=keyword) | Q(spec__icontains=keyword) | Q(category__icontains=keyword))
        qs = qs.order_by('name')
        page = max(1, int(request.query_params.get('page', 1) or 1))
        page_size = max(1, int(request.query_params.get('page_size', 20) or 20))
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        items_page = list(qs[(page - 1) * page_size: page * page_size])
        from django.db.models import Sum
        stock_in = dict(MaterialRequirementItem.objects.filter(
            requirement__tenant=tenant, requirement__status='stocked'
        ).values('item_name').annotate(t=Sum('quantity')).values_list('item_name', 't'))
        stock_out = dict(MaterialRequisitionItem.objects.filter(
            requisition__tenant=tenant, requisition__status='approved'
        ).values('item_name').annotate(t=Sum('quantity')).values_list('item_name', 't'))
        data = [{
            'id': i.id, 'name': i.name, 'spec': i.spec, 'unit': i.unit,
            'category': i.category, 'price': str(i.price or ''),
            'stock': float((stock_in.get(i.name, 0) or 0) - (stock_out.get(i.name, 0) or 0)),
        } for i in items_page]
        return Response({'results': data, 'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages})

    def item_search(self, request):
        """物品名称联想（表单录入用）"""
        tenant = self._tenant(request)
        keyword = request.query_params.get('search', '').strip()
        qs = MaterialItem.objects.filter(tenant=tenant, is_active=True)
        if keyword:
            from django.db.models import Q
            qs = qs.filter(Q(name__icontains=keyword) | Q(spec__icontains=keyword))
        qs = qs.order_by('name')[:10]
        data = [{'id': i.id, 'name': i.name, 'spec': i.spec, 'unit': i.unit,
                 'price': str(i.price or '')} for i in qs]
        return Response({'results': data})

    def create_item(self, request):
        if not self._admin(request):
            return Response({'error': '仅企业管理员可操作'}, status=403)
        tenant = self._tenant(request)
        name = (request.data.get('name') or '').strip()
        if not name:
            return Response({'error': '物品名称不能为空'}, status=400)
        item = MaterialItem.objects.create(
            tenant=tenant, name=name,
            spec=(request.data.get('spec') or '').strip(),
            unit=(request.data.get('unit') or '').strip(),
            category=(request.data.get('category') or '').strip(),
            price=request.data.get('price') or None,
            created_by=request.user,
        )
        return Response({'message': 'ok', 'id': item.id}, status=201)

    def retrieve_item(self, request, pk=None):
        try:
            item = MaterialItem.objects.get(id=pk)
        except MaterialItem.DoesNotExist:
            return Response({'error': '物品不存在'}, status=404)
        return Response({'id': item.id, 'name': item.name, 'spec': item.spec, 'unit': item.unit,
                         'category': item.category, 'price': str(item.price or '')})

    def update_item(self, request, pk=None):
        if not self._admin(request):
            return Response({'error': '仅企业管理员可操作'}, status=403)
        try:
            item = MaterialItem.objects.get(id=pk)
        except MaterialItem.DoesNotExist:
            return Response({'error': '物品不存在'}, status=404)
        if request.data.get('name') is not None:
            item.name = (request.data.get('name') or '').strip()
        if 'spec' in request.data:
            item.spec = (request.data.get('spec') or '').strip()
        if 'unit' in request.data:
            item.unit = (request.data.get('unit') or '').strip()
        if 'category' in request.data:
            item.category = (request.data.get('category') or '').strip()
        if 'price' in request.data:
            item.price = request.data.get('price') or None
        if 'is_active' in request.data:
            item.is_active = bool(request.data.get('is_active', True))
        item.save()
        return Response({'message': 'ok'})

    def delete_item(self, request, pk=None):
        if not self._admin(request):
            return Response({'error': '仅企业管理员可操作'}, status=403)
        try:
            item = MaterialItem.objects.get(id=pk)
        except MaterialItem.DoesNotExist:
            return Response({'error': '物品不存在'}, status=404)
        item.delete()
        return Response({'message': 'ok'})

    # ===== 需求单联动 =====
    def requirement_search(self, request):
        """需求单搜索：返回本企业需求单并标注是否可领用（已入库且剩余>0 才能关联），
        未入库的需求单仅展示（灰色、不可选），让申请人清楚需求单已生成及当前状态"""
        tenant = self._tenant(request)
        keyword = request.query_params.get('search', '').strip()
        qs = MaterialRequirement.objects.filter(tenant=tenant).prefetch_related('items')
        if keyword:
            from django.db.models import Q
            qs = qs.filter(Q(doc_no__icontains=keyword) | Q(purpose__icontains=keyword)
                          | Q(branch_dept__name__icontains=keyword) | Q(items__item_name__icontains=keyword))
        qs = qs.distinct().order_by('-updated_at')[:50]
        status_labels = dict(MaterialRequirement.STATUS_CHOICES)
        data = []
        for r in qs:
            total_remain = float(sum((i.quantity - i.requisitioned_quantity) for i in r.items.all()))
            data.append({
                'id': r.id, 'doc_no': r.doc_no,
                'branch_dept': r.branch_dept.name if r.branch_dept else '',
                'purpose': r.purpose,
                'item_count': r.items.count(),
                'remaining': total_remain,
                'status': r.status,
                'status_label': status_labels.get(r.status, r.status),
                'linkable': r.status == 'stocked' and total_remain > 0,
                'created_at': r.created_at.strftime('%Y-%m-%d'),
            })
        return Response({'results': data})

    def requirement_detail(self, request):
        """需求单详情 + 明细（含剩余可领），供领用单自动带出"""
        rid = request.query_params.get('id', '').strip()
        if not rid:
            return Response({'error': '缺少需求单ID'}, status=400)
        try:
            r = MaterialRequirement.objects.get(id=int(rid))
        except (ValueError, TypeError, MaterialRequirement.DoesNotExist):
            return Response({'error': '需求单不存在'}, status=404)
        items = [{
            'item_name': i.item_name, 'spec': i.spec, 'unit': i.unit,
            'price': float(i.price) if i.price is not None else None,
            'quantity': float(i.quantity),
            'remaining': float(i.quantity - i.requisitioned_quantity),
            'remark': i.remark,
        } for i in r.items.all()]
        status_labels = dict(MaterialRequirement.STATUS_CHOICES)
        # 预估金额：优先取审批金额字段，兜底取审批 form_data
        amount = None
        if r.request_id:
            _req = ApprovalRequest.objects.filter(id=r.request_id).first()
            if _req:
                if _req.amount:
                    amount = float(_req.amount)
                elif _req.form_data.get('amount'):
                    try:
                        amount = float(_req.form_data['amount'])
                    except (ValueError, TypeError):
                        amount = None
        return Response({'encrypt': True, 'data': encrypt_data({
            'id': r.id, 'doc_no': r.doc_no,
            'branch_dept': r.branch_dept.name if r.branch_dept else '',
            'purpose': r.purpose,
            'status': r.status,
            'status_label': status_labels.get(r.status, r.status),
            'applicant': (r.created_by.real_name or r.created_by.username) if r.created_by else '',
            'request_id': r.request_id,
            'amount': amount,
            'items': items,
        })})

    def stock_in(self, request):
        """入库确认：需求单已采购入库，之后才可被领用（仅企业管理员/超管）"""
        if not self._admin(request):
            return Response({'error': '仅企业管理员可操作'}, status=403)
        rid = request.data.get('id') or request.data.get('requirement_id')
        if not rid:
            return Response({'error': '缺少需求单ID'}, status=400)
        try:
            r = MaterialRequirement.objects.get(id=int(rid))
        except (ValueError, TypeError, MaterialRequirement.DoesNotExist):
            return Response({'error': '需求单不存在'}, status=404)
        if r.request.status != 'approved':
            return Response({'error': '需求单审批通过后才能确认入库'}, status=400)
        r.status = 'stocked'
        r.save(update_fields=['status'])
        try:
            if r.created_by_id:
                send_work_notification(
                    user_id=r.created_by_id,
                    title='物资已入库',
                    content=f'需求单 {r.doc_no} 已入库，可发起领用',
                    notification_type='approval',
                    related_url=f'/oa/approval/?approval_id={r.request_id}',
                    extra_data={'approval_id': r.request_id},
                )
        except Exception:
            pass
        return Response({'message': '已确认入库'})

    def requirements(self, request):
        """物资需求单列表（物资管理界面，可搜索单号/分公司/用途 + 分页）"""
        from django.db.models import Q
        tenant = self._tenant(request)
        qs = MaterialRequirement.objects.filter(tenant=tenant).select_related('branch_dept', 'created_by') \
            .prefetch_related('items')
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(doc_no__icontains=search) | Q(purpose__icontains=search)
                          | Q(branch_dept__name__icontains=search))
        qs = qs.order_by('-updated_at')
        page = max(1, int(request.query_params.get('page', 1) or 1))
        page_size = max(1, int(request.query_params.get('page_size', 20) or 20))
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        status_labels = dict(MaterialRequirement.STATUS_CHOICES)
        data = []
        for r in list(qs[(page - 1) * page_size: page * page_size]):
            items = list(r.items.all())
            total_remain = float(sum((i.quantity - i.requisitioned_quantity) for i in items))
            data.append({
                'id': r.id, 'doc_no': r.doc_no,
                'branch_dept': r.branch_dept.name if r.branch_dept else '',
                'purpose': r.purpose,
                'status': r.status,
                'status_label': status_labels.get(r.status, r.status),
                'item_count': len(items),
                'remaining': round(total_remain, 2),
                'applicant': (r.created_by.real_name or r.created_by.username) if r.created_by else '',
                'request_status': r.request.status if r.request_id else '',
                'request_id': r.request_id,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M'),
            })
        return Response({'encrypt': True, 'data': encrypt_data(
            {'results': data, 'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages})})

    def requisitions(self, request):
        """物资领用单列表（物资管理界面，可搜索单号/需求单号 + 分页）"""
        from django.db.models import Q
        tenant = self._tenant(request)
        qs = MaterialRequisition.objects.filter(tenant=tenant).select_related('requirement', 'branch_dept', 'created_by') \
            .prefetch_related('items')
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(doc_no__icontains=search) | Q(requirement_doc_no__icontains=search)
                          | Q(branch_dept__name__icontains=search))
        qs = qs.order_by('-updated_at')
        page = max(1, int(request.query_params.get('page', 1) or 1))
        page_size = max(1, int(request.query_params.get('page_size', 20) or 20))
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        status_labels = dict(MaterialRequisition.STATUS_CHOICES)
        data = []
        for r in list(qs[(page - 1) * page_size: page * page_size]):
            data.append({
                'id': r.id, 'doc_no': r.doc_no,
                'requirement_doc_no': r.requirement_doc_no or (r.requirement.doc_no if r.requirement else ''),
                'branch_dept': r.branch_dept.name if r.branch_dept else '',
                'purpose': r.purpose,
                'status': r.status,
                'status_label': status_labels.get(r.status, r.status),
                'item_count': r.items.count(),
                'applicant': (r.created_by.real_name or r.created_by.username) if r.created_by else '',
                'request_status': r.request.status if r.request_id else '',
                'request_id': r.request_id,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M'),
            })
        return Response({'encrypt': True, 'data': encrypt_data(
            {'results': data, 'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages})})

    def requirement_status(self, request):
        """需求单状态流转：已通过(待采购) → 采购中 → 已入库(可领用)（仅企业管理员/超管）"""
        if not self._admin(request):
            return Response({'error': '仅企业管理员可操作'}, status=403)
        rid = request.data.get('id') or request.data.get('requirement_id')
        status = (request.data.get('status') or '').strip()
        if not rid:
            return Response({'error': '缺少需求单ID'}, status=400)
        if status not in ('purchasing', 'stocked'):
            return Response({'error': '不支持的状态'}, status=400)
        try:
            r = MaterialRequirement.objects.get(id=int(rid))
        except (ValueError, TypeError, MaterialRequirement.DoesNotExist):
            return Response({'error': '需求单不存在'}, status=404)
        if r.request.status != 'approved':
            return Response({'error': '需求单审批通过后才能进入采购/入库'}, status=400)
        allowed = {'purchasing': {'approved'}, 'stocked': {'approved', 'purchasing'}}
        if r.status not in allowed.get(status, set()):
            return Response({'error': f'当前状态({r.status})不能流转到({status})'}, status=400)
        r.status = status
        r.save(update_fields=['status'])
        try:
            if r.created_by_id:
                label = '已入库(可领用)' if status == 'stocked' else '采购中'
                send_work_notification(
                    user_id=r.created_by_id,
                    title='物资需求单状态更新',
                    content=f'需求单 {r.doc_no} {label}',
                    notification_type='approval',
                    related_url=f'/oa/approval/?approval_id={r.request_id}',
                    extra_data={'approval_id': r.request_id},
                )
        except Exception:
            pass
        return Response({'message': '已更新'})


class WatermarkViewSet(viewsets.ViewSet):
    """企业水印配置：管理控制台维护（仅超管），各页面加载渲染水印（显性+隐性）"""
    permission_classes = [permissions.IsAuthenticated]

    def _tenant(self, request):
        return getattr(request, 'tenant', None) or request.user.get_active_tenant()

    def _config_data(self, cfg):
        return {
            'enabled': bool(cfg.enabled),
            'company_name': cfg.company_name,
            'text': cfg.text,
            'font_size': cfg.font_size,
            'font_color': cfg.font_color,
            'font_style': cfg.font_style,
            'rotation': cfg.rotation,
            'opacity': float(cfg.opacity),
            'position': cfg.position,
            'shape': cfg.shape,
            'hidden_enabled': bool(cfg.hidden_enabled),
            'hidden_opacity': float(cfg.hidden_opacity),
            'print_enabled': bool(cfg.print_enabled),
            'page_enabled': cfg.page_enabled or {},
        }

    def config(self, request):
        """获取当前企业水印配置（各页面加载水印用；附当前登录 IP 供隐性水印溯源）"""
        tenant = self._tenant(request)
        # 无有效企业（如系统级账号/未归属企业）时不渲染水印，避免建配置报错
        if not tenant:
            return Response({'encrypt': True, 'data': encrypt_data({
                'config': {'enabled': False, 'page_enabled': {}},
                'extra': {'ip': get_request_ip(request)},
            })})
        defaults = {'page_enabled': {p: True for p in DEFAULT_WATERMARK_PAGES}}
        cfg, _ = WatermarkConfig.objects.get_or_create(tenant=tenant, defaults=defaults)
        if not cfg.page_enabled:
            cfg.page_enabled = {p: True for p in DEFAULT_WATERMARK_PAGES}
            cfg.save(update_fields=['page_enabled'])
        return Response({'encrypt': True, 'data': encrypt_data({
            'config': self._config_data(cfg),
            'extra': {'ip': get_request_ip(request)},
        })})

    def save_config(self, request):
        """保存水印配置（仅超级管理员）"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可配置'}, status=403)
        tenant = self._tenant(request)
        if not tenant:
            return Response({'error': '未找到所属企业，无法配置水印'}, status=400)
        cfg, _ = WatermarkConfig.objects.get_or_create(
            tenant=tenant, defaults={'page_enabled': {p: True for p in DEFAULT_WATERMARK_PAGES}})
        if 'enabled' in request.data:
            cfg.enabled = bool(request.data.get('enabled'))
        if request.data.get('company_name') not in (None, ''):
            cfg.company_name = str(request.data.get('company_name')).strip()
        if 'text' in request.data:
            cfg.text = str(request.data.get('text') or '')
        if 'font_size' in request.data:
            try:
                cfg.font_size = max(8, int(request.data.get('font_size')))
            except (ValueError, TypeError):
                pass
        if request.data.get('font_color') not in (None, ''):
            cfg.font_color = str(request.data.get('font_color')).strip()
        if 'font_style' in request.data:
            s = str(request.data.get('font_style'))
            cfg.font_style = s if s in ('normal', 'bold', 'italic') else 'normal'
        if 'rotation' in request.data:
            try:
                cfg.rotation = int(request.data.get('rotation'))
            except (ValueError, TypeError):
                pass
        if 'opacity' in request.data:
            try:
                cfg.opacity = max(0.0, min(1.0, float(request.data.get('opacity'))))
            except (ValueError, TypeError):
                pass
        if 'position' in request.data:
            cfg.position = str(request.data.get('position'))
        if 'shape' in request.data:
            cfg.shape = str(request.data.get('shape'))
        if 'hidden_enabled' in request.data:
            cfg.hidden_enabled = bool(request.data.get('hidden_enabled'))
        if 'hidden_opacity' in request.data:
            try:
                cfg.hidden_opacity = max(0.0, min(1.0, float(request.data.get('hidden_opacity'))))
            except (ValueError, TypeError):
                pass
        if 'print_enabled' in request.data:
            cfg.print_enabled = bool(request.data.get('print_enabled'))
        if 'page_enabled' in request.data and isinstance(request.data.get('page_enabled'), dict):
            merged = dict(cfg.page_enabled or {})
            for k, v in request.data.get('page_enabled').items():
                merged[str(k)] = bool(v)
            cfg.page_enabled = merged
        cfg.save()
        return Response({'encrypt': True, 'data': encrypt_data(self._config_data(cfg))})


class PrintLogViewSet(viewsets.ViewSet):
    """打印操作留痕：记录用户打印了哪个页面的什么内容，供打印统计/打印权限分配使用"""
    permission_classes = [permissions.IsAuthenticated]

    def create(self, request):
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        try:
            count = int(request.data.get('count') or 0)
        except (TypeError, ValueError):
            count = 0
        # 校验「允许打印」权限（用户自定义覆盖全局），结果随响应返回，前端据 allowed=false 阻止打印
        allowed = get_effective_operation_permission(request.user, 'allow_print', tenant)
        # 打印操作一律记录到打印日志（含无权限的尝试），便于管理员查看打印统计与权限分配
        try:
            PrintLog.objects.create(
                tenant=tenant,
                user=request.user,
                page=str(request.data.get('page') or 'other')[:50],
                target_type=str(request.data.get('target_type') or '')[:50],
                target_id=str(request.data.get('target_id') or '')[:100],
                count=max(count, 0),
                ip=get_request_ip(request),
            )
        except Exception as e:
            logger.warning(f'记录打印日志失败: {e}')
        return Response({'encrypt': True, 'data': encrypt_data({'ok': True, 'allowed': bool(allowed)})})


def _build_summary_pdf(s, dept_name):
    """生成每日工作总结真实 PDF 文件（优先 reportlab 内置中文 CID 字体，可正确渲染中文）"""
    import html as _html
    title = '每日工作总结'
    lines = [
        f'员工：{s.user.real_name or s.user.username}',
        f'所属部门：{dept_name}',
        f'职位：{s.position}',
        f'日期：{s.summary_date}',
        f'状态：{s.get_status_display()}',
        '',
        '当日工作总结：',
        s.content or '（未填写总结文字）',
        '',
        '大模型分析建议：',
        s.analysis_result or '（模型未返回内容）',
    ]
    text = '\n'.join(lines)
    # 1) reportlab：内置 STSong-Light CID 字体，原生支持中文
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        from reportlab.pdfgen import canvas
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
        buf = BytesIO()
        width, height = A4
        margin = 20 * mm
        c = canvas.Canvas(buf, pagesize=A4)
        y = height - margin
        c.setFont('STSong-Light', 16)
        c.drawString(margin, y, title)
        y -= 34
        c.setFont('STSong-Light', 11)
        for ln in text.split('\n'):
            if y < margin:
                c.showPage()
                c.setFont('STSong-Light', 11)
                y = height - margin
            c.drawString(margin, y, ln)
            y -= 16
        c.showPage()
        c.save()
        return buf.getvalue()
    except Exception:
        pass
    # 2) weasyprint：HTML → PDF（可含中文，需系统依赖）
    try:
        from weasyprint import HTML
        body = ''.join(f'<p>{_html.escape(ln)}</p>' for ln in lines)
        html_doc = f'<html><head><meta charset="utf-8"><style>body{{font-family:"SimSun","Microsoft YaHei",sans-serif;line-height:1.7;}}</style></head><body><h2>{_html.escape(title)}</h2>{body}</body></html>'
        return HTML(string=html_doc).write_pdf()
    except Exception:
        pass
    # 3) fpdf2：仅英文兼容（无中文字体时中文会乱码，作为最后兜底）
    try:
        from fpdf import FPDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('helvetica', 'B', 16)
        pdf.cell(0, 10, 'Daily Work Summary')
        pdf.ln(12)
        pdf.set_font('helvetica', '', 10)
        for ln in lines:
            pdf.multi_cell(0, 6, ln)
            pdf.ln(1)
        return bytes(pdf.output())
    except Exception:
        pass
    return {'error': '服务器未安装 PDF 生成库（reportlab/fpdf2/weasyprint），请管理员执行 pip install reportlab'}


class DailyWorkSummaryViewSet(viewsets.ViewSet):
    """每日工作总结：员工上传当日工作数据（图片/文档/表格）+ 总结文字，
    系统按职位调用火山方舟大模型（Celery 流式）分析推理，结果保存到数据库供前端流式展示。"""
    permission_classes = [permissions.IsAuthenticated]

    def _tenant(self, request):
        return getattr(request, 'tenant', None) or request.user.get_active_tenant()

    def _primary_dept_name(self, user):
        try:
            from org.models import UserDepartment
            ud = UserDepartment.objects.filter(user=user, is_primary=True).select_related('department').first()
            if ud and ud.department:
                return ud.department.name
        except Exception:
            pass
        return (user.department.name if user.department else '') or ''

    def _can_view_all(self, user):
        """部门负责人 / 超级管理员可查看团队总结"""
        if user.user_type == 'super_admin':
            return True
        try:
            from accounts.models import Department
            return Department.objects.filter(Q(manager=user) | Q(deputy_managers=user)).exists()
        except Exception:
            return False

    def _managed_dept_ids(self, user):
        """用户负责的部门（含其所有子部门）id 集合"""
        from accounts.models import Department
        ids = set()
        roots = list(Department.objects.filter(Q(manager=user) | Q(deputy_managers=user)))
        stack = list(roots)
        while stack:
            d = stack.pop()
            if d.id in ids:
                continue
            ids.add(d.id)
            stack.extend(d.children.all())
        return ids

    def _scoped_user_ids(self, user):
        """部门负责人可见的成员用户 id 集合（含子部门成员），超管返回 None 表示不限制"""
        if user.user_type == 'super_admin':
            return None
        from org.models import UserDepartment
        dept_ids = self._managed_dept_ids(user)
        if not dept_ids:
            return set()
        return set(UserDepartment.objects.filter(department_id__in=dept_ids).values_list('user_id', flat=True))

    def _can_view_summary(self, user, s):
        """判断某用户能否查看该总结：本人 / 超管 / 其负责部门成员"""
        if s.user_id == user.id:
            return True
        if user.user_type == 'super_admin':
            return True
        from org.models import UserDepartment
        dept_ids = self._managed_dept_ids(user)
        if not dept_ids:
            return False
        return UserDepartment.objects.filter(user_id=s.user_id, department_id__in=dept_ids).exists()

    def _summary_data(self, s, with_result=False):
        d = {
            'id': s.id,
            'user': s.user_id,
            'user_name': s.user.real_name or s.user.username,
            'user_avatar': s.user.get_avatar_url() if hasattr(s.user, 'get_avatar_url') else '',
            'department': self._primary_dept_name(s.user),
            'position': s.position,
            'summary_date': str(s.summary_date),
            'content': s.content,
            'files': s.files or [],
            'status': s.status,
            'created_at': s.created_at.isoformat(),
            'analyzed_at': s.analyzed_at.isoformat() if s.analyzed_at else None,
            'error_message': s.error_message,
        }
        if with_result:
            d['analysis_result'] = s.analysis_result
            d['prompt_text'] = s.prompt_text
        return d

    def _file_type(self, name):
        ext = os.path.splitext(name or '')[1].lower()
        if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
            return 'image'
        if ext in ('.xls', '.xlsx', '.csv'):
            return 'table'
        if ext in ('.doc', '.docx', '.pdf', '.txt', '.md'):
            return 'doc'
        return 'file'

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload(self, request):
        files = request.FILES.getlist('files')
        if not files and request.FILES.get('file'):
            files = [request.FILES['file']]
        if not files:
            return Response({'error': '请选择要上传的文件'}, status=400)
        from django.core.files.storage import default_storage
        out = []
        try:
            for f in files:
                ext = os.path.splitext(f.name)[1].lower() or ''
                path = 'work_summary/' + timezone.now().strftime('%Y%m') + '/' + uuid.uuid4().hex + ext
                saved = default_storage.save(path, f)
                out.append({'name': f.name, 'url': default_storage.url(saved), 'type': self._file_type(f.name), 'size': f.size})
        except Exception as e:
            logger.warning(f'每日总结文件上传失败: {e}')
            return Response({'error': f'上传失败：{e}'}, status=500)
        return Response({'encrypt': True, 'data': encrypt_data({'files': out})})

    def create(self, request):
        from datetime import date as dt_date
        summary_date = request.data.get('summary_date')
        content = (request.data.get('content') or '').strip()
        if not summary_date:
            return Response({'error': '请选择工作总结日期'}, status=400)
        try:
            d = dt_date.fromisoformat(str(summary_date))
        except (ValueError, TypeError):
            return Response({'error': '日期格式错误'}, status=400)
        files = request.data.get('files')
        if isinstance(files, str):
            try:
                files = json.loads(files)
            except Exception:
                files = []
        if not isinstance(files, list):
            files = []
        user = request.user
        tenant = self._tenant(request)
        from .models import WorkSummaryConfig
        cfg = WorkSummaryConfig.get_config(tenant)
        status = 'pending' if cfg.enabled else 'disabled'
        if status == 'pending':
            # 灰度试点：不在分析范围内的员工仅保存总结，不调用大模型
            if not cfg.in_scope(user):
                status = 'not_allowed'
            else:
                # 当日调用量/成本限额已满：降级为仅保存总结
                from .tasks import check_summary_quota
                blocked, _reason = check_summary_quota(cfg)
                if blocked:
                    status = 'limited'
        s = DailyWorkSummary.objects.create(
            user=user, tenant=tenant, summary_date=d, content=content,
            position=user.position or '', files=files, status=status)
        if status == 'pending':
            try:
                from .tasks import analyze_work_summary_task
                analyze_work_summary_task.delay(s.id)
            except Exception as e:
                logger.warning(f'触发每日总结分析失败: {e}')
        return Response({'encrypt': True, 'data': encrypt_data(self._summary_data(s))}, status=201)

    def list(self, request):
        """我的每日工作总结（可按日期/总结内容/账号名搜索）"""
        user = request.user
        qs = DailyWorkSummary.objects.filter(user=user).select_related('user').order_by('-summary_date', '-created_at')
        date_from = request.query_params.get('date_from', '').strip()
        date_to = request.query_params.get('date_to', '').strip()
        if date_from:
            qs = qs.filter(summary_date__gte=date_from)
        if date_to:
            qs = qs.filter(summary_date__lte=date_to)
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(content__icontains=search) | Q(user__username__icontains=search) | Q(user__real_name__icontains=search))
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        items = qs[(page - 1) * page_size:page * page_size]
        return Response({'encrypt': True, 'data': encrypt_data({
            'results': [self._summary_data(x, True) for x in items],
            'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages,
        })})

    @action(detail=False, methods=['get'])
    def all(self, request):
        """团队每日工作总结（部门负责人/超管）：部门过滤 + 日期过滤 + 搜索员工"""
        if not self._can_view_all(request.user):
            return Response({'error': '仅部门负责人或超级管理员可查看'}, status=403)
        tenant = self._tenant(request)
        qs = DailyWorkSummary.objects.select_related('user').order_by('-summary_date', '-created_at')
        if tenant:
            ids = [tenant.id]
            try:
                ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
            except Exception:
                pass
            qs = qs.filter(tenant_id__in=ids)
        # 部门负责人（非超管）：只能查看自己负责部门（含子部门）成员的总结
        if request.user.user_type != 'super_admin':
            scoped = self._scoped_user_ids(request.user)
            qs = qs.filter(user_id__in=scoped) if scoped else qs.none()
        dept_id = request.query_params.get('department_id', '').strip()
        if dept_id:
            try:
                from org.models import UserDepartment
                member_ids = list(UserDepartment.objects.filter(department_id=int(dept_id)).values_list('user_id', flat=True))
                qs = qs.filter(user_id__in=member_ids)
            except Exception:
                pass
        date_from = request.query_params.get('date_from', '').strip()
        date_to = request.query_params.get('date_to', '').strip()
        if date_from:
            qs = qs.filter(summary_date__gte=date_from)
        if date_to:
            qs = qs.filter(summary_date__lte=date_to)
        user_id = request.query_params.get('user_id', '').strip()
        if user_id:
            try:
                qs = qs.filter(user_id=int(user_id))
            except (ValueError, TypeError):
                pass
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(user__username__icontains=search) | Q(user__real_name__icontains=search))
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        items = qs[(page - 1) * page_size:page * page_size]
        return Response({'encrypt': True, 'data': encrypt_data({
            'results': [self._summary_data(x, True) for x in items],
            'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages,
        })})

    def retrieve(self, request, pk=None):
        try:
            s = DailyWorkSummary.objects.select_related('user').get(id=pk)
        except DailyWorkSummary.DoesNotExist:
            return Response({'error': '工作总结不存在'}, status=404)
        if not self._can_view_summary(request.user, s):
            return Response({'error': '无权查看该工作总结'}, status=403)
        return Response({'encrypt': True, 'data': encrypt_data(self._summary_data(s, True))})

    def destroy(self, request, pk=None):
        """仅超级管理员可删除工作总结"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可删除工作总结'}, status=403)
        try:
            s = DailyWorkSummary.objects.get(id=pk)
        except DailyWorkSummary.DoesNotExist:
            return Response({'error': '工作总结不存在'}, status=404)
        s.delete()
        return Response({'encrypt': True, 'data': encrypt_data({'ok': True})})

    @action(detail=True, methods=['post'])
    def analyze(self, request, pk=None):
        """重新触发大模型分析"""
        try:
            s = DailyWorkSummary.objects.select_related('user').get(id=pk)
        except DailyWorkSummary.DoesNotExist:
            return Response({'error': '工作总结不存在'}, status=404)
        if not self._can_view_summary(request.user, s):
            return Response({'error': '无权操作'}, status=403)
        try:
            from .tasks import analyze_work_summary_task
            analyze_work_summary_task.delay(s.id)
        except Exception as e:
            logger.warning(f'重跑每日总结分析失败: {e}')
            return Response({'error': f'触发失败：{e}'}, status=500)
        return Response({'encrypt': True, 'data': encrypt_data({'ok': True})})

    @action(detail=True, methods=['post'])
    def share(self, request, pk=None):
        """将工作总结及其分析结果以私聊卡片形式分享给指定用户（对方点击卡片直达总结详情）"""
        import json as _json
        from chat.models import ChatRoom, Message
        user = request.user
        s = DailyWorkSummary.objects.select_related('user').filter(id=pk).first()
        if not s:
            return Response({'error': '工作总结不存在'}, status=404)
        if not self._can_view_summary(request.user, s):
            return Response({'error': '无权分享该工作总结'}, status=403)
        target_id = request.data.get('target_user_id')
        if not target_id:
            return Response({'error': '请选择要分享的用户'}, status=400)
        try:
            from accounts.models import CustomUser
            target = CustomUser.objects.get(id=int(target_id), is_active=True)
        except (ValueError, TypeError, CustomUser.DoesNotExist):
            return Response({'error': '目标用户不存在'}, status=404)
        if target.id == user.id:
            return Response({'error': '不能分享给自己'}, status=400)
        # get-or-create 私聊房间（与审批私聊保持一致）
        room = ChatRoom.objects.filter(room_type='private', members=user).filter(members__id=target.id).first()
        if not room:
            room = ChatRoom.objects.create(room_type='private', name='', creator=user)
            room.members.add(user, target)
        card_data = {
            'summary_id': s.id,
            'user_name': s.user.real_name or s.user.username,
            'summary_date': str(s.summary_date),
            'position': s.position,
            'status': s.status,
            'content': (s.content or '')[:200],
            'analysis': (s.analysis_result or '')[:200],
            'sender_name': user.real_name or user.username,
            'sender_id': user.id,
        }
        preview_text = f"📄 {card_data['sender_name']} 分享了每日工作总结（{s.summary_date}）"
        message = Message.objects.create(
            chat_room=room, sender=user,
            content=_json.dumps(card_data, ensure_ascii=False),
            message_type='work_summary_card',
        )
        try:
            from asgiref.sync import async_to_sync
            from channels.layers import get_channel_layer
            async_to_sync(get_channel_layer().group_send)(
                f'chat_{room.id}',
                {
                    'type': 'chat_message',
                    'chat_room': room.id,
                    'message_id': str(message.id),
                    'sender': {
                        'id': user.id,
                        'username': user.username,
                        'real_name': user.real_name,
                        'avatar': user.avatar.url if getattr(user, 'avatar', None) else None,
                    },
                    'sender_id': user.id,
                    'sender_name': user.real_name or user.username,
                    'content': preview_text,
                    'message_type': 'work_summary_card',
                    'timestamp': message.timestamp.isoformat(),
                    'work_summary_data': card_data,
                }
            )
        except Exception as e:
            logger.warning(f'分享工作总结 WebSocket 广播失败: {e}')
        # 同步发一条工作通知，点击直达总结详情
        send_work_notification(
            target.id,
            '每日工作总结分享',
            f'{user.real_name or user.username} 给您分享了一条每日工作总结（{s.summary_date}），请查看',
            notification_type='work_summary',
            related_url=f'/oa/work-summary/?id={s.id}',
            extra_data={'summary_id': s.id, 'action': 'shared'},
        )
        return Response({'encrypt': True, 'data': encrypt_data({'success': True, 'message': '已分享'})})

    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """将工作总结及分析结果导出为真实 PDF 文件（非打印）"""
        from django.http import HttpResponse
        from urllib.parse import quote
        s = DailyWorkSummary.objects.select_related('user').filter(id=pk).first()
        if not s:
            return Response({'error': '工作总结不存在'}, status=404)
        if not self._can_view_summary(request.user, s):
            return Response({'error': '无权导出该工作总结'}, status=403)
        pdf_bytes = _build_summary_pdf(s, self._primary_dept_name(s.user))
        if isinstance(pdf_bytes, dict) and pdf_bytes.get('error'):
            return Response(pdf_bytes, status=500)
        filename = f"每日工作总结_{s.summary_date}_{s.user.real_name or s.user.username}.pdf"
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(filename)}'
        return resp

    def _config_data(self, request, cfg):
        from utils.ark_llm import ARK_MODEL_PRESETS
        try:
            cfg.ensure_today()
        except Exception:
            pass
        return {
            'enabled': cfg.enabled,
            'model_id': cfg.model_id,
            'effective_model': cfg.effective_model(),
            'presets': ARK_MODEL_PRESETS,
            'is_super_admin': request.user.user_type == 'super_admin',
            'updated_at': cfg.updated_at.isoformat() if cfg.updated_at else None,
            'limit_enabled': cfg.limit_enabled,
            'daily_call_limit': cfg.daily_call_limit,
            'daily_cost_limit': float(cfg.daily_cost_limit or 0),
            'cost_per_1k_tokens': float(cfg.cost_per_1k_tokens or 0),
            'today_call_count': cfg.today_call_count,
            'today_cost': float(cfg.today_cost or 0),
            'mask_sensitive': cfg.mask_sensitive,
            'scope_type': cfg.scope_type,
            'scope_value': list(cfg.scope_value or []),
            'scope_users_info': self._scope_users_info(cfg),
        }

    def _scope_users_info(self, cfg):
        """灰度范围-指定用户：解析用户ID为姓名/头像，供前端标签展示"""
        if (cfg.scope_type or 'all') != 'users':
            return []
        ids = [x for x in (cfg.scope_value or [])]
        if not ids:
            return []
        from accounts.models import CustomUser
        try:
            us = CustomUser.objects.filter(id__in=ids)
            return [{'id': u.id, 'name': u.real_name or u.username,
                     'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else ''} for u in us]
        except Exception:
            return []

    @action(detail=False, methods=['get', 'post'])
    def config(self, request):
        """每日总结大模型配置：GET 读取（登录即可）；POST 保存（仅超级管理员）"""
        from .models import WorkSummaryConfig
        tenant = self._tenant(request)
        if not tenant:
            from utils.ark_llm import ARK_MODEL_PRESETS
            return Response({'encrypt': True, 'data': encrypt_data({
                'enabled': True, 'model_id': '',
                'effective_model': settings.ARK_MODEL or 'doubao-seed-1-6-250615',
                'presets': ARK_MODEL_PRESETS,
                'is_super_admin': request.user.user_type == 'super_admin',
                'updated_at': None,
            })})
        if request.method == 'POST':
            if request.user.user_type != 'super_admin':
                return Response({'error': '仅超级管理员可配置'}, status=403)
            cfg = WorkSummaryConfig.get_config(tenant)
            if 'enabled' in request.data:
                cfg.enabled = bool(request.data.get('enabled'))
            model_id = (request.data.get('model_id') or '').strip()
            if model_id:
                cfg.model_id = model_id[:100]
            elif 'model_id' in request.data:
                cfg.model_id = ''
            # —— 风险管控：调用限额 / 成本 / 数据脱敏 / 灰度范围 ——
            for f in ('limit_enabled', 'mask_sensitive'):
                if f in request.data:
                    setattr(cfg, f, bool(request.data.get(f)))
            for f in ('daily_call_limit', 'daily_cost_limit', 'cost_per_1k_tokens'):
                if f in request.data and request.data.get(f) not in (None, ''):
                    try:
                        setattr(cfg, f, request.data.get(f))
                    except Exception:
                        pass
            scope_type = (request.data.get('scope_type') or 'all').strip()
            if scope_type in dict(WorkSummaryConfig.SCOPE_CHOICES):
                cfg.scope_type = scope_type
            if 'scope_value' in request.data:
                sv = request.data.get('scope_value')
                cfg.scope_value = sv if isinstance(sv, list) else []
            cfg.updated_by = request.user
            cfg.save()
            logger.info(f'超管 {request.user} 更新每日总结模型配置: enabled={cfg.enabled}, model={cfg.model_id or "默认"}')
            return Response({'encrypt': True, 'data': encrypt_data(self._config_data(request, cfg))})
        cfg = WorkSummaryConfig.get_config(tenant)
        return Response({'encrypt': True, 'data': encrypt_data(self._config_data(request, cfg))})

    @action(detail=False, methods=['get'])
    def members(self, request):
        """当前企业成员列表，供范围分析选择员工（部门负责人/超管）"""
        if not self._can_view_all(request.user):
            return Response({'error': '仅部门负责人或超级管理员可查看'}, status=403)
        from accounts.models import CustomUser
        from django.db.models import Q as QF
        tenant = self._tenant(request)
        keyword = request.query_params.get('search', '').strip()
        qs = CustomUser.objects.filter(is_active=True)
        if tenant:
            ids = [tenant.id]
            try:
                ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
            except Exception:
                pass
            qs = qs.filter(tenant_memberships__tenant_id__in=ids, tenant_memberships__is_active=True)
        if keyword:
            qs = qs.filter(QF(real_name__icontains=keyword) | QF(username__icontains=keyword))
        # 部门负责人（非超管）只能选择自己负责部门（含子部门）成员
        if request.user.user_type != 'super_admin':
            scoped = self._scoped_user_ids(request.user)
            qs = qs.filter(id__in=scoped) if scoped else qs.none()
        qs = qs.distinct().order_by('real_name', 'username')[:100]
        ids = [u.id for u in qs]
        from org.models import UserDepartment
        dept_map = dict(UserDepartment.objects.filter(user_id__in=ids, is_primary=True).select_related('department').values_list('user_id', 'department__name'))
        results = [{
            'id': u.id,
            'name': u.real_name or u.username,
            'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
            'position': u.position or '',
            'department_name': dept_map.get(u.id) or '',
        } for u in qs]
        return Response({'encrypt': True, 'data': encrypt_data({'results': results})})

    def _range_data(self, a):
        return {
            'id': a.id,
            'target_user': a.target_user_id,
            'target_name': a.target_user.real_name or a.target_user.username,
            'target_avatar': a.target_user.get_avatar_url() if hasattr(a.target_user, 'get_avatar_url') else '',
            'target_position': a.target_user.position or '',
            'requester': a.requester_id,
            'requester_name': (a.requester.real_name or a.requester.username) if a.requester_id else '',
            'date_from': str(a.date_from),
            'date_to': str(a.date_to),
            'summary_count': a.summary_count,
            'status': a.status,
            'analysis_result': a.analysis_result,
            'prompt_text': a.prompt_text,
            'error_message': a.error_message,
            'created_at': a.created_at.isoformat(),
            'analyzed_at': a.analyzed_at.isoformat() if a.analyzed_at else None,
        }

    @action(detail=False, methods=['get', 'post'])
    def range(self, request):
        """范围分析：GET 我的批量分析记录列表；POST 触发指定员工+日期范围的批量分析"""
        from .models import WorkSummaryRangeAnalysis
        if not self._can_view_all(request.user):
            return Response({'error': '仅部门负责人或超级管理员可查看'}, status=403)
        if request.method == 'POST':
            from accounts.models import CustomUser
            from datetime import date as dt_date
            user_id = request.data.get('user_id')
            date_from = request.data.get('date_from', '').strip()
            date_to = request.data.get('date_to', '').strip()
            if not user_id:
                return Response({'error': '请选择要分析的员工'}, status=400)
            try:
                target = CustomUser.objects.get(id=int(user_id), is_active=True)
            except (ValueError, TypeError, CustomUser.DoesNotExist):
                return Response({'error': '员工不存在'}, status=400)
            tenant = self._tenant(request)
            if tenant:
                ids = [tenant.id]
                try:
                    ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
                except Exception:
                    pass
                if not target.tenant_memberships.filter(tenant_id__in=ids, is_active=True).exists():
                    return Response({'error': '该员工不属于当前企业'}, status=400)
            # 部门负责人（非超管）只能分析自己负责部门成员
            if request.user.user_type != 'super_admin':
                scoped = self._scoped_user_ids(request.user)
                if target.id not in scoped:
                    return Response({'error': '该员工不属于您负责的部门，无权分析'}, status=403)
            try:
                d_from = dt_date.fromisoformat(date_from)
                d_to = dt_date.fromisoformat(date_to)
            except (ValueError, TypeError):
                return Response({'error': '日期格式错误'}, status=400)
            if d_from > d_to:
                return Response({'error': '开始日期不能晚于结束日期'}, status=400)
            from .models import WorkSummaryConfig
            cfg = WorkSummaryConfig.get_config(tenant)
            status = 'pending' if cfg.enabled else 'disabled'
            if status == 'pending':
                if not cfg.in_scope(target):
                    status = 'not_allowed'
                else:
                    from .tasks import check_summary_quota
                    blocked, _reason = check_summary_quota(cfg)
                    if blocked:
                        status = 'limited'
            a = WorkSummaryRangeAnalysis.objects.create(
                requester=request.user, target_user=target, tenant=tenant,
                date_from=d_from, date_to=d_to, status=status)
            if status == 'pending':
                try:
                    from .tasks import analyze_work_summary_range_task
                    analyze_work_summary_range_task.delay(a.id)
                except Exception as e:
                    logger.warning(f'触发每日总结范围分析失败: {e}')
            return Response({'encrypt': True, 'data': encrypt_data(self._range_data(a))}, status=201)
        # GET 列表：超管可查看本企业（含子公司）全部批量分析记录；部门负责人/普通查看自己发起的
        qs = WorkSummaryRangeAnalysis.objects.select_related('target_user', 'requester').order_by('-created_at')
        if request.user.user_type == 'super_admin':
            tenant = self._tenant(request)
            if tenant:
                tids = [tenant.id]
                try:
                    tids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
                except Exception:
                    pass
                qs = qs.filter(tenant_id__in=tids)
        else:
            qs = qs.filter(requester=request.user)
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        items = qs[(page - 1) * page_size:page * page_size]
        return Response({'encrypt': True, 'data': encrypt_data({
            'results': [self._range_data(x) for x in items],
            'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages,
        })})

    @action(detail=True, methods=['get'])
    def range_detail(self, request, pk=None):
        """范围分析详情（供前端轮询流式结果）"""
        from .models import WorkSummaryRangeAnalysis
        a = WorkSummaryRangeAnalysis.objects.select_related('target_user').filter(id=pk).first()
        if not a:
            return Response({'error': '范围分析记录不存在'}, status=404)
        if a.requester_id != request.user.id and not self._can_view_all(request.user):
            return Response({'error': '无权查看'}, status=403)
        return Response({'encrypt': True, 'data': encrypt_data(self._range_data(a))})


class WorkCalendarViewSet(viewsets.ViewSet):
    """工作日历：按日聚合工作事件 + 每日工作汇总通知配置"""
    permission_classes = [permissions.IsAuthenticated]

    # 网盘/协作文档操作 → (事件类型, 图标, 动作文案)
    _CLOUD_OP_EVENT = {
        'upload': ('cloud', 'fas fa-upload', '上传文件'),
        'download': ('cloud', 'fas fa-download', '下载文件'),
        'preview': ('cloud', 'fas fa-eye', '预览文件'),
        'create': ('cloud', 'fas fa-file-alt', '新建文件'),
        'update': ('cloud', 'fas fa-edit', '更新文件'),
        'rename': ('cloud', 'fas fa-i-cursor', '重命名文件'),
        'move': ('cloud', 'fas fa-folder-open', '移动文件'),
        'batch_move': ('cloud', 'fas fa-folder-open', '批量移动'),
        'delete': ('cloud', 'fas fa-trash', '删除文件'),
        'restore': ('cloud', 'fas fa-trash-restore', '恢复文件'),
        'permanent_delete': ('cloud', 'fas fa-trash', '永久删除'),
        'batch_delete': ('cloud', 'fas fa-trash', '批量删除'),
        'empty_trash': ('cloud', 'fas fa-trash', '清空回收站'),
        'share': ('cloud', 'fas fa-share-alt', '分享文件'),
        'unshare': ('cloud', 'fas fa-ban', '取消分享'),
        'share_access': ('cloud', 'fas fa-eye', '访问分享'),
        'save_from_share': ('cloud', 'fas fa-cloud-download-alt', '从分享保存'),
        'save_from_chat': ('cloud', 'fas fa-cloud-upload-alt', '从聊天保存'),
        'sync_to_cloud': ('cloud', 'fas fa-cloud-upload-alt', '同步到云'),
        'edit_save': ('doc', 'fas fa-edit', '编辑保存'),
        'add_collaborator': ('cloud', 'fas fa-user-plus', '添加协作者'),
        'remove_collaborator': ('cloud', 'fas fa-user-minus', '移除协作者'),
        'update_collaborator': ('cloud', 'fas fa-user-cog', '更新协作者'),
        'update_collaboration_status': ('doc', 'fas fa-users', '协作状态'),
        'restore_version': ('doc', 'fas fa-history', '恢复版本'),
        'version_download': ('cloud', 'fas fa-download', '下载版本'),
        'add_folder_member': ('cloud', 'fas fa-user-plus', '添加文件夹成员'),
        'remove_folder_member': ('cloud', 'fas fa-user-minus', '移除文件夹成员'),
    }
    # 组织架构操作 → (图标, 动作文案)
    _ORG_ACTION_MAP = {
        'create_dept': ('fas fa-plus', '创建部门'),
        'update_dept': ('fas fa-edit', '修改部门'),
        'delete_dept': ('fas fa-trash', '删除部门'),
        'move_dept': ('fas fa-arrows-alt', '移动部门'),
        'add_member': ('fas fa-user-plus', '添加成员'),
        'remove_member': ('fas fa-user-minus', '移除成员'),
        'set_leader': ('fas fa-user-tie', '设置负责人'),
        'switch_tenant': ('fas fa-exchange-alt', '切换企业'),
    }

    def _tenant(self, request):
        return getattr(request, 'tenant', None) or request.user.get_active_tenant()

    def _resolve_target_user(self, request):
        """超管可通过 user_id 指定查看某成员工作日历；否则返回当前用户"""
        user = request.user
        user_id = request.query_params.get('user_id', '').strip()
        if not user_id or user.user_type != 'super_admin':
            return user
        try:
            from accounts.models import CustomUser
            tenant = self._tenant(request)
            target = CustomUser.objects.get(id=int(user_id), is_active=True)
            if tenant:
                ids = [tenant.id]
                try:
                    ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
                except Exception:
                    pass
                if not target.tenant_memberships.filter(tenant_id__in=ids, is_active=True).exists():
                    return user
            return target
        except Exception:
            return user

    def _summary(self, request, target=None):
        from .tasks import compute_user_work_summary
        u = target or request.user
        return compute_user_work_summary(u, self._tenant(request))

    def _target_info(self, target):
        return {'id': target.id, 'name': target.real_name or target.username,
                'avatar': target.get_avatar_url() if hasattr(target, 'get_avatar_url') else ''}

    def _primary_dept_name(self, user):
        try:
            from org.models import UserDepartment
            ud = UserDepartment.objects.filter(user=user, is_primary=True).select_related('department').first()
            if ud and ud.department:
                return ud.department.name
        except Exception:
            pass
        return (user.department.name if user.department else '') or ''

    def _user_brief(self, user):
        return {
            'name': user.real_name or user.username,
            'avatar': user.get_avatar_url() if hasattr(user, 'get_avatar_url') else '',
            'department': self._primary_dept_name(user),
            'position': user.position or '',
        }

    def list(self, request):
        """月度汇总：当前待办 + 该月每日事件计数"""
        now = timezone.localtime()
        year = int(request.query_params.get('year', now.year))
        month = int(request.query_params.get('month', now.month))
        target = self._resolve_target_user(request)
        pending = self._summary(request, target)
        days = self._month_days(target, year, month)
        return Response({'encrypt': True, 'data': encrypt_data({
            'pending': pending, 'year': year, 'month': month, 'days': days,
            'target_user': self._target_info(target),
        })})

    def _month_days(self, user, year, month):
        import calendar as cal
        from datetime import date
        from tasks.models import Task
        from cloud.models import CloudFile, FileOperationLog
        from org.models import OrgChangeLog
        from .models import ApprovalRequest, SubsidyApplication, SubsidyWithdrawal, AttendanceRecord
        _, last = cal.monthrange(year, month)
        start, end = date(year, month, 1), date(year, month, last)
        days = {}
        for d in range(1, last + 1):
            days[f'{year:04d}-{month:02d}-{d:02d}'] = {'approvals': 0, 'invoices': 0, 'withdrawals': 0, 'tasks': 0, 'docs': 0, 'cloud': 0, 'org': 0, 'work_summary': 0, 'clock_in': None, 'clock_out': None}

        def _bump(qs, key):
            for dt in qs:
                ds = dt.strftime('%Y-%m-%d')
                if ds in days:
                    days[ds][key] += 1

        _bump(ApprovalRequest.objects.filter(applicant=user, created_at__date__range=[start, end]).values_list('created_at', flat=True), 'approvals')
        _bump(DailyWorkSummary.objects.filter(user=user, summary_date__range=[start, end]).values_list('summary_date', flat=True), 'work_summary')
        _bump(SubsidyApplication.objects.filter(applicant=user, created_at__date__range=[start, end]).values_list('created_at', flat=True), 'invoices')
        _bump(SubsidyWithdrawal.objects.filter(user=user, requested_at__date__range=[start, end]).values_list('requested_at', flat=True), 'withdrawals')
        _bump(Task.objects.filter(Q(assignee=user) | Q(creator=user), status='done', updated_at__date__range=[start, end]).values_list('updated_at', flat=True), 'tasks')
        _bump(CloudFile.objects.filter(owner=user, is_document=True, created_at__date__range=[start, end]).values_list('created_at', flat=True), 'docs')
        # 网盘/协作文档操作 + 组织架构操作（按操作者归属）
        _bump(FileOperationLog.objects.filter(user=user, created_at__date__range=[start, end]).values_list('created_at', flat=True), 'cloud')
        _bump(OrgChangeLog.objects.filter(operator=user, created_at__date__range=[start, end]).values_list('created_at', flat=True), 'org')
        for r in AttendanceRecord.objects.filter(user=user, date__range=[start, end]).order_by('date', 'clock_time'):
            ds = r.date.strftime('%Y-%m-%d')
            t = timezone.localtime(r.clock_time).strftime('%H:%M')
            if r.clock_type == 'clock_in' and not days[ds]['clock_in']:
                days[ds]['clock_in'] = t
            elif r.clock_type == 'clock_out':
                days[ds]['clock_out'] = t
        return days

    def day(self, request):
        """某日工作汇总 + 当日事件明细"""
        from datetime import date as dt_date
        date_str = request.query_params.get('date', '')
        try:
            d = dt_date.fromisoformat(date_str)
        except (ValueError, TypeError):
            return Response({'error': '日期格式错误'}, status=400)
        target = self._resolve_target_user(request)
        pending = self._summary(request, target)
        events = self._day_events(target, d)
        return Response({'encrypt': True, 'data': encrypt_data({
            'pending': pending, 'date': date_str, 'events': events,
            'target_user': self._target_info(target),
        })})

    def _day_events(self, user, d):
        from tasks.models import Task
        from cloud.models import CloudFile, FileOperationLog
        from org.models import OrgChangeLog
        from .models import ApprovalRequest, ApprovalLog, SubsidyApplication, SubsidyWithdrawal, AttendanceRecord
        events = []

        def fmt(iso):
            return timezone.localtime(iso).strftime('%H:%M') if iso else ''

        def add(etype, icon, title, t, url):
            events.append({'type': etype, 'icon': icon, 'title': title, 'time': t, 'url': url})

        for a in ApprovalRequest.objects.filter(applicant=user, created_at__date=d):
            add('approval', 'fas fa-check-double', f'提交审批：{a.title}', fmt(a.created_at), f'/oa/approval/?approval_id={a.id}')
        action_map = dict(ApprovalLog.ACTION_CHOICES)
        for lg in ApprovalLog.objects.filter(operator=user, created_at__date=d).select_related('request'):
            add('approval', 'fas fa-check-double', f'审批{action_map.get(lg.action, lg.action)}：{lg.request.title}', fmt(lg.created_at), f'/oa/approval/?approval_id={lg.request_id}')
        for s in SubsidyApplication.objects.filter(applicant=user, created_at__date=d):
            add('subsidy', 'fas fa-hand-holding-usd', f'提交补贴申领：{s.application_no}', fmt(s.created_at), '/oa/subsidy/')
        for s in SubsidyApplication.objects.filter(verified_by=user, verified_at__date=d):
            add('subsidy', 'fas fa-clipboard-check', f'核验补贴申领：{s.application_no}', fmt(s.verified_at), '/oa/subsidy-verify/')
        for w in SubsidyWithdrawal.objects.filter(user=user, requested_at__date=d):
            add('subsidy', 'fas fa-money-check-alt', f'发起提现：¥{float(w.amount):.2f}', fmt(w.requested_at), '/oa/subsidy/')
        for w in SubsidyWithdrawal.objects.filter(paid_by=user, paid_at__date=d):
            add('subsidy', 'fas fa-wallet', f'支付提现：¥{float(w.amount):.2f}', fmt(w.paid_at), '/oa/subsidy-pay/')
        for t in Task.objects.filter(Q(assignee=user) | Q(creator=user), status='done', updated_at__date=d):
            add('task', 'fas fa-tasks', f'完成任务：{t.title}', fmt(t.updated_at), f'/tasks/?task_id={t.id}')
        for c in CloudFile.objects.filter(owner=user, is_document=True, created_at__date=d):
            add('doc', 'fas fa-file-alt', f'新建协作文档：{c.name}', fmt(c.created_at), '/cloud/')
        # 网盘/协作文档操作：谁操作记谁（协作者编辑归属到编辑者）
        for lg in FileOperationLog.objects.filter(user=user, created_at__date=d).order_by('created_at'):
            ev = self._CLOUD_OP_EVENT.get(lg.operation)
            if ev:
                title = lg.description or ev[2]
                add(ev[0], ev[1], title, fmt(lg.created_at), '/cloud/')
        # 组织架构操作：谁操作记谁
        for oc in OrgChangeLog.objects.filter(operator=user, created_at__date=d).select_related('department'):
            icon, label = self._ORG_ACTION_MAP.get(oc.action, ('fas fa-sitemap', oc.get_action_display() or oc.action))
            dept_name = oc.department.name if oc.department else '企业'
            add('org', icon, f'{label}：{dept_name}', fmt(oc.created_at), '/org/')
        for r in AttendanceRecord.objects.filter(user=user, date=d).order_by('clock_time'):
            add('attendance', 'fas fa-clock', f'{r.get_clock_type_display()}打卡', fmt(r.clock_time), '/oa/attendance/')
        for ws in DailyWorkSummary.objects.filter(user=user, summary_date=d).order_by('created_at'):
            add('work_summary', 'fas fa-file-signature', f'每日工作总结（{ws.position or "未填职位"}）', fmt(ws.created_at), f'/oa/work-summary/?id={ws.id}')
        events.sort(key=lambda e: e['time'], reverse=True)
        return events

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """时间线统计：某时间范围内按日/月聚合的处理量（审批/核验发票/支付提现/任务 + 网盘/组织操作）"""
        from datetime import date as dt_date, timedelta
        from .models import ApprovalAssignee, SubsidyApplication, SubsidyWithdrawal
        from tasks.models import Task
        from cloud.models import FileOperationLog
        from org.models import OrgChangeLog
        target = self._resolve_target_user(request)
        now_date = timezone.localdate()
        start = end = None
        try:
            start_s = request.query_params.get('start', '').strip()
            if start_s:
                start = dt_date.fromisoformat(start_s)
            end_s = request.query_params.get('end', '').strip()
            if end_s:
                end = dt_date.fromisoformat(end_s)
        except ValueError:
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD'}, status=400)
        if start is None:
            start = now_date - timedelta(days=29)
        if end is None:
            end = now_date
        if start > end:
            start, end = end, start
        days_count = (end - start).days + 1

        if days_count <= 62:
            buckets = [start + timedelta(days=i) for i in range(days_count)]
            labels = [b.strftime('%m-%d') for b in buckets]
            key_fn = lambda v: v.date()
        else:
            cur = start.replace(day=1)
            buckets = []
            while cur <= end:
                buckets.append(cur)
                nxt = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)
                cur = nxt
            labels = [b.strftime('%Y-%m') for b in buckets]
            key_fn = lambda v: v.date().replace(day=1)
        bucket_map = {b: i for i, b in enumerate(buckets)}

        def _arr():
            return [0] * len(buckets)

        approvals, invoices, withdrawals = _arr(), _arr(), _arr()
        tasks, cloud, org = _arr(), _arr(), _arr()

        def _count(qs, arr):
            for v in qs:
                idx = bucket_map.get(key_fn(v))
                if idx is not None:
                    arr[idx] += 1

        _count(ApprovalAssignee.objects.filter(
            user=target, status__in=['approved', 'rejected', 'deferred'],
            operated_at__isnull=False, operated_at__date__gte=start, operated_at__date__lte=end,
        ).values_list('operated_at', flat=True), approvals)
        _count(SubsidyApplication.objects.filter(
            verified_by=target, verified_at__isnull=False, verified_at__date__gte=start, verified_at__date__lte=end,
        ).values_list('verified_at', flat=True), invoices)
        _count(SubsidyWithdrawal.objects.filter(
            paid_by=target, paid_at__isnull=False, paid_at__date__gte=start, paid_at__date__lte=end,
        ).values_list('paid_at', flat=True), withdrawals)
        _count(Task.objects.filter(
            Q(assignee=target) | Q(creator=target), status='done',
            updated_at__date__gte=start, updated_at__date__lte=end,
        ).values_list('updated_at', flat=True), tasks)
        _count(FileOperationLog.objects.filter(
            user=target, created_at__date__gte=start, created_at__date__lte=end,
        ).values_list('created_at', flat=True), cloud)
        _count(OrgChangeLog.objects.filter(
            operator=target, created_at__date__gte=start, created_at__date__lte=end,
        ).values_list('created_at', flat=True), org)

        return Response({'encrypt': True, 'data': encrypt_data({
            'start': start.isoformat(), 'end': end.isoformat(),
            'labels': labels, 'approvals': approvals, 'invoices': invoices,
            'withdrawals': withdrawals, 'tasks': tasks, 'cloud': cloud, 'org': org,
            'target_user': self._target_info(target),
        })})

    @action(detail=False, methods=['get'])
    def org_activity(self, request):
        """成员关系网络 + 活跃度聚合：静态组织架构 × 动态行为数据（聊天/审批/考勤/总结/任务/网盘/文档）
        GET /api/oa/work-calendar/org-activity/?start=YYYY-MM-DD&end=YYYY-MM-DD
        返回 org_tree（架构树）、network（关系图节点/边）、members（雷达选择）、activity（每人分维度活跃度）"""
        from datetime import date as dt_date, timedelta
        from accounts.models import CustomUser, Tenant, Department as OrgDept
        from org.models import UserDepartment
        from chat.models import Message, ChatRoom
        from tasks.models import Task
        from cloud.models import FileOperationLog
        from .models import ApprovalRequest, ApprovalAssignee, AttendanceRecord, DailyWorkSummary

        now_date = timezone.localdate()
        start = end = None
        try:
            start_s = request.query_params.get('start', '').strip()
            if start_s:
                start = dt_date.fromisoformat(start_s)
            end_s = request.query_params.get('end', '').strip()
            if end_s:
                end = dt_date.fromisoformat(end_s)
        except ValueError:
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD'}, status=400)
        if start is None:
            start = now_date - timedelta(days=29)
        if end is None:
            end = now_date
        if start > end:
            start, end = end, start

        tenant = self._tenant(request)
        tenant_ids = [tenant.id] if tenant else []
        if tenant:
            try:
                tenant_ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
            except Exception:
                pass

        members_qs = CustomUser.objects.filter(is_active=True)
        if tenant_ids:
            members_qs = members_qs.filter(
                tenant_memberships__tenant_id__in=tenant_ids,
                tenant_memberships__is_active=True,
            ).distinct()
        members = list(members_qs)
        user_ids = [u.id for u in members]
        id_set = set(user_ids)
        empty = {'encrypt': True, 'data': encrypt_data({
            'org_tree': [], 'network': {'nodes': [], 'links': []},
            'members': [], 'activity': {}, 'range': {'start': start.isoformat(), 'end': end.isoformat()},
        })}
        if not user_ids:
            return Response(empty)

        # 成员基础信息 + 主部门
        ud_map = {}
        for ud in UserDepartment.objects.filter(user_id__in=user_ids, is_primary=True).select_related('department'):
            ud_map[ud.user_id] = {'dept_id': ud.department_id,
                                  'dept_name': ud.department.name if ud.department else ''}
        member_info = {}
        for u in members:
            d = ud_map.get(u.id, {})
            member_info[u.id] = {
                'id': u.id, 'name': u.real_name or u.username,
                'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
                'department': d.get('dept_name', ''), 'dept_id': d.get('dept_id'),
                'position': u.position or '',
            }

        act = {uid: {'chat': 0, 'approval': 0, 'attendance': 0, 'summary': 0,
                     'task': 0, 'cloud': 0, 'doc': 0} for uid in user_ids}

        def _add(qs, key):
            for uid in qs:
                if uid in act:
                    act[uid][key] += 1

        _add(Message.objects.filter(sender_id__in=user_ids, timestamp__date__range=[start, end], is_deleted=False)
             .values_list('sender_id', flat=True), 'chat')
        _add(ApprovalRequest.objects.filter(applicant_id__in=user_ids, created_at__date__range=[start, end])
             .values_list('applicant_id', flat=True), 'approval')
        _add(ApprovalAssignee.objects.filter(user_id__in=user_ids, operated_at__isnull=False,
                                             operated_at__date__range=[start, end])
             .values_list('user_id', flat=True), 'approval')
        _add(AttendanceRecord.objects.filter(user_id__in=user_ids, date__range=[start, end])
             .values_list('user_id', flat=True), 'attendance')
        _add(DailyWorkSummary.objects.filter(user_id__in=user_ids, summary_date__range=[start, end])
             .values_list('user_id', flat=True), 'summary')
        _add(Task.objects.filter(creator_id__in=user_ids, created_at__date__range=[start, end])
             .values_list('creator_id', flat=True), 'task')
        _add(Task.objects.filter(assignee_id__in=user_ids, status='done', updated_at__date__range=[start, end])
             .values_list('assignee_id', flat=True), 'task')
        _add(FileOperationLog.objects.filter(user_id__in=user_ids, created_at__date__range=[start, end])
             .exclude(operation='edit_save').values_list('user_id', flat=True), 'cloud')
        _add(FileOperationLog.objects.filter(user_id__in=user_ids, operation='edit_save',
                                             created_at__date__range=[start, end])
             .values_list('user_id', flat=True), 'doc')

        # —— 两两互动边：每次互动生成一条线段（含类型/时间/标题），逐条显示、点击可看该次互动 ——
        from django.db.models import Max as MaxAgg
        MAX_EDGE_PER_PAIR = 15
        MAX_EDGES = 600
        edges = []
        pair_count = {}

        def _push_edge(a, b, etype, time_iso, title):
            if a in id_set and b in id_set and a != b and len(edges) < MAX_EDGES:
                key = (a, b) if a < b else (b, a)
                n = pair_count.get(key, 0)
                if n >= MAX_EDGE_PER_PAIR:
                    return
                pair_count[key] = n + 1
                edges.append({'source': a, 'target': b, 'type': etype,
                              'time': time_iso, 'title': title or ''})

        # 私聊：每条消息一条线段
        room_pair = {}
        for rm in ChatRoom.objects.filter(room_type='private', members__id__in=user_ids).distinct().prefetch_related('members'):
            mids = [m.id for m in rm.members.all() if m.id in id_set]
            if len(mids) == 2:
                room_pair[rm.id] = (mids[0], mids[1])
        if room_pair:
            for rid, mts in Message.objects.filter(chat_room_id__in=list(room_pair.keys()),
                                                   timestamp__date__range=[start, end], is_deleted=False) \
                    .order_by('-timestamp') \
                    .values_list('chat_room_id', 'timestamp')[:2000]:
                p = room_pair[rid]
                _push_edge(p[0], p[1], 'chat', mts.isoformat(), '私聊消息')
                if len(edges) >= MAX_EDGES:
                    break
        # 审批：发起人 → 审批人（每次审批一条线段）
        for asg in ApprovalAssignee.objects.filter(
                user_id__in=user_ids, status__in=['approved', 'rejected'],
                operated_at__isnull=False, operated_at__date__range=[start, end],
        ).select_related('node__request__applicant'):
            _push_edge(asg.node.request.applicant_id, asg.user_id, 'approval',
                       timezone.localtime(asg.operated_at).isoformat(), asg.node.request.title)
        # 任务：创建人 → 负责人（每次指派一条线段）
        for t in Task.objects.filter(created_at__date__range=[start, end],
                                     creator_id__in=user_ids, assignee_id__in=user_ids) \
                .only('creator_id', 'assignee_id', 'created_at', 'title'):
            _push_edge(t.creator_id, t.assignee_id, 'task',
                       timezone.localtime(t.created_at).isoformat(), t.title)
        # 协作文档：同一文档多名编辑者两两协作（每次协作一条线段）
        edit_rows = FileOperationLog.objects.filter(operation='edit_save', user_id__in=user_ids,
                                                    created_at__date__range=[start, end]) \
            .values('file_id', 'user_id').annotate(last=MaxAgg('created_at'))
        file_editors = {}
        file_time = {}
        for row in edit_rows:
            fid, uid = row['file_id'], row['user_id']
            if fid is None:
                continue
            file_editors.setdefault(fid, set()).add(uid)
            if fid not in file_time or (row['last'] and row['last'] > file_time[fid]):
                file_time[fid] = row['last']
        processed_files = 0
        for fid, uid_set in file_editors.items():
            ul = list(uid_set)
            if len(ul) >= 2:
                processed_files += 1
                if processed_files > 60:
                    break
                t_iso = timezone.localtime(file_time[fid]).isoformat() if file_time.get(fid) else None
                for i in range(len(ul)):
                    for j in range(i + 1, len(ul)):
                        _push_edge(ul[i], ul[j], 'doc', t_iso, '协作文档')
        # 网盘分享：分享者 → 指定用户（每次分享一条线段）
        from cloud.models import FileShare
        share_qs = FileShare.objects.filter(owner_id__in=user_ids, is_active=True,
                                            created_at__date__range=[start, end]) \
            .prefetch_related('allowed_users')
        for share in share_qs:
            for au in share.allowed_users.all():
                if au.id in id_set and au.id != share.owner_id:
                    _push_edge(share.owner_id, au.id, 'share',
                               timezone.localtime(share.created_at).isoformat(), '网盘分享')

        # —— 组织架构树：根=当前企业/集团名，部门按类型着色（公司/子公司类型区别于普通部门，不单独列出） ——
        depts = list(OrgDept.objects.filter(tenant_id__in=tenant_ids))
        dept_by_id = {d.id: d for d in depts}
        children_of = {}
        for d in depts:
            children_of.setdefault(d.parent_id, []).append(d)

        def _build_dept(did):
            d = dept_by_id[did]
            node = {'name': d.name, 'type': 'dept', 'dept_type': d.department_type, 'children': []}
            for c in children_of.get(did, []):
                node['children'].append(_build_dept(c.id))
            for uid, info in member_info.items():
                if info['dept_id'] == did:
                    a = act[uid]
                    node['children'].append({'name': info['name'], 'id': uid, 'type': 'member',
                                             'avatar': info['avatar'], 'value': sum(a.values())})
            return node

        root_children = []
        for c in children_of.get(None, []):
            root_children.append(_build_dept(c.id))
        # 未分组成员（无主部门）挂在根下
        orphans = [info for info in member_info.values() if info['dept_id'] not in dept_by_id]
        if orphans:
            root_children.append({'name': '未分组', 'type': 'virtual', 'children': [
                {'name': o['name'], 'id': o['id'], 'type': 'member', 'avatar': o['avatar'],
                 'value': sum(act[o['id']].values())} for o in orphans]})
        if not root_children:
            root_children.append({'name': '暂无部门', 'type': 'virtual', 'children': []})
        ttype_label = dict(Tenant.TENANT_TYPE_CHOICES).get(tenant.tenant_type, '企业') if tenant else '企业'
        org_tree = [{'name': (tenant.name if tenant else '企业'), 'type': 'tenant',
                     'label': ttype_label, 'children': root_children}]

        # —— 关系网络节点/边：展示全部成员，保证任意互动对均有节点可连线 ——
        all_nodes = []
        for uid in user_ids:
            a = act[uid]
            all_nodes.append({'id': uid, 'name': member_info[uid]['name'],
                              'avatar': member_info[uid]['avatar'],
                              'category': member_info[uid]['department'] or '未分组',
                              'position': member_info[uid]['position'],
                              'value': sum(a.values()), 'activity': a})
        all_nodes.sort(key=lambda n: n['value'], reverse=True)
        top_nodes = all_nodes
        top_ids = {n['id'] for n in top_nodes}
        # 逐条互动边：每条边=一次互动（含类型/时间/标题），仅保留两端都在展示节点的边
        links = [e for e in edges if e['source'] in top_ids and e['target'] in top_ids]

        return Response({'encrypt': True, 'data': encrypt_data({
            'org_tree': org_tree,
            'network': {'nodes': top_nodes, 'links': links, 'total_members': len(user_ids)},
            'members': [{'id': info['id'], 'name': info['name'], 'avatar': info['avatar'],
                         'department': info['department'], 'position': info['position']}
                        for info in member_info.values()],
            'activity': {str(uid): a for uid, a in act.items()},
            'range': {'start': start.isoformat(), 'end': end.isoformat()},
        })})

    @action(detail=False, methods=['get'])
    def member_search(self, request):
        """按姓名/账号搜索企业成员（含子企业），供个人雷达选择成员（所有登录用户可用）"""
        from accounts.models import CustomUser
        from django.db.models import Q
        tenant = self._tenant(request)
        tenant_ids = [tenant.id] if tenant else []
        if tenant:
            try:
                tenant_ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
            except Exception:
                pass
        kw = request.query_params.get('q', '').strip()
        qs = CustomUser.objects.filter(is_active=True)
        if tenant_ids:
            qs = qs.filter(tenant_memberships__tenant_id__in=tenant_ids,
                           tenant_memberships__is_active=True).distinct()
        if kw:
            qs = qs.filter(Q(real_name__icontains=kw) | Q(username__icontains=kw))
        qs = qs.order_by('real_name', 'username')[:30]
        uids = [u.id for u in qs]
        from org.models import UserDepartment
        dept_map = dict(UserDepartment.objects.filter(
            user_id__in=uids, is_primary=True).select_related('department')
            .values_list('user_id', 'department__name'))
        results = [{
            'id': u.id, 'name': u.real_name or u.username,
            'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
            'department': dept_map.get(u.id, ''), 'position': u.position or '',
        } for u in qs]
        return Response({'encrypt': True, 'data': encrypt_data({'results': results})})

    def _node_reach_times(self, request_ids):
        """节点到达时间：节点1=最近一次提交/重新提交时间；后续节点=上一节点最后一个操作时间。
        撤销后重新提交（re_edit）会删除并重建全部审批人与节点，
        因此"最近一次到达"以当前审批人记录 created_at 的最小值（最近一次提交时间）为准，
        而不是原始 request.created_at。"""
        from .models import ApprovalRequest, ApprovalNode
        result = {}
        if not request_ids:
            return result
        nodes = ApprovalNode.objects.filter(request_id__in=request_ids).prefetch_related('assignees')
        by_req = {}
        for n in nodes:
            by_req.setdefault(n.request_id, []).append(n)
        req_times = {r.id: r.created_at for r in ApprovalRequest.objects.filter(id__in=request_ids)}
        for rid, nl in by_req.items():
            ordered = sorted(nl, key=lambda n: n.order)
            # 最近一次提交/重新提交时间：当前审批人记录创建时间的最小值（撤销重提会删除重建，故为最近一次）
            submit_time = None
            for n in ordered:
                for a in n.assignees.all():
                    if submit_time is None or a.created_at < submit_time:
                        submit_time = a.created_at
            if submit_time is None:
                submit_time = req_times.get(rid)
            last_op_of_prev = None
            for node in ordered:
                arrival = submit_time if last_op_of_prev is None else last_op_of_prev
                if arrival:
                    result[node.id] = arrival
                lop = None
                for a in node.assignees.all():
                    if a.operated_at and (lop is None or a.operated_at > lop):
                        lop = a.operated_at
                if lop:
                    last_op_of_prev = lop
        return result

    @action(detail=False, methods=['get'])
    def approval_efficiency(self, request):
        """审批效率：审批到达自己节点 → 自己审批通过 花费的时间"""
        from datetime import date as dt_date, timedelta
        from .models import ApprovalAssignee
        target = self._resolve_target_user(request)
        now_date = timezone.localdate()
        start = end = None
        try:
            start_s = request.query_params.get('start', '').strip()
            if start_s:
                start = dt_date.fromisoformat(start_s)
            end_s = request.query_params.get('end', '').strip()
            if end_s:
                end = dt_date.fromisoformat(end_s)
        except ValueError:
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD'}, status=400)
        if start is None:
            start = now_date - timedelta(days=29)
        if end is None:
            end = now_date
        if start > end:
            start, end = end, start

        assignees = ApprovalAssignee.objects.filter(
            user=target, status='approved', operated_at__isnull=False,
            operated_at__date__gte=start, operated_at__date__lte=end,
        ).select_related('node__request__applicant').order_by('operated_at')
        assignee_list = list(assignees)
        request_ids = {a.node.request_id for a in assignee_list}
        reach = self._node_reach_times(request_ids)

        items = []
        for a in assignee_list:
            arrival = reach.get(a.node_id)
            operated = a.operated_at
            if not arrival or operated <= arrival:
                continue
            minutes = int((operated - arrival).total_seconds() // 60)
            items.append({
                'request_id': a.node.request_id,
                'title': a.node.request.title,
                'node_order': a.node.order,
                'arrival': arrival.isoformat(),
                'approved_at': operated.isoformat(),
                'minutes': minutes,
                'applicant': self._user_brief(a.node.request.applicant),
            })
        avg_min = max_min = 0
        if items:
            total_min = sum(i['minutes'] for i in items)
            avg_min = int(total_min / len(items))
            max_min = max(i['minutes'] for i in items)

        return Response({'encrypt': True, 'data': encrypt_data({
            'start': start.isoformat(), 'end': end.isoformat(),
            'count': len(items), 'avg_minutes': avg_min, 'max_minutes': max_min,
            'items': items, 'target_user': self._target_info(target),
        })})

    @action(detail=False, methods=['get'])
    def approval_leaderboard(self, request):
        """审批效率负责人排行榜：企业审批人在时间范围内按平均用时升序，含审批数量/总用时/平均用时"""
        from datetime import date as dt_date, timedelta
        from collections import defaultdict
        from .models import ApprovalAssignee
        from accounts.models import CustomUser
        from org.models import UserDepartment
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可查看'}, status=403)
        tenant = self._tenant(request)
        now_date = timezone.localdate()
        start = end = None
        try:
            start_s = request.query_params.get('start', '').strip()
            if start_s:
                start = dt_date.fromisoformat(start_s)
            end_s = request.query_params.get('end', '').strip()
            if end_s:
                end = dt_date.fromisoformat(end_s)
        except ValueError:
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD'}, status=400)
        if start is None:
            start = now_date - timedelta(days=29)
        if end is None:
            end = now_date
        if start > end:
            start, end = end, start

        qs = ApprovalAssignee.objects.filter(
            status='approved', operated_at__isnull=False,
            operated_at__date__gte=start, operated_at__date__lte=end,
        ).select_related('user', 'node__request')
        if tenant:
            ids = [tenant.id]
            try:
                ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
            except Exception:
                pass
            qs = qs.filter(user__tenant_memberships__tenant_id__in=ids, user__tenant_memberships__is_active=True)
        qs = qs.distinct()
        assignee_list = list(qs)
        request_ids = {a.node.request_id for a in assignee_list}
        reach = self._node_reach_times(request_ids)

        agg = defaultdict(lambda: {'count': 0, 'total': 0})
        for a in assignee_list:
            arrival = reach.get(a.node_id)
            if not arrival or a.operated_at <= arrival:
                continue
            mins = int((a.operated_at - arrival).total_seconds() // 60)
            agg[a.user_id]['count'] += 1
            agg[a.user_id]['total'] += mins

        user_ids = list(agg.keys())
        users = {u.id: u for u in CustomUser.objects.filter(id__in=user_ids)}
        dept_map = dict(UserDepartment.objects.filter(
            user_id__in=user_ids, is_primary=True).select_related('department').values_list('user_id', 'department__name'))

        results = []
        for uid, data in agg.items():
            u = users.get(uid)
            if not u:
                continue
            results.append({
                'user_id': uid,
                'name': u.real_name or u.username,
                'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
                'department': dept_map.get(uid) or (u.department.name if u.department else '') or '',
                'position': u.position or '',
                'count': data['count'],
                'total_minutes': data['total'],
                'avg_minutes': int(data['total'] / data['count']) if data['count'] else 0,
            })
        results.sort(key=lambda r: r['avg_minutes'])

        return Response({'encrypt': True, 'data': encrypt_data({
            'start': start.isoformat(), 'end': end.isoformat(),
            'results': results,
        })})

    @action(detail=False, methods=['get'])
    def work_summary_stats(self, request):
        """全员每日工作总结完成情况统计（仅超管）：按日趋势 / 部门 / 岗位 / 成员维度，用于监督总结完成与沉淀经验"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可查看'}, status=403)
        tenant = self._tenant(request)
        if not tenant:
            return Response({'error': '未找到所属企业'}, status=400)
        from datetime import date as dt_date, timedelta
        from accounts.models import CustomUser
        from org.models import UserDepartment
        from .models import DailyWorkSummary
        tenant_ids = [tenant.id]
        try:
            tenant_ids += [t.id for t in tenant.sub_tenants.filter(is_active=True)]
        except Exception:
            pass
        now_date = timezone.localdate()
        try:
            start_s = request.query_params.get('start', '').strip()
            end_s = request.query_params.get('end', '').strip()
            start = dt_date.fromisoformat(start_s) if start_s else (now_date - timedelta(days=29))
            end = dt_date.fromisoformat(end_s) if end_s else now_date
        except ValueError:
            return Response({'error': '日期格式错误，请使用 YYYY-MM-DD'}, status=400)
        if start > end:
            start, end = end, start
        dates = [start + timedelta(days=i) for i in range((end - start).days + 1)]
        total_days = len(dates)
        # 活跃成员（含子企业）
        members = list(CustomUser.objects.filter(
            is_active=True, tenant_memberships__tenant_id__in=tenant_ids, tenant_memberships__is_active=True,
        ).distinct())
        member_ids = [u.id for u in members]
        member_count = len(member_ids)
        # 主部门映射
        dept_map = dict(UserDepartment.objects.filter(
            user_id__in=member_ids, is_primary=True).select_related('department').values_list('user_id', 'department__name'))
        # 范围内全部总结（user_id -> 日期集合）
        submitted_dates = {}
        for r in DailyWorkSummary.objects.filter(
                user_id__in=member_ids, summary_date__range=[start, end]).values('user_id', 'summary_date'):
            submitted_dates.setdefault(r['user_id'], set()).add(r['summary_date'])
        # 每日趋势
        daily = []
        for d in dates:
            cnt = sum(1 for u in members if d in submitted_dates.get(u.id, set()))
            daily.append({'date': str(d), 'submitted': cnt, 'total': member_count,
                          'rate': round(cnt / max(member_count, 1) * 100, 1)})
        # 成员维度（完成率升序，便于监督落后者）
        by_member = []
        for u in members:
            s = len(submitted_dates.get(u.id, set()))
            missed = [d for d in dates if d not in submitted_dates.get(u.id, set())]
            by_member.append({
                'user_id': u.id, 'name': u.real_name or u.username,
                'avatar': u.get_avatar_url() if hasattr(u, 'get_avatar_url') else '',
                'department': dept_map.get(u.id) or '', 'position': u.position or '',
                'submitted': s, 'total': total_days,
                'rate': round(s / max(total_days, 1) * 100, 1),
                'missed_count': len(missed),
                'missed_days': [str(d) for d in missed][:10],
            })
        by_member.sort(key=lambda x: (x['rate'], x['name']))
        # 部门维度
        dept_stats = {}
        for u in members:
            dept = dept_map.get(u.id) or '未分组'
            d = dept_stats.setdefault(dept, {'member_count': 0, 'submitted_days': 0})
            d['member_count'] += 1
            d['submitted_days'] += len(submitted_dates.get(u.id, set()))
        by_department = [{'name': k, 'member_count': v['member_count'], 'submitted_days': v['submitted_days'],
                          'total_days': total_days,
                          'rate': round(v['submitted_days'] / max(v['member_count'] * total_days, 1) * 100, 1)}
                         for k, v in dept_stats.items()]
        by_department.sort(key=lambda x: x['rate'])
        # 岗位维度
        pos_stats = {}
        for u in members:
            pos = u.position or '未填职位'
            d = pos_stats.setdefault(pos, {'member_count': 0, 'submitted_days': 0})
            d['member_count'] += 1
            d['submitted_days'] += len(submitted_dates.get(u.id, set()))
        by_position = [{'name': k, 'member_count': v['member_count'], 'submitted_days': v['submitted_days'],
                        'total_days': total_days,
                        'rate': round(v['submitted_days'] / max(v['member_count'] * total_days, 1) * 100, 1)}
                       for k, v in pos_stats.items()]
        by_position.sort(key=lambda x: x['rate'])
        # 总体
        total_submitted_days = sum(len(submitted_dates.get(u.id, set())) for u in members)
        submitted_members = sum(1 for u in members if submitted_dates.get(u.id))
        return Response({'encrypt': True, 'data': encrypt_data({
            'range': {'start': str(start), 'end': str(end), 'days': total_days},
            'overview': {
                'member_count': member_count, 'submitted_members': submitted_members,
                'total_submitted_days': total_submitted_days,
                'overall_rate': round(total_submitted_days / max(member_count * total_days, 1) * 100, 1),
            },
            'daily': daily,
            'by_department': by_department,
            'by_position': by_position,
            'by_member': by_member[:200],
        })})

    # ===== 每日通知配置 =====
    def digest_config(self, request):
        """每日通知配置：GET 读取 / POST 保存（仅超级管理员）"""
        tenant = self._tenant(request)
        if request.method == 'POST':
            if request.user.user_type != 'super_admin':
                return Response({'error': '仅超级管理员可配置'}, status=403)
            from .models import DailyDigestConfig
            from datetime import time as dt_time
            send_time = request.data.get('send_time') or '09:00'
            try:
                parts = str(send_time).split(':')
                send_time_obj = dt_time(int(parts[0]), int(parts[1]), 0)
            except (ValueError, IndexError):
                return Response({'error': '发送时间格式错误，请使用 HH:MM'}, status=400)
            DailyDigestConfig.objects.update_or_create(
                tenant=tenant,
                defaults={
                    'enabled': bool(request.data.get('enabled', False)),
                    'send_time': send_time_obj,
                    'auto_send': bool(request.data.get('auto_send', False)),
                },
            )
            cfg = DailyDigestConfig.objects.filter(tenant=tenant).first()
            if cfg and cfg.enabled and cfg.auto_send:
                self._ensure_digest_periodic_task()
            return Response({'encrypt': True, 'data': encrypt_data(self._digest_config_data(tenant))})
        return Response({'encrypt': True, 'data': encrypt_data(self._digest_config_data(tenant))})

    def _digest_config_data(self, tenant):
        from .models import DailyDigestConfig
        cfg = DailyDigestConfig.objects.filter(tenant=tenant).first()
        return {
            'enabled': bool(cfg and cfg.enabled),
            'send_time': (cfg.send_time.strftime('%H:%M') if cfg and cfg.send_time else '09:00'),
            'auto_send': bool(cfg and cfg.auto_send),
        }

    def _ensure_digest_periodic_task(self):
        """确保 celery beat 有一个每5分钟轮询每日通知的定时任务（幂等）"""
        try:
            from django_celery_beat.models import PeriodicTask, CrontabSchedule
            ct, _ = CrontabSchedule.objects.get_or_create(
                minute='*/5', hour='*', day_of_week='*', day_of_month='*', month_of_year='*')
            pt, created = PeriodicTask.objects.get_or_create(
                name='run-daily-digest',
                defaults={'task': 'oa.tasks.run_daily_digest', 'crontab': ct, 'enabled': True},
            )
            if not created:
                pt.task = 'oa.tasks.run_daily_digest'
                pt.crontab = ct
                pt.enabled = True
                pt.save()
        except Exception as e:
            logger.warning(f'确保每日通知定时任务失败: {e}')

    def digest_send(self, request):
        """立即手动发送每日工作汇总给企业所有活跃用户（仅超级管理员）"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可操作'}, status=403)
        tenant = self._tenant(request)
        if not tenant:
            return Response({'error': '未找到所属企业'}, status=400)
        from accounts.models import CustomUser
        from .tasks import send_daily_digest_to_user
        users = CustomUser.objects.filter(
            tenant_memberships__tenant=tenant,
            tenant_memberships__is_active=True,
            is_active=True,
        ).distinct()
        count = 0
        for u in users:
            if send_daily_digest_to_user(u, tenant):
                count += 1
        return Response({'message': f'已发送 {count} 条每日工作汇总'})


class WorkNotificationViewSet(viewsets.ViewSet):
    """工作通知视图集"""
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        """通知列表（分页）"""
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        read_filter = request.query_params.get('read_filter', '').strip()
        qs = WorkNotification.objects.filter(recipient=request.user)
        if read_filter == 'unread':
            qs = qs.filter(is_read=False)
        elif read_filter == 'read':
            qs = qs.filter(is_read=True)
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        start = (page - 1) * page_size
        items = qs[start:start + page_size]
        results = [{
            'id': n.id, 'type': n.notification_type,
            'title': n.title, 'content': n.content,
            'related_url': n.related_url,
            'is_read': n.is_read,
            'created_at': n.created_at.isoformat() if n.created_at else '',
            'extra_data': n.extra_data,
        } for n in items]
        return Response({'results': results, 'count': total, 'page': page, 'total_pages': total_pages})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """未读通知数"""
        count = WorkNotification.objects.filter(recipient=request.user, is_read=False).count()
        return Response({'count': count})

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """标记单条通知为已读"""
        try:
            note = WorkNotification.objects.get(id=pk, recipient=request.user)
            note.is_read = True
            note.save()
            return Response({'message': 'ok'})
        except WorkNotification.DoesNotExist:
            return Response({'error': '通知不存在'}, status=404)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """标记所有通知为已读"""
        WorkNotification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
        return Response({'message': 'ok'})


def _parse_ocr_raw(raw):
    """解析前端回传的 OCR 原始返回（JSON 字符串或 dict），供 SubsidyApplication.ocr_raw_data 存储"""
    if raw is None:
        return None
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            return json.loads(raw)
        except Exception:
            return None
    return None


def _baidu_invoice_type(invoice_type, ocr_raw_data=None):
    """推导百度验真接口的发票种类代码（invoice_type）。

    优先用 OCR 原始返回的发票类型文本判断：
      电子发票/全电发票（专/普）→ elec_invoice_special / elec_invoice_normal（校验码可为空、金额填价税合计）；
      纸质专票/普票 → special_vat_invoice / normal_invoice；
    无 OCR 数据时按系统类型 special/ordinary 兜底为全电发票（当前以电子发票为主）。
    """
    org = ''
    if ocr_raw_data and isinstance(ocr_raw_data, dict):
        wr = ocr_raw_data.get('words_result') or {}
        if isinstance(wr, dict):
            org = str(wr.get('InvoiceTypeOrg') or wr.get('InvoiceType') or '')
    if org:
        if '专用' in org:
            return 'elec_invoice_special' if ('电子发票' in org or '全电' in org) else 'special_vat_invoice'
        if '普通' in org or '晋通' in org:
            return 'elec_invoice_normal' if ('电子发票' in org or '全电' in org) else 'normal_invoice'
    if invoice_type == 'special':
        return 'elec_invoice_special'
    if invoice_type == 'ordinary':
        return 'elec_invoice_normal'
    return ''


class SubsidyViewSet(viewsets.ViewSet):
    """员工消费普惠补贴"""
    permission_classes = [permissions.IsAuthenticated]
    RATE_MAP = {'special': 0.0100, 'ordinary': 0.0050}

    @staticmethod
    def _gen_application_no():
        import random
        for _ in range(10):
            no = 'BF' + timezone.now().strftime('%Y%m%d%H%M%S') + str(random.randint(1000, 9999))
            if not SubsidyApplication.objects.filter(application_no=no).exists():
                return no
        return 'BF' + uuid.uuid4().hex[:12].upper()

    @staticmethod
    def _get_parent_tenant(tenant):
        return tenant.parent if tenant else None

    @staticmethod
    def _applicant_primary_dept_id(user):
        from org.models import UserDepartment
        rec = UserDepartment.objects.filter(user=user, is_primary=True).select_related('department').first()
        return rec.department_id if rec else None

    def _get_subsidy_config(self, tenant, department_id=None):
        """按优先级获取普惠补贴配置：部门配置 > 子公司配置 > 企业默认 > 父级回溯"""
        if not tenant:
            return None
        if department_id:
            try:
                return SubsidyConfig.objects.get(tenant=tenant, department_id=department_id)
            except SubsidyConfig.DoesNotExist:
                pass
        if tenant.parent:
            try:
                return SubsidyConfig.objects.get(
                    tenant=tenant.parent, sub_tenant=tenant, department__isnull=True)
            except SubsidyConfig.DoesNotExist:
                pass
        try:
            return SubsidyConfig.objects.get(
                tenant=tenant, sub_tenant__isnull=True, department__isnull=True)
        except SubsidyConfig.DoesNotExist:
            pass
        if tenant.parent:
            parent = self._get_parent_tenant(tenant)
            if parent:
                return self._get_subsidy_config(parent)
        return None

    def _can_manage_subsidy_config(self, request):
        """是否可管理普惠补贴配置：超级管理员、配置的财务核验人员或财务支付人员"""
        user = request.user
        if not user or user.user_type == 'super_admin':
            return True
        tenant = getattr(request, 'tenant', None) or user.get_active_tenant()
        return self._is_verifier(user, tenant) or self._is_payment_staff(user, tenant)

    def _is_verifier(self, user, tenant):
        """判断用户是否为当前企业范围内配置的财务核验人员"""
        if not tenant or not user:
            return False
        configs = SubsidyConfig.objects.filter(verifiers=user)
        for cfg in configs:
            if cfg.department:
                if cfg.department.tenant_id == tenant.id:
                    return True
                continue
            if cfg.sub_tenant:
                if cfg.sub_tenant_id == tenant.id:
                    return True
                continue
            t = tenant
            while t:
                if t.id == cfg.tenant_id:
                    return True
                t = t.parent
        return False

    def _is_payment_staff(self, user, tenant):
        """判断用户是否为当前企业范围内配置的财务支付人员"""
        if not tenant or not user:
            return False
        configs = SubsidyConfig.objects.filter(payment_staff=user)
        for cfg in configs:
            if cfg.department:
                if cfg.department.tenant_id == tenant.id:
                    return True
                continue
            if cfg.sub_tenant:
                if cfg.sub_tenant_id == tenant.id:
                    return True
                continue
            t = tenant
            while t:
                if t.id == cfg.tenant_id:
                    return True
                t = t.parent
        return False

    @staticmethod
    def _user_primary_dept_name(user):
        if not user:
            return ''
        try:
            from org.models import UserDepartment
            rec = UserDepartment.objects.filter(user=user, is_primary=True).select_related('department').first()
            if rec and rec.department:
                return rec.department.name
        except Exception:
            pass
        return user.department.name if user.department else ''

    @staticmethod
    def _user_wallet_balance(user, tenant):
        try:
            w = SubsidyWallet.objects.filter(user=user, tenant=tenant).first()
            return float(w.balance) if w else 0.0
        except Exception:
            return 0.0

    @staticmethod
    def _get_or_create_wallet(user, tenant):
        from decimal import Decimal as _D
        wallet, _ = SubsidyWallet.objects.get_or_create(user=user, tenant=tenant or user.get_active_tenant())
        return wallet

    @staticmethod
    def _credit_wallet(user, tenant, amount, note=''):
        """核验通过补贴入账：balance + amount, total_in + amount（事务内调用）"""
        from decimal import Decimal as _D
        from django.db.models import F
        amount = _D(str(amount))
        wallet, _ = SubsidyWallet.objects.get_or_create(user=user, tenant=tenant or user.get_active_tenant())
        SubsidyWallet.objects.filter(pk=wallet.pk).update(
            balance=F('balance') + amount, total_in=F('total_in') + amount)
        return wallet

    @staticmethod
    def _debit_wallet(user, tenant, amount):
        """提现扣减：余额足够才扣（原子防超扣），余额不足返回 False"""
        from decimal import Decimal as _D
        from django.db.models import F
        amount = _D(str(amount))
        wallet, _ = SubsidyWallet.objects.get_or_create(user=user, tenant=tenant or user.get_active_tenant())
        updated = SubsidyWallet.objects.filter(pk=wallet.pk, balance__gte=amount).update(
            balance=F('balance') - amount, total_out=F('total_out') + amount)
        return updated > 0

    @staticmethod
    def _refund_wallet(user, tenant, amount):
        """提现驳回返还余额（total_out 保持累计提现口径不变）"""
        from decimal import Decimal as _D
        from django.db.models import F
        amount = _D(str(amount))
        wallet, _ = SubsidyWallet.objects.get_or_create(user=user, tenant=tenant or user.get_active_tenant())
        SubsidyWallet.objects.filter(pk=wallet.pk).update(balance=F('balance') + amount)
        return wallet

    def _notify_verifiers_pending(self, app):
        """新申领通知配置的核验人员（跳转财务核验页）；未配置核验人员时回退通知企业管理员"""
        dept_id = self._applicant_primary_dept_id(app.applicant)
        config = self._get_subsidy_config(app.tenant, dept_id)
        verify_url = f'/oa/subsidy-verify/?application_id={app.id}'
        if config:
            verifiers = list(config.verifiers.filter(is_active=True).exclude(id=app.applicant_id)[:20])
            if verifiers:
                for v in verifiers:
                    send_work_notification(
                        user_id=v.id,
                        title='新补贴申领待核验',
                        content=f'{app.applicant.real_name or app.applicant.username} 提交了普惠补贴申领（{app.application_no}），请及时核验',
                        notification_type='subsidy_apply', related_url=verify_url,
                        extra_data={'application_no': app.application_no, 'action': 'pending'},
                    )
                return
        from accounts.models import CustomUser
        qs = CustomUser.objects.filter(user_type='super_admin')
        if app.tenant:
            qs = qs.filter(tenant_memberships__tenant=app.tenant, tenant_memberships__is_active=True)
        for admin in qs.distinct()[:20]:
            send_work_notification(
                user_id=admin.id,
                title='新补贴申领待核验',
                content=f'{app.applicant.real_name or app.applicant.username} 提交了普惠补贴申领（{app.application_no}），请及时核验',
                notification_type='subsidy_apply', related_url=verify_url,
                extra_data={'application_no': app.application_no, 'action': 'pending'},
            )

    def _notify_payment_staff_pending(self, wd):
        """新提现申请通知配置的财务支付人员（跳转财务支付页）；未配置时回退通知企业超级管理员"""
        from accounts.models import CustomUser
        tenant = wd.tenant or wd.user.get_active_tenant()
        config = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(wd.user))
        pay_url = f'/oa/subsidy-pay/?withdrawal_id={wd.id}'
        notified = set()
        if config:
            staff = list(config.payment_staff.filter(is_active=True).exclude(id=wd.user_id)[:20])
            if staff:
                for s in staff:
                    if s.id in notified:
                        continue
                    notified.add(s.id)
                    send_work_notification(
                        user_id=s.id,
                        title='新提现申请待支付',
                        content=f'{wd.user.real_name or wd.user.username} 提交了提现申请 {wd.amount} 元，请及时支付',
                        notification_type='subsidy_withdraw', related_url=pay_url,
                        extra_data={'withdrawal_id': wd.id, 'action': 'withdraw'},
                    )
                return
        qs = CustomUser.objects.filter(user_type='super_admin')
        if tenant:
            qs = qs.filter(tenant_memberships__tenant=tenant, tenant_memberships__is_active=True)
        for admin in qs.distinct()[:20]:
            if admin.id in notified or admin.id == wd.user_id:
                continue
            notified.add(admin.id)
            send_work_notification(
                user_id=admin.id,
                title='新提现申请待支付',
                content=f'{wd.user.real_name or wd.user.username} 提交了提现申请 {wd.amount} 元，请及时支付',
                notification_type='subsidy_withdraw', related_url=pay_url,
                extra_data={'withdrawal_id': wd.id, 'action': 'withdraw'},
            )

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_invoice(self, request):
        """上传发票文件（图片/PDF）"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请选择文件'}, status=400)
        ext = os.path.splitext(file.name)[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'):
            return Response({'error': '仅支持图片或 PDF 文件'}, status=400)
        if file.size > 50 * 1024 * 1024:
            return Response({'error': '文件大小不能超过50MB'}, status=400)
        from django.core.files.storage import default_storage
        saved_path = default_storage.save(f'subsidy_invoices/{uuid.uuid4().hex}{ext}', file)
        return Response({'url': settings.MEDIA_URL + saved_path, 'name': file.name})

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_voucher(self, request):
        """上传支付凭证（付款截图），图片/PDF"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请选择文件'}, status=400)
        ext = os.path.splitext(file.name)[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'):
            return Response({'error': '仅支持图片或 PDF 文件'}, status=400)
        if file.size > 50 * 1024 * 1024:
            return Response({'error': '文件大小不能超过50MB'}, status=400)
        from django.core.files.storage import default_storage
        saved_path = default_storage.save(f'subsidy_payment_vouchers/{uuid.uuid4().hex}{ext}', file)
        return Response({'url': settings.MEDIA_URL + saved_path, 'name': file.name})

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_proof(self, request):
        """通用上传：支付截图 / 支付宝/微信收款码（图片/PDF），返回 url + name"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请选择文件'}, status=400)
        ext = os.path.splitext(file.name)[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'):
            return Response({'error': '仅支持图片或 PDF 文件'}, status=400)
        if file.size > 50 * 1024 * 1024:
            return Response({'error': '文件大小不能超过50MB'}, status=400)
        from django.core.files.storage import default_storage
        saved_path = default_storage.save(f'subsidy_proofs/{uuid.uuid4().hex}{ext}', file)
        return Response({'url': settings.MEDIA_URL + saved_path, 'name': file.name})

    @staticmethod
    def _user_has_payment_info(user):
        """判断用户是否已填写有效收款账号（收款人姓名 + 至少一种收款方式）"""
        if not user or not (user.payee_name or '').strip():
            return False
        return bool((user.bank_card or '').strip() or (user.alipay_account or '').strip()
                    or (user.wechat_account or '').strip())

    @action(detail=False, methods=['get', 'post'])
    def payment_info(self, request):
        """获取/更新当前用户收款账号信息"""
        u = request.user
        if request.method == 'GET':
            return Response({'encrypt': True, 'data': encrypt_data({
                'payee_name': u.payee_name or '',
                'bank_card': u.bank_card or '',
                'bank_name': u.bank_name or '',
                'bank_address': u.bank_address or '',
                'alipay_account': u.alipay_account or '',
                'wechat_account': u.wechat_account or '',
                'alipay_qr': u.alipay_qr or '',
                'wechat_qr': u.wechat_qr or '',
            })})
        u.payee_name = (request.data.get('payee_name') or '').strip()
        u.bank_card = (request.data.get('bank_card') or '').strip()
        u.bank_name = (request.data.get('bank_name') or '').strip()
        u.bank_address = (request.data.get('bank_address') or '').strip()
        u.alipay_account = (request.data.get('alipay_account') or '').strip()
        u.wechat_account = (request.data.get('wechat_account') or '').strip()
        u.alipay_qr = (request.data.get('alipay_qr') or '').strip()
        u.wechat_qr = (request.data.get('wechat_qr') or '').strip()
        u.save(update_fields=['payee_name', 'bank_card', 'bank_name', 'bank_address',
                              'alipay_account', 'wechat_account', 'alipay_qr', 'wechat_qr'])
        return Response({'encrypt': True, 'data': encrypt_data({
            'payee_name': u.payee_name, 'bank_card': u.bank_card,
            'bank_name': u.bank_name, 'bank_address': u.bank_address,
            'alipay_account': u.alipay_account, 'wechat_account': u.wechat_account,
            'alipay_qr': u.alipay_qr, 'wechat_qr': u.wechat_qr,
        })})

    @action(detail=False, methods=['post'])
    def ocr_invoice(self, request):
        """票据OCR识别（Celery 异步 + MD5去重缓存）：
        同一票据（文件MD5）+ 同一 ocr_version 已识别过则直接返回缓存结果，否则入队后台识别。
        ocr_version: baidu_vat(默认) / baidu_general / paddle"""
        import hashlib
        from django.core.cache import cache
        # 未指定识别版本时，使用配置的默认OCR识别版本
        ocr_version = (request.data.get('ocr_version') or '').strip()
        if not ocr_version:
            tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
            cfg = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(request.user))
            ocr_version = (cfg.default_ocr_version if cfg and cfg.default_ocr_version else 'paddle')
        ocr_version = ocr_version.strip() or 'paddle'
        if ocr_version not in ('baidu_vat', 'baidu_general', 'paddle'):
            return Response({'error': '不支持的OCR识别版本'}, status=400)
        # 税率阈值（用于按税率判定发票类型：识别税率 >= 阈值 → 增值税专用发票，否则普通发票）
        # 与发票识别缓存时间（秒）均取自补贴配置（默认阈值 6%、缓存 7 天）
        cfg = None
        try:
            tax_rate_threshold = float(request.data.get('tax_rate_threshold') or 0)
        except (ValueError, TypeError):
            tax_rate_threshold = 0
        if tax_rate_threshold <= 0:
            try:
                tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
                cfg = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(request.user))
                tax_rate_threshold = float(cfg.tax_rate_threshold) if cfg and cfg.tax_rate_threshold else 0.06
            except Exception as e:
                logger.warning(f'解析税率阈值失败: {e}')
                tax_rate_threshold = 0.06
        try:
            ocr_cache_ttl = int(cfg.ocr_cache_ttl) if cfg and cfg.ocr_cache_ttl else 604800
        except (ValueError, TypeError):
            ocr_cache_ttl = 604800
        if ocr_cache_ttl < 60:
            ocr_cache_ttl = 60
        url = (request.data.get('url') or request.data.get('file_url') or '').strip()
        file = request.FILES.get('file')
        image_path = None
        delete_after = False
        # 尽量解析为服务器本地文件路径，避免 Worker 重复下载
        image_path = self._resolve_media_path(url) if url else None
        image_data = None
        if image_path:
            try:
                with open(image_path, 'rb') as f:
                    image_data = f.read()
            except Exception as e:
                logger.warning(f'读取票据文件失败: {e}')
                image_data = None
        if image_data is None:
            if file:
                image_data = file.read()
            elif url:
                full = url if not url.startswith('/') else f"{request.scheme}://{request.get_host()}{url}"
                try:
                    resp = requests.get(full, timeout=30)
                    resp.raise_for_status()
                    image_data = resp.content
                except Exception as e:
                    return Response({'error': f'下载票据文件失败: {e}'}, status=400)
            if image_data is not None:
                import tempfile
                ext = os.path.splitext(url or '')[-1] or '.img'
                tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
                tmp.write(image_data)
                tmp.close()
                image_path = tmp.name
                delete_after = True
        if image_data is None or not image_path or not os.path.exists(image_path):
            return Response({'error': '缺少票据文件'}, status=400)
        # 缓存键：票据文件 MD5 + OCR 版本，防止同一票据重复识别
        md5 = hashlib.md5(image_data).hexdigest()
        cache_key = f'subsidy_ocr:{md5}:{ocr_version}'
        cached = cache.get(cache_key)
        if cached:
            logger.info(f'OCR缓存命中 md5={md5[:8]} version={ocr_version}')
            return Response({'encrypt': True, 'data': encrypt_data({'task_id': '', 'result': cached, 'cached': True})})
        try:
            from .tasks import subsidy_ocr_task
            logger.info(f'OCR任务入队 md5={md5[:8]} version={ocr_version} tax_rate_threshold={tax_rate_threshold} cache_ttl={ocr_cache_ttl} user={request.user}')
            task = subsidy_ocr_task.delay(image_path, ocr_version, cache_key, delete_after, tax_rate_threshold, ocr_cache_ttl)
        except Exception as e:
            logger.error(f'OCR任务入队失败: {e}')
            return Response({'error': f'识别任务提交失败: {e}'}, status=500)
        return Response({'encrypt': True, 'data': encrypt_data({'task_id': task.id})})

    @action(detail=False, methods=['get'])
    def ocr_status(self, request):
        """轮询异步OCR任务状态：PENDING / PROGRESS / SUCCESS(result) / FAILURE(error)"""
        task_id = request.query_params.get('task_id', '').strip()
        if not task_id:
            return Response({'error': '缺少 task_id 参数'}, status=400)
        try:
            from celery.result import AsyncResult
            res = AsyncResult(task_id)
        except Exception as e:
            return Response({'error': f'查询任务失败: {e}'}, status=500)
        state = res.state
        if state == 'SUCCESS':
            result = res.result or {}
            if isinstance(result, dict) and result.get('error'):
                return Response({'encrypt': True, 'data': encrypt_data({'state': 'FAILURE', 'error': result['error']})})
            return Response({'encrypt': True, 'data': encrypt_data({'state': 'SUCCESS', 'result': result})})
        if state == 'FAILURE':
            return Response({'encrypt': True, 'data': encrypt_data({'state': 'FAILURE', 'error': str(res.result or '识别失败')})})
        return Response({'encrypt': True, 'data': encrypt_data({'state': 'PENDING'})})

    @staticmethod
    def _resolve_media_path(url):
        """将票据 URL 解析为服务器本地文件路径

        兼容多种存储形式：带/不带前导斜杠（/media/... 与 media/...）、
        完整 http(s) URL、纯相对路径（subsidy_invoices/...），
        并做路径穿越防护（结果必须位于 MEDIA_ROOT 内）。
        """
        if not url:
            return None
        try:
            if url.startswith(('http://', 'https://')):
                from urllib.parse import urlparse
                url = urlparse(url).path
            rel = url.split('?', 1)[0].lstrip('/')
            media_prefix = (settings.MEDIA_URL or '').lstrip('/')
            if media_prefix and rel.startswith(media_prefix):
                rel = rel[len(media_prefix):]
            candidate = os.path.abspath(os.path.join(settings.MEDIA_ROOT, rel))
            root = os.path.abspath(settings.MEDIA_ROOT)
            if not candidate.startswith(root):
                return None
            if os.path.exists(candidate) and os.path.isfile(candidate):
                return candidate
        except Exception as e:
            logger.warning(f'解析媒体路径失败: {e}')
        return None

    def _save_pdf_image(self, invoice_file):
        """若票据为 PDF，将第一页转成 PNG 存入 MEDIA，返回图片 URL；非 PDF 或转换失败返回 ''"""
        if not invoice_file:
            return ''
        if not invoice_file.lower().endswith('.pdf'):
            return ''
        path = self._resolve_media_path(invoice_file)
        if not path:
            return ''
        try:
            with open(path, 'rb') as f:
                pdf_bytes = f.read()
            png = self._render_pdf_preview_png(pdf_bytes)
            if not png:
                return ''
            from django.core.files.storage import default_storage
            from django.core.files.base import ContentFile
            saved = default_storage.save(
                f'subsidy_invoice_images/{uuid.uuid4().hex}.png', ContentFile(png))
            return settings.MEDIA_URL + saved
        except Exception as e:
            logger.warning(f'PDF转图片保存失败: {e}')
            return ''

    @staticmethod
    def _render_pdf_preview_png(pdf_bytes):
        """将 PDF 第一页渲染为 PNG 字节（供详情预览）"""
        # 方式一：PyMuPDF
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype='pdf')
            page = doc.load_page(0)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            png = pix.tobytes('png')
            doc.close()
            return png
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f'PyMuPDF 渲染预览失败: {e}')
        # 方式二：pypdfium2 + PIL
        try:
            import pypdfium2 as pdfium
            from PIL import Image
            import io
            pdf = pdfium.PdfDocument(pdf_bytes)
            pil = pdf[0].render(scale=2).to_pil()
            buf = io.BytesIO()
            pil.save(buf, format='PNG')
            pdf.close()
            return buf.getvalue()
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f'pypdfium2 渲染预览失败: {e}')
        # 方式三：pdf2image（依赖 poppler）
        try:
            import io as _io
            from pdf2image import convert_from_bytes
            images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1, dpi=150)
            if images:
                buf = _io.BytesIO()
                images[0].save(buf, format='PNG')
                return buf.getvalue()
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f'pdf2image 渲染预览失败: {e}')
        return None

    @action(detail=False, methods=['get'])
    def invoice_preview(self, request):
        """票据预览图：图片直接返回原图；PDF 渲染第一页为 PNG"""
        from django.http import HttpResponse
        url = request.query_params.get('url', '').strip()
        if not url:
            return Response({'error': '缺少 url 参数'}, status=400)
        image_path = self._resolve_media_path(url)
        if not image_path:
            return Response({'error': '票据文件不存在'}, status=404)
        if image_path.lower().endswith('.pdf'):
            try:
                with open(image_path, 'rb') as f:
                    pdf_bytes = f.read()
            except Exception as e:
                return Response({'error': f'读取PDF失败: {e}'}, status=500)
            png = self._render_pdf_preview_png(pdf_bytes)
            if not png:
                return Response({'error': '服务器未安装 PDF 渲染库，无法预览该PDF'}, status=500)
            return HttpResponse(png, content_type='image/png')
        # 图片直接返回
        ext = os.path.splitext(image_path)[1].lower().lstrip('.')
        ctype = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
                 'gif': 'image/gif', 'webp': 'image/webp'}.get(ext, 'application/octet-stream')
        try:
            with open(image_path, 'rb') as f:
                return HttpResponse(f.read(), content_type=ctype)
        except Exception as e:
            return Response({'error': f'读取票据失败: {e}'}, status=500)

    @action(detail=False, methods=['post'])
    def qr_scan(self, request):
        """扫描票据二维码：解码二维码原文 + 解析发票字段，供核验人员核验参考"""
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if request.user.user_type != 'super_admin' and not self._is_verifier(request.user, tenant):
            return Response({'error': '仅超级管理员或财务核验人员可操作'}, status=403)
        url = (request.data.get('url') or '').strip()
        if not url:
            return Response({'error': '缺少票据url参数'}, status=400)
        image_path = self._resolve_media_path(url)
        if not image_path:
            return Response({'error': '票据文件不存在'}, status=404)
        try:
            with open(image_path, 'rb') as f:
                image_data = f.read()
        except Exception as e:
            return Response({'error': f'读取票据失败: {e}'}, status=500)
        # PDF 先渲染第一页为 PNG 再解码
        if image_path.lower().endswith('.pdf'):
            png = self._render_pdf_preview_png(image_data)
            if not png:
                return Response({'error': 'PDF渲染失败，无法扫描二维码'}, status=500)
            image_data = png
        from utils.qr_scan import scan_qr_strings, parse_qr_fields
        try:
            qr_strings = scan_qr_strings(image_data)
        except Exception as e:
            logger.warning(f'二维码扫描失败: {e}')
            qr_strings = []
        if not qr_strings:
            return Response({'encrypt': True, 'data': encrypt_data({'qr_strings': [], 'parsed': {}})})
        parsed = parse_qr_fields(qr_strings[0])
        return Response({'encrypt': True, 'data': encrypt_data({'qr_strings': qr_strings, 'parsed': parsed})})

    def create(self, request):
        """提交补贴申领（需先完善收款账号）"""
        from decimal import Decimal
        # 校验收款账号：未填写则提示先完善收款信息
        if not self._user_has_payment_info(request.user):
            return Response({'error': '您还未填写收款账号，请先完善收款信息后再提交申领',
                             'need_payment_info': True}, status=400)
        invoice_type = request.data.get('invoice_type')
        if invoice_type not in ('special', 'ordinary'):
            return Response({'error': '发票类型必须为 special(专用) 或 ordinary(普通)'}, status=400)
        try:
            amount = Decimal(str(request.data.get('invoice_amount') or 0))
        except Exception:
            return Response({'error': '开票金额不合法'}, status=400)
        if amount <= 0:
            return Response({'error': '开票金额必须大于0'}, status=400)
        invoice_file = (request.data.get('invoice_file') or '').strip()
        if not invoice_file:
            return Response({'error': '请先上传票据文件'}, status=400)
        invoice_number = (request.data.get('invoice_number') or '').strip()
        if not invoice_number:
            return Response({'error': '发票号码为必填项'}, status=400)
        if SubsidyApplication.objects.filter(invoice_number=invoice_number).exists():
            return Response({'error': f'{invoice_number} 该发票号码已存在'}, status=400)
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        # 依据配置优先级解析补贴比例与开启状态
        dept_id = self._applicant_primary_dept_id(request.user)
        config = self._get_subsidy_config(tenant, dept_id)
        if config is not None and not config.enabled:
            return Response({'error': '普惠补贴暂未开启，无法提交申领'}, status=400)
        if config is not None:
            rate = config.special_rate if invoice_type == 'special' else config.ordinary_rate
            rate = Decimal(str(rate))
        else:
            rate = Decimal(str(self.RATE_MAP[invoice_type]))
        subsidy_amount = (amount * rate).quantize(Decimal('0.01'))
        invoice_date = request.data.get('invoice_date') or None
        # PDF 票据转首页 PNG 存储，供前端渲染与二维码扫描
        invoice_image = self._save_pdf_image(invoice_file)
        app = SubsidyApplication.objects.create(
            applicant=request.user, tenant=tenant,
            application_no=self._gen_application_no(),
            invoice_number=invoice_number,
            invoice_type=invoice_type,
            invoice_code=(request.data.get('invoice_code') or '').strip(),
            invoice_amount=amount,
            invoice_date=invoice_date,
            tax_rate=(request.data.get('tax_rate') or '').strip(),
            invoice_issuer=(request.data.get('invoice_issuer') or '').strip(),
            invoice_file=invoice_file,
            invoice_original_name=(request.data.get('invoice_original_name') or '').strip(),
            invoice_image=invoice_image,
            buyer_name=(request.data.get('buyer_name') or '').strip(),
            buyer_tax_no=(request.data.get('buyer_tax_no') or '').strip(),
            seller_name=(request.data.get('seller_name') or '').strip(),
            seller_tax_no=(request.data.get('seller_tax_no') or '').strip(),
            drawer=(request.data.get('drawer') or '').strip(),
            payment_proof=(request.data.get('payment_proof') or '').strip(),
            payment_proof_name=(request.data.get('payment_proof_name') or '').strip(),
            ocr_raw_data=_parse_ocr_raw(request.data.get('ocr_raw_data')),
            subsidy_rate=rate, subsidy_amount=subsidy_amount, status='pending',
        )
        try:
            self._notify_verifiers_pending(app)
        except Exception as e:
            logger.warning(f'通知核验人员待核验失败: {e}')
        from .serializers import SubsidyApplicationSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyApplicationSerializer(app).data)}, status=201)

    @action(detail=True, methods=['post'])
    def re_submit(self, request, pk=None):
        """被驳回的申领修改后重新提交（保留原申领编号）"""
        from decimal import Decimal
        try:
            app = SubsidyApplication.objects.get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '申领不存在'}, status=404)
        if app.applicant_id != request.user.id:
            return Response({'error': '仅申领本人可操作'}, status=403)
        if app.status != 'rejected':
            return Response({'error': '仅被驳回的申领可重新提交'}, status=400)
        invoice_type = request.data.get('invoice_type')
        if invoice_type not in ('special', 'ordinary'):
            return Response({'error': '发票类型必须为 special(专用) 或 ordinary(普通)'}, status=400)
        try:
            amount = Decimal(str(request.data.get('invoice_amount') or 0))
        except Exception:
            return Response({'error': '开票金额不合法'}, status=400)
        if amount <= 0:
            return Response({'error': '开票金额必须大于0'}, status=400)
        invoice_file = (request.data.get('invoice_file') or '').strip()
        if not invoice_file:
            return Response({'error': '请先上传票据文件'}, status=400)
        invoice_number = (request.data.get('invoice_number') or '').strip()
        if not invoice_number:
            return Response({'error': '发票号码为必填项'}, status=400)
        # 唯一性校验：排除自身
        if SubsidyApplication.objects.filter(invoice_number=invoice_number).exclude(id=app.id).exists():
            return Response({'error': f'{invoice_number} 该发票号码已存在'}, status=400)
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        dept_id = self._applicant_primary_dept_id(request.user)
        config = self._get_subsidy_config(tenant, dept_id)
        if config is not None and not config.enabled:
            return Response({'error': '普惠补贴暂未开启，无法提交申领'}, status=400)
        if config is not None:
            rate = config.special_rate if invoice_type == 'special' else config.ordinary_rate
            rate = Decimal(str(rate))
        else:
            rate = Decimal(str(self.RATE_MAP[invoice_type]))
        subsidy_amount = (amount * rate).quantize(Decimal('0.01'))
        invoice_date = request.data.get('invoice_date') or None
        app.invoice_number = invoice_number
        app.invoice_type = invoice_type
        app.invoice_code = (request.data.get('invoice_code') or '').strip()
        app.invoice_amount = amount
        app.invoice_date = invoice_date
        app.tax_rate = (request.data.get('tax_rate') or '').strip()
        app.invoice_issuer = (request.data.get('invoice_issuer') or '').strip()
        app.invoice_file = invoice_file
        app.invoice_original_name = (request.data.get('invoice_original_name') or '').strip()
        app.invoice_image = self._save_pdf_image(invoice_file)
        app.buyer_name = (request.data.get('buyer_name') or '').strip()
        app.buyer_tax_no = (request.data.get('buyer_tax_no') or '').strip()
        app.seller_name = (request.data.get('seller_name') or '').strip()
        app.seller_tax_no = (request.data.get('seller_tax_no') or '').strip()
        app.drawer = (request.data.get('drawer') or '').strip()
        app.payment_proof = (request.data.get('payment_proof') or '').strip()
        app.payment_proof_name = (request.data.get('payment_proof_name') or '').strip()
        app.ocr_raw_data = _parse_ocr_raw(request.data.get('ocr_raw_data')) or app.ocr_raw_data
        app.subsidy_rate = rate
        app.subsidy_amount = subsidy_amount
        app.status = 'pending'
        app.reject_reason = ''
        app.verified_by = None
        app.verified_at = None
        app.save()
        try:
            self._notify_verifiers_pending(app)
        except Exception as e:
            logger.warning(f'重新提交后通知核验人员失败: {e}')
        from .serializers import SubsidyApplicationSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyApplicationSerializer(app).data)})

    @action(detail=True, methods=['delete'])
    def delete_my(self, request, pk=None):
        """删除被驳回的申领（仅本人可操作）"""
        try:
            app = SubsidyApplication.objects.get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '申领不存在'}, status=404)
        if app.applicant_id != request.user.id:
            return Response({'error': '仅申领本人可删除'}, status=403)
        if app.status != 'rejected':
            return Response({'error': '仅被驳回的申领可删除'}, status=400)
        app.delete()
        return Response({'message': 'ok'})

    def list(self, request):
        """我的申领列表"""
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        status = request.query_params.get('status', '').strip()
        search = request.query_params.get('search', '').strip()
        qs = SubsidyApplication.objects.filter(applicant=request.user)
        if status:
            qs = qs.filter(status=status)
        if search:
            qs = qs.filter(Q(application_no__icontains=search) | Q(invoice_issuer__icontains=search) | Q(invoice_code__icontains=search) | Q(invoice_number__icontains=search))
        qs = qs.order_by('-created_at')
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        items = qs[(page - 1) * page_size:page * page_size]
        from .serializers import SubsidyApplicationSerializer
        data = SubsidyApplicationSerializer(items, many=True).data
        return Response({'encrypt': True, 'data': encrypt_data({
            'results': data, 'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages,
        })})

    def retrieve(self, request, pk=None):
        """申领详情"""
        try:
            app = SubsidyApplication.objects.select_related('applicant', 'verified_by').get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '该条申领不存在或者已经删除'}, status=404)
        if app.applicant != request.user and request.user.user_type != 'super_admin' \
                and not self._is_verifier(request.user, app.tenant):
            return Response({'error': '无权查看'}, status=403)
        from .serializers import SubsidyApplicationSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyApplicationSerializer(app).data)})

    @action(detail=False, methods=['get'])
    def account(self, request):
        """我的补贴账户汇总 + 钱包 + 角色身份 + 配置（发票抬头/税率阈值/OCR版本/验真开关）"""
        qs = SubsidyApplication.objects.filter(applicant=request.user)
        total = qs.filter(status='approved').aggregate(s=Sum('subsidy_amount'))['s'] or 0
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        config = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(request.user))
        ih = {}
        if config:
            for f in ('invoice_header_name', 'invoice_header_tax_no', 'invoice_header_address',
                      'invoice_header_phone', 'invoice_header_bank', 'invoice_header_bank_account',
                      'invoice_header_bank_name', 'company_name', 'company_tax_no'):
                ih[f] = getattr(config, f, '') or ''
        wallet, _ = SubsidyWallet.objects.get_or_create(user=request.user, tenant=tenant)
        data = {
            'total_subsidy': float(total),
            'pending_count': qs.filter(status='pending').count(),
            'approved_count': qs.filter(status='approved').count(),
            'rejected_count': qs.filter(status='rejected').count(),
            'is_verifier': self._is_verifier(request.user, tenant),
            'is_payment_staff': self._is_payment_staff(request.user, tenant),
            'wallet_balance': float(wallet.balance),
            'wallet_total_in': float(wallet.total_in),
            'wallet_total_out': float(wallet.total_out),
            'min_withdraw_amount': float(config.min_withdraw_amount) if config and config.min_withdraw_amount is not None else 0.0,
            'default_ocr_version': (config.default_ocr_version if config and config.default_ocr_version else 'paddle'),
            'invoice_verify_enabled': bool(config and config.invoice_verify_enabled),
            'max_invoices': config.max_invoices if config and config.max_invoices else 10,
            'has_payment_info': self._user_has_payment_info(request.user),
            'show_invoice_header': bool(config and config.show_invoice_header),
            'invoice_header_show': (config.invoice_header_show or {}) if config else {},
            'tax_rate_threshold': float(config.tax_rate_threshold) if config and config.tax_rate_threshold else 0.06,
            'invoice_header': ih,
        }
        return Response({'encrypt': True, 'data': encrypt_data(data)})

    @action(detail=False, methods=['get'])
    def payments(self, request):
        """我的补贴发放记录"""
        qs = SubsidyPayment.objects.filter(user=request.user).select_related('application').order_by('-paid_at')
        from .serializers import SubsidyPaymentSerializer
        return Response({'encrypt': True, 'data': encrypt_data({'results': SubsidyPaymentSerializer(qs, many=True).data})})

    def _build_all_queryset(self, request):
        """构建核验列表过滤后的 queryset（供列表/导出共用），返回 (qs, tenant)
        仅超级管理员或配置的财务核验人员可访问核验列表"""
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if request.user.user_type != 'super_admin' and not self._is_verifier(request.user, tenant):
            return None, tenant
        tenant_ids = [tenant.id] if tenant else []
        if tenant:
            try:
                sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                tenant_ids.extend(sub_ids)
            except Exception:
                pass
        qs = SubsidyApplication.objects.select_related('applicant', 'tenant', 'verified_by').all()
        if tenant_ids:
            qs = qs.filter(Q(tenant_id__in=tenant_ids) | Q(tenant__isnull=True))
        record_ids = request.query_params.get('record_ids', '').strip()
        if record_ids:
            ids = [int(x) for x in record_ids.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(id__in=ids)
        status = request.query_params.get('status', '').strip()
        if status:
            qs = qs.filter(status=status)
        invoice_type = request.query_params.get('invoice_type', '').strip()
        if invoice_type:
            qs = qs.filter(invoice_type=invoice_type)
        search = request.query_params.get('search', '').strip()
        if search:
            from accounts.models import CustomUser
            applicant_ids = list(CustomUser.objects.filter(
                Q(real_name__icontains=search) | Q(username__icontains=search)
            ).values_list('id', flat=True)[:500])
            qs = qs.filter(
                Q(applicant_id__in=applicant_ids)
                | Q(application_no__icontains=search)
                | Q(invoice_issuer__icontains=search)
                | Q(invoice_code__icontains=search)
                | Q(invoice_number__icontains=search)
            )
        invoice_number = request.query_params.get('invoice_number', '').strip()
        if invoice_number:
            qs = qs.filter(invoice_number__icontains=invoice_number)
        date_from = request.query_params.get('date_from', '').strip()
        if date_from:
            try:
                qs = qs.filter(created_at__date__gte=timezone.datetime.strptime(date_from, '%Y-%m-%d').date())
            except (ValueError, TypeError):
                pass
        date_to = request.query_params.get('date_to', '').strip()
        if date_to:
            try:
                qs = qs.filter(created_at__date__lte=timezone.datetime.strptime(date_to, '%Y-%m-%d').date())
            except (ValueError, TypeError):
                pass
        min_amount = request.query_params.get('min_amount', '').strip()
        if min_amount:
            try:
                qs = qs.filter(invoice_amount__gte=float(min_amount))
            except (ValueError, TypeError):
                pass
        max_amount = request.query_params.get('max_amount', '').strip()
        if max_amount:
            try:
                qs = qs.filter(invoice_amount__lte=float(max_amount))
            except (ValueError, TypeError):
                pass
        return qs.order_by('-created_at'), tenant

    @action(detail=False, methods=['get'])
    def all(self, request):
        """核验用列表（当前企业 + 子企业），管理员与配置的核验人员可查看，支持多条件搜索"""
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        qs, tenant = self._build_all_queryset(request)
        if qs is None:
            return Response({'error': '仅管理员或财务核验人员可查看'}, status=403)
        total = qs.count()
        total_pages = max(1, (total + page_size - 1) // page_size)
        items = qs[(page - 1) * page_size:page * page_size]
        from .serializers import SubsidyApplicationSerializer
        data = SubsidyApplicationSerializer(items, many=True).data
        return Response({'encrypt': True, 'data': encrypt_data({
            'results': data, 'count': total, 'page': page, 'page_size': page_size, 'total_pages': total_pages,
        })})

    @action(detail=False, methods=['get'])
    def export(self, request):
        """核验列表导出 Excel（openpyxl），支持与列表相同的筛选条件"""
        from django.http import FileResponse
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response({'error': '服务器缺少 openpyxl 依赖，请联系管理员安装'}, status=500)
        qs, tenant = self._build_all_queryset(request)
        if qs is None:
            return Response({'error': '仅管理员或财务核验人员可导出'}, status=403)
        items = qs[:10000]
        status_map = {'pending': '待核验', 'approved': '已通过', 'rejected': '已驳回'}
        field_map = {
            'application_no': lambda r: r.application_no,
            'applicant_name': lambda r: (r.applicant.real_name or r.applicant.username) if r.applicant else '',
            'tenant_name': lambda r: (r.tenant.short_name or r.tenant.name) if r.tenant else '',
            'invoice_type': lambda r: r.get_invoice_type_display(),
            'invoice_code': lambda r: r.invoice_code,
            'invoice_number': lambda r: r.invoice_number or '',
            'invoice_amount': lambda r: float(r.invoice_amount),
            'subsidy_rate': lambda r: f'{float(r.subsidy_rate) * 100:g}%',
            'subsidy_amount': lambda r: float(r.subsidy_amount),
            'status': lambda r: status_map.get(r.status, r.status),
            'invoice_date': lambda r: r.invoice_date.strftime('%Y-%m-%d') if r.invoice_date else '',
            'invoice_issuer': lambda r: r.invoice_issuer,
            'drawer': lambda r: r.drawer or '',
            'buyer_name': lambda r: r.buyer_name or '',
            'buyer_tax_no': lambda r: r.buyer_tax_no or '',
            'seller_tax_no': lambda r: r.seller_tax_no or '',
            'verified_by': lambda r: (r.verified_by.real_name or r.verified_by.username) if r.verified_by else '',
            'verified_at': lambda r: timezone.localtime(r.verified_at).strftime('%Y-%m-%d %H:%M') if r.verified_at else '',
            'created_at': lambda r: timezone.localtime(r.created_at).strftime('%Y-%m-%d %H:%M') if r.created_at else '',
            'reject_reason': lambda r: r.reject_reason or '',
        }
        field_labels = {
            'application_no': '申领编号', 'applicant_name': '申请人', 'tenant_name': '所属企业',
            'invoice_type': '发票类型', 'invoice_code': '票据代码', 'invoice_number': '发票号码',
            'invoice_amount': '开票金额(元)', 'subsidy_rate': '补贴比例', 'subsidy_amount': '补贴金额(元)',
            'status': '状态', 'invoice_date': '开票日期', 'invoice_issuer': '开票主体',
            'drawer': '开票人', 'buyer_name': '购买方名称', 'buyer_tax_no': '购买方纳税人识别号',
            'seller_tax_no': '销售方纳税人识别号', 'verified_by': '核验人', 'verified_at': '核验时间',
            'created_at': '申请时间', 'reject_reason': '驳回原因',
        }
        fields_param = request.query_params.get('fields', '').strip()
        if fields_param:
            selected_fields = [f.strip() for f in fields_param.split(',') if f.strip() in field_map]
        else:
            selected_fields = list(field_map.keys())
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '普惠补贴申领'
            headers = [field_labels.get(f, f) for f in selected_fields]
            header_fill = PatternFill('solid', fgColor='409EFF')
            header_font = Font(color='FFFFFF', bold=True)
            thin = Side(style='thin', color='D9D9D9')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            for app in items:
                try:
                    row = [field_map[f](app) for f in selected_fields]
                except Exception:
                    continue
                ws.append(row)
            for col, width in enumerate([22, 18, 20, 18, 16, 16, 14, 12, 14, 12, 14, 30, 12, 22, 22, 22, 18, 20, 20, 30], start=1):
                ws.column_dimensions[get_column_letter(col)].width = width
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
        except Exception as e:
            logger.error(f'导出普惠补贴申领失败: {e}')
            return Response({'error': f'导出失败: {str(e)}'}, status=500)
        now = timezone.localtime(timezone.now())
        filename = '普惠补贴申领_%s.xlsx' % now.strftime('%Y%m%d_%H%M')
        return FileResponse(
            output,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """财务核验：approve 通过（金额转入用户钱包）/ reject 驳回填写原因（核验与支付分离）"""
        try:
            app = SubsidyApplication.objects.get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '申领不存在'}, status=404)
        if request.user.user_type != 'super_admin' and not self._is_verifier(request.user, app.tenant):
            return Response({'error': '仅超级管理员或财务核验人员可核验'}, status=403)
        if app.status != 'pending':
            return Response({'error': '仅待核验的申领可操作'}, status=400)
        action = request.data.get('action')
        reason = (request.data.get('reason') or '').strip()
        if action == 'approve':
            app.status = 'approved'
            app.verified_by = request.user
            app.verified_at = timezone.now()
            app.payment_voucher = ''
            app.payment_voucher_name = ''
            with transaction.atomic():
                app.save(update_fields=['status', 'verified_by', 'verified_at', 'payment_voucher', 'payment_voucher_name'])
                self._credit_wallet(app.applicant, app.tenant, app.subsidy_amount)
            send_work_notification(
                user_id=app.applicant_id,
                title='普惠补贴核验通过',
                content=f'您申领的{app.get_invoice_type_display()}普惠补贴 {app.subsidy_amount} 元已核验通过，已转入您的钱包，可前往普惠补贴页面提现。',
                notification_type='subsidy_result', related_url=f'/oa/subsidy/?application_id={app.id}',
                extra_data={'application_no': app.application_no, 'action': 'approved'},
            )
        elif action == 'reject':
            if not reason:
                return Response({'error': '驳回必须填写原因'}, status=400)
            app.status = 'rejected'
            app.reject_reason = reason
            app.verified_by = request.user
            app.verified_at = timezone.now()
            app.save(update_fields=['status', 'reject_reason', 'verified_by', 'verified_at'])
            send_work_notification(
                user_id=app.applicant_id,
                title='补贴申领被驳回',
                content=f'您的申领 {app.application_no} 未通过核验，原因：{reason}。可修改或更换票据后重新申领。',
                notification_type='subsidy_result', related_url=f'/oa/subsidy/?application_id={app.id}',
                extra_data={'application_no': app.application_no, 'action': 'rejected'},
            )
        else:
            return Response({'error': 'action 必须为 approve 或 reject'}, status=400)
        from .serializers import SubsidyApplicationSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyApplicationSerializer(app).data)})

    @action(detail=False, methods=['get'])
    def wallet(self, request):
        """我的钱包：余额 + 汇总 + 提现最小额度"""
        from decimal import Decimal
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        wallet, _ = SubsidyWallet.objects.get_or_create(user=request.user, tenant=tenant)
        config = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(request.user))
        min_wd = float(config.min_withdraw_amount) if config and config.min_withdraw_amount is not None else 0.0
        return Response({'encrypt': True, 'data': encrypt_data({
            'balance': float(wallet.balance),
            'total_in': float(wallet.total_in),
            'total_out': float(wallet.total_out),
            'min_withdraw_amount': min_wd,
            'updated_at': wallet.updated_at.isoformat() if wallet.updated_at else '',
        })})

    @action(detail=False, methods=['post'])
    def withdraw(self, request):
        """用户提现：金额>0、≤余额、≥最小额度；扣减余额、建提现申请、通知财务支付人员"""
        from decimal import Decimal
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        try:
            amount = Decimal(str(request.data.get('amount') or 0))
        except Exception:
            return Response({'error': '提现金额不合法'}, status=400)
        if amount <= 0:
            return Response({'error': '提现金额必须大于0'}, status=400)
        wallet, _ = SubsidyWallet.objects.get_or_create(user=request.user, tenant=tenant)
        if amount > wallet.balance:
            return Response({'error': '提现金额不能大于钱包余额'}, status=400)
        config = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(request.user))
        min_wd = config.min_withdraw_amount if config and config.min_withdraw_amount is not None else Decimal('0')
        if amount < min_wd:
            return Response({'error': f'提现金额不能低于最小额度 {min_wd} 元'}, status=400)
        with transaction.atomic():
            ok = self._debit_wallet(request.user, tenant, amount)
            if not ok:
                return Response({'error': '余额不足，无法提现'}, status=400)
            wd = SubsidyWithdrawal.objects.create(
                user=request.user, tenant=tenant, amount=amount,
                status='pending', note=(request.data.get('note') or '').strip())
        self._notify_payment_staff_pending(wd)
        from .serializers import SubsidyWithdrawalSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyWithdrawalSerializer(wd).data)}, status=201)

    @action(detail=False, methods=['get'])
    def withdrawals(self, request):
        """我的提现记录"""
        qs = SubsidyWithdrawal.objects.filter(user=request.user).order_by('-requested_at')
        from .serializers import SubsidyWithdrawalSerializer
        return Response({'encrypt': True, 'data': encrypt_data({'results': SubsidyWithdrawalSerializer(qs, many=True).data})})

    @action(detail=False, methods=['get'])
    def withdrawals_all(self, request):
        """财务支付列表（super_admin 或财务支付人员）"""
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if request.user.user_type != 'super_admin' and not self._is_payment_staff(request.user, tenant):
            return Response({'error': '仅超级管理员或财务支付人员可查看'}, status=403)
        qs = SubsidyWithdrawal.objects.select_related('user', 'tenant', 'paid_by').all()
        tenant_ids = [tenant.id] if tenant else []
        if tenant:
            try:
                sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                tenant_ids.extend(sub_ids)
            except Exception:
                pass
        qs = qs.filter(Q(tenant_id__in=tenant_ids) | Q(tenant__isnull=True))
        status = request.query_params.get('status', '').strip()
        if status:
            qs = qs.filter(status=status)
        user_id = request.query_params.get('user_id', '').strip()
        if user_id and user_id.isdigit():
            qs = qs.filter(user_id=int(user_id))
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(user__username__icontains=search) | Q(user__real_name__icontains=search))
        qs = qs.order_by('-requested_at')
        from .serializers import SubsidyWithdrawalSerializer
        return Response({'encrypt': True, 'data': encrypt_data({'results': SubsidyWithdrawalSerializer(qs, many=True).data})})

    @action(detail=False, methods=['get'])
    def withdrawals_export(self, request):
        """提现申请导出 Excel（super_admin 或财务支付人员），支持 record_ids + 字段选择"""
        import io
        from django.http import FileResponse
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if request.user.user_type != 'super_admin' and not self._is_payment_staff(request.user, tenant):
            return Response({'error': '仅超级管理员或财务支付人员可导出'}, status=403)
        qs = SubsidyWithdrawal.objects.select_related('user', 'tenant', 'paid_by').all()
        tenant_ids = [tenant.id] if tenant else []
        if tenant:
            try:
                sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
                tenant_ids.extend(sub_ids)
            except Exception:
                pass
        qs = qs.filter(Q(tenant_id__in=tenant_ids) | Q(tenant__isnull=True))
        record_ids = request.query_params.get('record_ids', '').strip()
        if record_ids:
            ids = [int(x) for x in record_ids.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(id__in=ids)
        status = request.query_params.get('status', '').strip()
        if status:
            qs = qs.filter(status=status)
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(user__username__icontains=search) | Q(user__real_name__icontains=search))
        qs = qs.order_by('-requested_at')[:10000]
        field_map = {
            'user_name': lambda w: w.user.real_name or w.user.username,
            'user_department': lambda w: self._user_primary_dept_name(w.user),
            'amount': lambda w: float(w.amount),
            'remaining_balance': lambda w: self._user_wallet_balance(w.user, w.tenant),
            'status': lambda w: w.get_status_display(),
            'requested_at': lambda w: w.requested_at.strftime('%Y-%m-%d %H:%M') if w.requested_at else '',
            'paid_at': lambda w: w.paid_at.strftime('%Y-%m-%d %H:%M') if w.paid_at else '',
            'paid_by': lambda w: (w.paid_by.real_name or w.paid_by.username) if w.paid_by else '',
            'note': lambda w: w.note or '',
            'reject_reason': lambda w: w.reject_reason or '',
        }
        field_labels = {
            'user_name': '提现人', 'user_department': '部门', 'amount': '提现金额(元)',
            'remaining_balance': '剩余金额(元)', 'status': '状态', 'requested_at': '申请时间',
            'paid_at': '支付时间', 'paid_by': '支付人员', 'note': '备注', 'reject_reason': '驳回原因',
        }
        fields_param = request.query_params.get('fields', '').strip()
        if fields_param:
            selected_fields = [f.strip() for f in fields_param.split(',') if f.strip() in field_map]
        else:
            selected_fields = list(field_map.keys())
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({'error': '服务器缺少 openpyxl 依赖'}, status=500)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '提现申请'
        ws.append([field_labels.get(f, f) for f in selected_fields])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor='E8F4FD')
            c.alignment = Alignment(horizontal='center')
        for w in qs:
            try:
                ws.append([field_map[f](w) for f in selected_fields])
            except Exception:
                continue
        for col in ws.columns:
            width = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 4, 8), 30)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from urllib.parse import quote
        now_dt = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')
        filename = f'提现申请_{now_dt}.xlsx'
        resp = FileResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return resp

    @action(detail=True, methods=['get'])
    def withdraw_detail(self, request, pk=None):
        """提现申请详情（申请人本人 / 财务支付人员 / 超级管理员）"""
        try:
            wd = SubsidyWithdrawal.objects.select_related('user', 'tenant', 'paid_by').get(id=pk)
        except SubsidyWithdrawal.DoesNotExist:
            return Response({'error': '提现申请不存在'}, status=404)
        if wd.user_id != request.user.id and request.user.user_type != 'super_admin' \
                and not self._is_payment_staff(request.user, wd.tenant):
            return Response({'error': '仅超级管理员或财务支付人员可查看'}, status=403)
        from .serializers import SubsidyWithdrawalSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyWithdrawalSerializer(wd).data)})

    @action(detail=True, methods=['post'])
    def withdraw_pay(self, request, pk=None):
        """财务支付：上传支付凭证，确认支付提现（生成发放记录并通知申请人）"""
        try:
            wd = SubsidyWithdrawal.objects.get(id=pk)
        except SubsidyWithdrawal.DoesNotExist:
            return Response({'error': '提现申请不存在'}, status=404)
        if request.user.user_type != 'super_admin' and not self._is_payment_staff(request.user, wd.tenant):
            return Response({'error': '仅超级管理员或财务支付人员可操作'}, status=403)
        if wd.status != 'pending':
            return Response({'error': '仅待支付的提现可操作'}, status=400)
        voucher = (request.data.get('payment_voucher') or '').strip()
        if not voucher:
            return Response({'error': '请先上传支付凭证（付款截图）'}, status=400)
        with transaction.atomic():
            wd.status = 'paid'
            wd.payment_voucher = voucher
            wd.payment_voucher_name = (request.data.get('payment_voucher_name') or '').strip()
            wd.paid_by = request.user
            wd.paid_at = timezone.now()
            wd.save()
            SubsidyPayment.objects.create(
                user=wd.user, tenant=wd.tenant, amount=wd.amount,
                note=f'提现申请 #{wd.id} 支付', withdrawal=wd)
        send_work_notification(
            user_id=wd.user_id,
            title='提现已支付',
            content=f'您申请的提现 {wd.amount} 元已由财务支付人员支付，请注意查收。',
            notification_type='subsidy_withdraw_result', related_url=f'/oa/subsidy/?withdrawal_id={wd.id}',
            extra_data={'withdrawal_id': wd.id, 'action': 'withdraw_paid'},
        )
        from .serializers import SubsidyWithdrawalSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyWithdrawalSerializer(wd).data)})

    @action(detail=True, methods=['post'])
    def withdraw_reject(self, request, pk=None):
        """财务支付：驳回提现（填写原因，金额返还钱包）"""
        try:
            wd = SubsidyWithdrawal.objects.get(id=pk)
        except SubsidyWithdrawal.DoesNotExist:
            return Response({'error': '提现申请不存在'}, status=404)
        if request.user.user_type != 'super_admin' and not self._is_payment_staff(request.user, wd.tenant):
            return Response({'error': '仅超级管理员或财务支付人员可操作'}, status=403)
        if wd.status != 'pending':
            return Response({'error': '仅待支付的提现可操作'}, status=400)
        reason = (request.data.get('reason') or '').strip()
        if not reason:
            return Response({'error': '驳回必须填写原因'}, status=400)
        with transaction.atomic():
            wd.status = 'rejected'
            wd.reject_reason = reason
            wd.paid_by = request.user
            wd.save()
            self._refund_wallet(wd.user, wd.tenant, wd.amount)
        send_work_notification(
            user_id=wd.user_id,
            title='提现被驳回',
            content=f'您申请的提现 {wd.amount} 元未通过，原因：{reason}，金额已退回钱包。',
            notification_type='subsidy_withdraw_result', related_url=f'/oa/subsidy/?withdrawal_id={wd.id}',
            extra_data={'withdrawal_id': wd.id, 'action': 'withdraw_rejected'},
        )
        from .serializers import SubsidyWithdrawalSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyWithdrawalSerializer(wd).data)})

    @action(detail=True, methods=['post'])
    def verify_invoice(self, request, pk=None):
        """发票验真（百度增值税发票验真，参数化接口 vat_invoice_verification）。

        凭发票代码/号码/开票日期/种类 + 校验码/金额与税务系统交叉核验；
        按发票要素（号码+代码+日期）去重，防止重复查验触发次数限制。
        """
        import hashlib
        try:
            app = SubsidyApplication.objects.get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '申领不存在'}, status=404)
        if request.user.user_type != 'super_admin' and not self._is_verifier(request.user, app.tenant):
            return Response({'error': '仅超级管理员或财务核验人员可验真'}, status=403)
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        config = self._get_subsidy_config(tenant, self._applicant_primary_dept_id(app.applicant))
        if config and not config.invoice_verify_enabled:
            return Response({'error': '未开启发票验真功能'}, status=400)
        if not app.invoice_number:
            return Response({'error': '缺少发票号码，无法验真'}, status=400)
        # 按发票要素去重（号码+代码+日期），同一发票只查验一次
        identity = f'{app.invoice_number}|{app.invoice_code}|{app.invoice_date}'
        md5 = hashlib.md5(identity.encode('utf-8')).hexdigest()
        rec = SubsidyInvoiceVerifyRecord.objects.filter(invoice_md5=md5).first()
        if rec:
            logger.info(f'[SubsidyInvoiceVerifyRecord] 重复查验 {md5}')
            return Response({'encrypt': True, 'data': encrypt_data({
                'result': rec.result, 'result_display': rec.get_result_display(),
                'message': rec.message, 'cached': True,
            })})
        # 从 OCR 原始返回推导百度发票种类/校验码/专票不含税金额
        raw = app.ocr_raw_data or {}
        wr_raw = raw.get('words_result') if isinstance(raw, dict) else {}
        if not isinstance(wr_raw, dict):
            wr_raw = {}
        baidu_type = _baidu_invoice_type(app.invoice_type, raw)
        check_code = str(wr_raw.get('CheckCode') or '').strip()
        total_amount = ''
        if baidu_type in ('elec_invoice_special', 'elec_invoice_normal'):
            # 全电发票（专/普）：金额填价税合计
            total_amount = str(app.invoice_amount)
        elif app.invoice_type == 'special':
            # 纸质/电子专票：填不含税金额（OCR 的 TotalAmount）
            total_amount = str(wr_raw.get('TotalAmount') or '').strip()
        # 普通发票（纸质/电子普票）total_amount 可为空
        from utils.baidu_ocr import verify_vat_invoice
        try:
            res = verify_vat_invoice(
                invoice_code=app.invoice_code or '',
                invoice_num=app.invoice_number or '',
                invoice_date=str(app.invoice_date) if app.invoice_date else '',
                invoice_type=baidu_type,
                check_code=check_code,
                total_amount=total_amount,
            )
        except Exception as e:
            return Response({'error': f'发票验真服务暂不可用：{e}'}, status=500)
        rec = SubsidyInvoiceVerifyRecord.objects.create(
            application=app, invoice_md5=md5, result=res['result'],
            message=res['message'], verify_data=res['data'], verified_by=request.user)
        return Response({'encrypt': True, 'data': encrypt_data({
            'result': rec.result, 'result_display': rec.get_result_display(),
            'message': rec.message, 'cached': False,
        })})

    @action(detail=True, methods=['get'])
    def invoice_verify_status(self, request, pk=None):
        """获取申领发票验真状态（供详情模态框回显）"""
        try:
            app = SubsidyApplication.objects.get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '申领不存在'}, status=404)
        rec = app.invoice_verify_records.first()
        if not rec:
            return Response({'encrypt': True, 'data': encrypt_data({'verified': False})})
        return Response({'encrypt': True, 'data': encrypt_data({
            'verified': True, 'result': rec.result,
            'result_display': rec.get_result_display(),
            'message': rec.message,
            'verified_at': rec.verified_at.isoformat() if rec.verified_at else '',
        })})

    @action(detail=True, methods=['post'])
    def update_invoice_type(self, request, pk=None):
        """核验人员修改发票类型，自动重算补贴比例与补贴金额"""
        from decimal import Decimal
        try:
            app = SubsidyApplication.objects.get(id=pk)
        except SubsidyApplication.DoesNotExist:
            return Response({'error': '申领不存在'}, status=404)
        if request.user.user_type != 'super_admin' and not self._is_verifier(request.user, app.tenant):
            return Response({'error': '仅超级管理员或财务核验人员可操作'}, status=403)
        if app.status != 'pending':
            return Response({'error': '仅待核验的申领可修改发票类型'}, status=400)
        invoice_type = (request.data.get('invoice_type') or '').strip()
        if invoice_type not in ('special', 'ordinary'):
            return Response({'error': '发票类型必须为 special(专用) 或 ordinary(普通)'}, status=400)
        config = self._get_subsidy_config(app.tenant, self._applicant_primary_dept_id(app.applicant))
        if config is not None:
            rate = config.special_rate if invoice_type == 'special' else config.ordinary_rate
            rate = Decimal(str(rate))
        else:
            rate = Decimal(str(self.RATE_MAP[invoice_type]))
        app.invoice_type = invoice_type
        app.subsidy_rate = rate
        app.subsidy_amount = (app.invoice_amount * rate).quantize(Decimal('0.01'))
        app.save(update_fields=['invoice_type', 'subsidy_rate', 'subsidy_amount'])
        from .serializers import SubsidyApplicationSerializer
        return Response({'encrypt': True, 'data': encrypt_data(SubsidyApplicationSerializer(app).data)})

    @action(detail=False, methods=['get'])
    def configs(self, request):
        """获取普惠补贴配置列表 + 子公司列表（仅超级管理员或财务核验人员可见）"""
        if not self._can_manage_subsidy_config(request):
            return Response({'results': [], 'sub_tenants': []})
        from .serializers import SubsidyConfigSerializer
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'results': [], 'sub_tenants': []})
        query = Q(tenant=tenant)
        try:
            sub_ids = list(tenant.sub_tenants.filter(is_active=True).values_list('id', flat=True))
            if sub_ids:
                query |= Q(tenant_id__in=sub_ids)
        except Exception:
            pass
        configs = SubsidyConfig.objects.filter(query).select_related('department', 'sub_tenant').prefetch_related('verifiers')
        data = SubsidyConfigSerializer(configs, many=True).data
        extra = {'sub_tenants': list(tenant.sub_tenants.filter(is_active=True).values(
            'id', 'name', 'short_name', 'tenant_type'))} if hasattr(tenant, 'sub_tenants') else {'sub_tenants': []}
        return Response({'results': data, **extra})

    @action(detail=False, methods=['post'])
    def save_config(self, request):
        """保存普惠补贴配置（集团/子公司/部门三级）"""
        if not self._can_manage_subsidy_config(request):
            return Response({'error': '仅超级管理员、财务核验或财务支付人员可操作'}, status=403)
        from accounts.models import Department, Tenant
        from decimal import Decimal
        tenant = getattr(request, 'tenant', None) or request.user.get_active_tenant()
        if not tenant:
            return Response({'error': '未找到所属企业'}, status=400)
        sub_tenant_id = request.data.get('sub_tenant_id')
        sub_tenant_obj = None
        if sub_tenant_id:
            try:
                sub_tenant_obj = Tenant.objects.get(id=int(sub_tenant_id))
                if sub_tenant_obj.parent_id != tenant.id:
                    return Response({'error': '指定的子公司不属于当前企业集团'}, status=400)
            except (ValueError, Tenant.DoesNotExist):
                return Response({'error': '子公司不存在'}, status=400)
        department_id = request.data.get('department_id')
        dept = None
        if department_id:
            try:
                dept = Department.objects.get(id=int(department_id), tenant=tenant)
            except (ValueError, Department.DoesNotExist):
                return Response({'error': '部门不存在'}, status=400)
        try:
            special_rate = Decimal(str(request.data.get('special_rate') or '0.0100'))
            ordinary_rate = Decimal(str(request.data.get('ordinary_rate') or '0.0050'))
        except Exception:
            return Response({'error': '补贴比例不合法'}, status=400)
        if special_rate < 0 or special_rate > 1 or ordinary_rate < 0 or ordinary_rate > 1:
            return Response({'error': '补贴比例应在 0 - 1 之间'}, status=400)
        enabled = str(request.data.get('enabled', 'true')).lower() in ('1', 'true', 'yes', 'on')
        try:
            max_invoices = int(request.data.get('max_invoices') or 10)
            if max_invoices < 1 or max_invoices > 50:
                return Response({'error': '一次上传最大票据数量应在 1 - 50 之间'}, status=400)
        except (ValueError, TypeError):
            max_invoices = 10
        # 税率阈值（小数，如 0.06 表示 6%）
        try:
            tax_rate_threshold = Decimal(str(request.data.get('tax_rate_threshold') or '0.0600'))
            if tax_rate_threshold < 0 or tax_rate_threshold > 1:
                return Response({'error': '税率阈值应在 0 - 1 之间'}, status=400)
        except Exception:
            tax_rate_threshold = Decimal('0.0600')
        verifier_ids = request.data.get('verifier_ids') or []
        if isinstance(verifier_ids, str):
            verifier_ids = [v for v in verifier_ids.split(',') if v.strip()]
        verifier_ids = [int(x) for x in verifier_ids if str(x).strip().isdigit()]
        # 财务支付人员（多选）
        payment_staff_ids = request.data.get('payment_staff_ids') or []
        if isinstance(payment_staff_ids, str):
            payment_staff_ids = [v for v in payment_staff_ids.split(',') if v.strip()]
        payment_staff_ids = [int(x) for x in payment_staff_ids if str(x).strip().isdigit()]
        # 提现最小额度
        try:
            min_withdraw_amount = Decimal(str(request.data.get('min_withdraw_amount') or '0'))
            if min_withdraw_amount < 0:
                return Response({'error': '提现最小额度不能为负'}, status=400)
        except Exception:
            min_withdraw_amount = Decimal('0')
        # 默认OCR识别版本
        default_ocr_version = (request.data.get('default_ocr_version') or 'paddle').strip()
        if default_ocr_version not in ('baidu_vat', 'baidu_general', 'paddle'):
            default_ocr_version = 'paddle'
        # 发票识别缓存时间（秒），默认 7 天
        try:
            ocr_cache_ttl = int(request.data.get('ocr_cache_ttl') or 604800)
        except (ValueError, TypeError):
            ocr_cache_ttl = 604800
        if ocr_cache_ttl < 60:
            ocr_cache_ttl = 60
        # 发票验真开关
        invoice_verify_enabled = str(request.data.get('invoice_verify_enabled', 'false')).lower() in ('1', 'true', 'yes', 'on')
        # 发票抬头字段显示开关（JSON：key→bool）
        try:
            ih_show_raw = request.data.get('invoice_header_show') or {}
            invoice_header_show = {}
            if isinstance(ih_show_raw, dict):
                for k, v in ih_show_raw.items():
                    invoice_header_show[str(k)] = str(v).lower() in ('1', 'true', 'yes', 'on')
        except Exception:
            invoice_header_show = {}
        try:
            config, created = SubsidyConfig.objects.update_or_create(
                tenant=tenant,
                sub_tenant=sub_tenant_obj,
                department=dept,
                defaults={
                    'enabled': enabled,
                    'special_rate': special_rate,
                    'ordinary_rate': ordinary_rate,
                    'max_invoices': max_invoices,
                    'show_invoice_header': str(request.data.get('show_invoice_header', 'false')).lower() in ('1', 'true', 'yes', 'on'),
                    'tax_rate_threshold': tax_rate_threshold,
                    'min_withdraw_amount': min_withdraw_amount,
                    'default_ocr_version': default_ocr_version,
                    'ocr_cache_ttl': ocr_cache_ttl,
                    'invoice_verify_enabled': invoice_verify_enabled,
                    'invoice_header_name': (request.data.get('invoice_header_name') or '').strip(),
                    'invoice_header_tax_no': (request.data.get('invoice_header_tax_no') or '').strip(),
                    'invoice_header_address': (request.data.get('invoice_header_address') or '').strip(),
                    'invoice_header_phone': (request.data.get('invoice_header_phone') or '').strip(),
                    'invoice_header_bank': (request.data.get('invoice_header_bank') or '').strip(),
                    'invoice_header_bank_account': (request.data.get('invoice_header_bank_account') or '').strip(),
                    'invoice_header_bank_name': (request.data.get('invoice_header_bank_name') or '').strip(),
                    'company_name': (request.data.get('company_name') or '').strip(),
                    'company_tax_no': (request.data.get('company_tax_no') or '').strip(),
                    'invoice_header_show': invoice_header_show,
                },
            )
            if verifier_ids:
                from accounts.models import CustomUser
                config.verifiers.set(CustomUser.objects.filter(id__in=verifier_ids))
            else:
                config.verifiers.clear()
            if payment_staff_ids:
                from accounts.models import CustomUser
                config.payment_staff.set(CustomUser.objects.filter(id__in=payment_staff_ids))
            else:
                config.payment_staff.clear()
            from .serializers import SubsidyConfigSerializer
            data = SubsidyConfigSerializer(config).data
            return Response({'encrypt': True, 'data': encrypt_data(data)}, status=201 if created else 200)
        except Exception as e:
            logger.error(f'保存普惠补贴配置失败: {e}')
            return Response({'error': f'保存配置失败: {str(e)}'}, status=400)

    @action(detail=True, methods=['delete'])
    def delete_config(self, request, pk=None):
        """删除普惠补贴配置"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可操作'}, status=403)
        try:
            config = SubsidyConfig.objects.get(id=pk)
            config.delete()
            return Response({'message': 'ok'})
        except SubsidyConfig.DoesNotExist:
            return Response({'error': '配置不存在'}, status=404)

    @action(detail=False, methods=['post'])
    def clear_ocr_cache(self, request):
        """清除发票识别缓存（修改识别版本/税率阈值后立即生效，无需等待缓存过期）"""
        if request.user.user_type != 'super_admin':
            return Response({'error': '仅超级管理员可操作'}, status=403)
        try:
            from django.core.cache import cache
            deleted = 0
            try:
                deleted = cache.delete_pattern('subsidy_ocr:*')
            except Exception as e:
                logger.warning(f'delete_pattern 不支持或失败: {e}')
            return Response({'encrypt': True, 'data': encrypt_data({
                'success': True, 'message': '发票识别缓存已清除，修改立即生效', 'deleted': deleted or 0})})
        except Exception as e:
            logger.error(f'清除发票识别缓存失败: {e}')
            return Response({'error': f'清除发票识别缓存失败: {str(e)}'}, status=500)
